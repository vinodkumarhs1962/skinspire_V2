# app/services/supplier_service.py 

from datetime import datetime, timezone, date, timedelta
import uuid
from typing import Dict, List, Optional, Tuple, Union
from decimal import Decimal

from sqlalchemy import and_, or_, func, desc
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError, SQLAlchemyError

from app.models.master import Supplier, Medicine, ChartOfAccounts, Hospital, Branch
from app.models.transaction import (
    PurchaseOrderHeader, PurchaseOrderLine,
    SupplierInvoice, SupplierInvoiceLine, SupplierPayment, PaymentDocumentAccessLog, PaymentDocument, GLTransaction, APSubledger, GLEntry, Inventory, PaymentDocument
)
from app.services.database_service import get_db_session, get_entity_dict
from app.services.gl_service import create_supplier_invoice_gl_entries, create_supplier_payment_gl_entries
from app.services.inventory_service import record_stock_from_supplier_invoice
from app.services.posting_config_service import get_posting_config, get_default_gl_account

# NEW: Import enhanced branch service functions
from app.services.branch_service import (
    get_default_branch_id,
    get_user_branch_id,
    validate_branch_access,
    get_user_accessible_branches,
    get_branch_for_supplier_operation,
    validate_entity_branch_access
)

from app.services.permission_service import (
    has_branch_permission,  
    has_cross_branch_permission,
    get_user_branch_context,
    get_user_accessible_branches, 
    apply_branch_filter_to_query
)

from app.utils.credit_note_utils import (
    is_credit_note_enabled, 
    generate_credit_note_number,
    validate_credit_amount,
    get_credit_note_description
)

from app.config import PAYMENT_CONFIG

from app.utils.unicode_logging import get_unicode_safe_logger
logger = get_unicode_safe_logger(__name__)


# ===================================================================
# Supplier Management - COMPLETE WITH BRANCH SUPPORT
# ===================================================================

def create_supplier(
    hospital_id: uuid.UUID,
    supplier_data: Dict,
    current_user_id: Optional[str] = None,
    branch_id: Optional[uuid.UUID] = None,  # NEW: Optional parameter
    session: Optional[Session] = None
) -> Dict:
    """
    Create a new supplier record - UPDATED to use branch_service
    BACKWARD COMPATIBLE: branch_id is optional and auto-determined
    """
    # NEW: Use branch service to determine branch
    determined_branch_id = get_branch_for_supplier_operation(
        current_user_id=current_user_id,
        hospital_id=hospital_id,
        specified_branch_id=branch_id
    )
    
    if not determined_branch_id:
        raise ValueError("Could not determine branch for supplier creation")
    
    # Add branch_id to supplier_data
    supplier_data['branch_id'] = determined_branch_id

    logger.info(f"Creating supplier for hospital {hospital_id}, branch {determined_branch_id}")
    logger.debug(f"Supplier data: {supplier_data}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _create_supplier(session, hospital_id, supplier_data, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _create_supplier(new_session, hospital_id, supplier_data, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing supplier creation: {result.get('supplier_id')}")
        new_session.commit()
        
        logger.info(f"Successfully created supplier: {result.get('supplier_name')}")
        return result

def _create_supplier(
    session: Session,
    hospital_id: uuid.UUID,
    supplier_data: Dict,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to create a new supplier record within a session - UPDATED with branch validation
    """
    try:
        branch_id = supplier_data.get('branch_id')
        if not branch_id:
            raise ValueError("Branch ID is required for supplier creation")
            
        # NEW: Validate branch belongs to hospital
        branch = session.query(Branch).filter_by(
            branch_id=branch_id,
            hospital_id=hospital_id,
            is_active=True
        ).first()
        
        if not branch:
            raise ValueError(f"Invalid branch ID {branch_id} for hospital {hospital_id}")
        
        logger.debug(f"Validating supplier name: {supplier_data.get('supplier_name')}")
        
        # Check for duplicate supplier name
        existing_supplier = session.query(Supplier).filter(
            Supplier.hospital_id == hospital_id,
            Supplier.supplier_name == supplier_data['supplier_name']
        ).first()
        
        if existing_supplier:
            error_msg = f"Supplier with name '{supplier_data['supplier_name']}' already exists"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Check for duplicate GST number if provided
        if supplier_data.get('gst_registration_number'):
            logger.debug(f"Validating GST number: {supplier_data.get('gst_registration_number')}")
            
            existing_gst = session.query(Supplier).filter(
                Supplier.hospital_id == hospital_id,
                Supplier.gst_registration_number == supplier_data['gst_registration_number']
            ).first()
            
            if existing_gst:
                error_msg = f"Supplier with GST number '{supplier_data['gst_registration_number']}' already exists"
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        logger.debug("Creating new supplier object")
        
        # Create new supplier with branch_id
        supplier = Supplier(
            hospital_id=hospital_id,
            branch_id=branch_id,  # NEW: Include branch_id
            supplier_name=supplier_data['supplier_name'],
            supplier_category=supplier_data.get('supplier_category'),
            supplier_address=supplier_data.get('supplier_address', {}),
            contact_person_name=supplier_data.get('contact_person_name'),
            contact_info=supplier_data.get('contact_info', {}),
            manager_name=supplier_data.get('manager_name'),
            manager_contact_info=supplier_data.get('manager_contact_info', {}),
            email=supplier_data.get('email'),
            black_listed=supplier_data.get('black_listed', False),
            performance_rating=supplier_data.get('performance_rating'),
            payment_terms=supplier_data.get('payment_terms'),
            gst_registration_number=supplier_data.get('gst_registration_number'),
            pan_number=supplier_data.get('pan_number'),
            tax_type=supplier_data.get('tax_type'),
            state_code=supplier_data.get('state_code'),
            bank_details=supplier_data.get('bank_details', {}),
            remarks=supplier_data.get('remarks'),
            status=supplier_data.get('status', 'active')
        )
        
        if current_user_id:
            supplier.created_by = current_user_id
            logger.debug(f"Setting created_by to: {current_user_id}")
            
        session.add(supplier)
        session.flush()
        
        logger.info(f"Supplier created with ID: {supplier.supplier_id}")
        
        # Return the created supplier
        return get_entity_dict(supplier)
        
    except IntegrityError as e:
        logger.error(f"Database integrity error creating supplier: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error creating supplier: {str(e)}", exc_info=True)
        session.rollback()
        raise

def update_supplier(
    supplier_id: uuid.UUID,
    supplier_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Update an existing supplier record - EXISTING FUNCTION (unchanged)
    """
    logger.info(f"Updating supplier {supplier_id} for hospital {hospital_id}")
    logger.debug(f"Update data: {supplier_data}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _update_supplier(session, supplier_id, supplier_data, hospital_id, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _update_supplier(new_session, supplier_id, supplier_data, hospital_id, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing supplier update: {supplier_id}")
        new_session.commit()
        
        logger.info(f"Successfully updated supplier: {result.get('supplier_name')}")
        return result

def _update_supplier(
    session: Session,
    supplier_id: uuid.UUID,
    supplier_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to update an existing supplier record within a session - EXISTING FUNCTION (unchanged)
    """
    try:
        logger.debug(f"Fetching supplier {supplier_id}")
        
        # Get the supplier
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Check for duplicate supplier name if being changed
        if 'supplier_name' in supplier_data and supplier_data['supplier_name'] != supplier.supplier_name:
            logger.debug(f"Checking for duplicate name: {supplier_data['supplier_name']}")
            
            existing_supplier = session.query(Supplier).filter(
                Supplier.hospital_id == hospital_id,
                Supplier.supplier_name == supplier_data['supplier_name'],
                Supplier.supplier_id != supplier_id
            ).first()
            
            if existing_supplier:
                error_msg = f"Supplier with name '{supplier_data['supplier_name']}' already exists"
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        # Check for duplicate GST number if being changed
        if 'gst_registration_number' in supplier_data and supplier_data['gst_registration_number'] != supplier.gst_registration_number:
            logger.debug(f"Checking for duplicate GST: {supplier_data['gst_registration_number']}")
            
            existing_gst = session.query(Supplier).filter(
                Supplier.hospital_id == hospital_id,
                Supplier.gst_registration_number == supplier_data['gst_registration_number'],
                Supplier.supplier_id != supplier_id
            ).first()
            
            if existing_gst:
                error_msg = f"Supplier with GST number '{supplier_data['gst_registration_number']}' already exists"
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        logger.debug("Updating supplier fields")
        
        # Update fields
        updatable_fields = [
            'supplier_name', 'supplier_category', 'supplier_address',
            'contact_person_name', 'contact_info', 'manager_name', 
            'manager_contact_info', 'email', 'black_listed', 
            'performance_rating', 'payment_terms', 'gst_registration_number',
            'pan_number', 'tax_type', 'state_code', 'bank_details',
            'remarks', 'status'
        ]
        
        for field in updatable_fields:
            if field in supplier_data:
                logger.debug(f"Updating field {field}: {supplier_data[field]}")
                setattr(supplier, field, supplier_data[field])
        
        if current_user_id:
            supplier.updated_by = current_user_id
            logger.debug(f"Setting updated_by to: {current_user_id}")
            
        # Update the timestamp
        supplier.updated_at = datetime.now(timezone.utc)
        
        session.flush()
        
        logger.info(f"Supplier {supplier_id} updated successfully")
        
        # Return the updated supplier
        return get_entity_dict(supplier)
        
    except IntegrityError as e:
        logger.error(f"Database integrity error updating supplier: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error updating supplier: {str(e)}", exc_info=True)
        session.rollback()
        raise

def get_supplier_by_id(
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Get a supplier by ID - EXISTING FUNCTION (unchanged)
    """
    logger.info(f"Fetching supplier {supplier_id} for hospital {hospital_id}")
    
    if session is not None:
        return _get_supplier_by_id(session, supplier_id, hospital_id)
    
    with get_db_session() as new_session:
        return _get_supplier_by_id(new_session, supplier_id, hospital_id)

def _get_supplier_by_id(
    session: Session,
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None

) -> Dict:
    """
    Internal function to get a supplier by ID within a session - EXISTING FUNCTION (unchanged)
    """
    try:
        logger.debug(f"Querying supplier {supplier_id}")
        
        # Get the supplier
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)

        if current_user_id and current_user_id != '7777777777':
            if not validate_entity_branch_access(current_user_id, hospital_id, supplier_id, 'supplier', 'view'):
                logger.warning(f"User {current_user_id} denied access to supplier {supplier_id}")
                raise ValueError("Access denied to supplier")


        logger.info(f"Found supplier: {supplier.supplier_name}")
        
        # Return the supplier as a dictionary
        return get_entity_dict(supplier)
        
    except Exception as e:
        logger.error(f"Error getting supplier: {str(e)}", exc_info=True)
        raise

def search_suppliers(
    hospital_id: uuid.UUID,
    name: Optional[str] = None,
    category: Optional[str] = None,
    gst_number: Optional[str] = None,
    status: Optional[str] = None,
    blacklisted: Optional[bool] = None,
    branch_id: Optional[uuid.UUID] = None,
    current_user_id: Optional[str] = None,
    page: int = 1,
    per_page: int = 20,
    session: Optional[Session] = None
) -> Dict:
    """
    Search suppliers with filtering and pagination - FIXED session management
    """
    logger.info(f"Searching suppliers for hospital {hospital_id}")
    logger.debug(f"Filters: name={name}, category={category}, gst={gst_number}, status={status}, blacklisted={blacklisted}, branch={branch_id}")
    logger.debug(f"Pagination: page={page}, per_page={per_page}")
    
    # CRITICAL FIX: Avoid nested session calls that cause conflicts
    if not branch_id and current_user_id:
        # Instead of calling get_user_accessible_branches (which opens its own session),
        # we'll handle branch logic within the main session
        pass  # We'll handle this in the _search_suppliers function
    
    if session is not None:
        return _search_suppliers(
            session, hospital_id, name, category, gst_number, 
            status, blacklisted, branch_id, current_user_id, page, per_page
        )
    
    with get_db_session() as new_session:
        return _search_suppliers(
            new_session, hospital_id, name, category, gst_number, 
            status, blacklisted, branch_id, current_user_id, page, per_page
        )

def _search_suppliers(
    session: Session,
    hospital_id: uuid.UUID,
    name: Optional[str] = None,
    category: Optional[str] = None,
    gst_number: Optional[str] = None,
    status: Optional[str] = None,
    blacklisted: Optional[bool] = None,
    branch_id: Optional[uuid.UUID] = None,
    current_user_id: Optional[str] = None,
    page: int = 1,
    per_page: int = 20
) -> Dict:
    """
    Internal function to search suppliers within a session - FIXED with branch support
    """
    try:
        logger.debug("Building supplier search query with branch support")
        
        # CRITICAL FIX: Handle user branch detection within the same session
        if not branch_id and current_user_id:
            try:
                from app.models.master import Staff
                from app.models.transaction import User
                
                # Get user's branch within this session
                user = session.query(User).filter_by(
                    user_id=current_user_id,
                    hospital_id=hospital_id
                ).first()
                
                if user and user.entity_type == 'staff' and user.entity_id:
                    staff = session.query(Staff).filter_by(
                        staff_id=user.entity_id,
                        hospital_id=hospital_id
                    ).first()
                    if staff and staff.branch_id:
                        branch_id = staff.branch_id
                        logger.debug(f"Auto-detected user branch: {branch_id}")
            except Exception as branch_error:
                logger.warning(f"Could not auto-detect user branch: {str(branch_error)}")
        
        # Base query
        query = session.query(Supplier).filter_by(hospital_id=hospital_id)
        
        # Add branch filter if specified or auto-detected
        if branch_id:
            logger.debug(f"Filtering by branch: {branch_id}")
            query = query.filter(Supplier.branch_id == branch_id)

        # Apply existing filters (unchanged)
        if name:
            logger.debug(f"Filtering by name: {name}")
            query = query.filter(Supplier.supplier_name.ilike(f"%{name}%"))
            
        if category:
            logger.debug(f"Filtering by category: {category}")
            query = query.filter(Supplier.supplier_category == category)
            
        if gst_number:
            logger.debug(f"Filtering by GST: {gst_number}")
            query = query.filter(Supplier.gst_registration_number.ilike(f"%{gst_number}%"))
            
        if status:
            logger.debug(f"Filtering by status: {status}")
            query = query.filter(Supplier.status == status)
            
        if blacklisted is not None:
            logger.debug(f"Filtering by blacklisted: {blacklisted}")
            query = query.filter(Supplier.black_listed == blacklisted)

        # NEW: Apply branch filtering if user provided (non-disruptive addition)
        if current_user_id and current_user_id != '7777777777':
            try:
                query = apply_branch_filter_to_query(query, current_user_id, hospital_id, 'supplier')
                logger.debug(f"Applied branch filtering for user {current_user_id}")
            except Exception as e:
                logger.warning(f"Branch filtering failed, proceeding without: {str(e)}")

        # Count total for pagination
        total_count = query.count()
        logger.debug(f"Total suppliers found: {total_count}")
        
        # Apply pagination
        query = query.order_by(Supplier.supplier_name)
        query = query.offset((page - 1) * per_page).limit(per_page)
        
        # Execute query
        suppliers = query.all()
        logger.debug(f"Retrieved {len(suppliers)} suppliers for current page")
        
        # Convert to dictionaries
        supplier_list = [get_entity_dict(supplier) for supplier in suppliers]
        
        # Prepare pagination info
        pagination = {
            'page': page,
            'per_page': per_page,
            'total_count': total_count,
            'total_pages': (total_count + per_page - 1) // per_page
        }
        
        logger.info(f"Search completed: {len(supplier_list)} suppliers returned")
        
        return {
            'suppliers': supplier_list,
            'pagination': pagination
        }
        
    except Exception as e:
        logger.error(f"Error searching suppliers: {str(e)}", exc_info=True)
        raise

# app/services/supplier_service.py - COMPLETE PART 3: Purchase Order Management Functions
# REPLACE all purchase order functions with these

# ===================================================================
# Purchase Order Management - COMPLETE WITH BRANCH SUPPORT
# ===================================================================

def validate_po_line_item(line_data: Dict) -> List[str]:
    """
    Enhanced validation - focus on business rules for free and regular items - EXISTING FUNCTION (unchanged)
    """
    errors = []
    
    try:
        # Extract values safely with proper type conversion
        quantity = float(line_data.get('units', 0))
        unit_rate = float(line_data.get('pack_purchase_price', 0))
        unit_mrp = float(line_data.get('pack_mrp', 0))
        discount_percent = float(line_data.get('discount_percent', 0))
        is_free_item = line_data.get('is_free_item', False)
        
        # Convert string boolean to actual boolean if needed
        if isinstance(is_free_item, str):
            is_free_item = is_free_item.lower() in ('true', '1', 'yes')
        
        # CRITICAL: Quantity validation applies to ALL items (including free items)
        if quantity <= 0:
            errors.append("Quantity must be greater than 0")
            
        # Discount validation
        if discount_percent > 100 or discount_percent < 0:
            errors.append("Discount must be between 0 and 100%")
        
        # Free item specific validation
        if is_free_item:
            # Free items: zero rate, zero discount, but quantity > 0 is allowed
            if unit_rate > 0:
                errors.append("Free items should have zero rate")
            if discount_percent > 0:
                errors.append("Free items cannot have discount")
            # Note: MRP can be non-zero for free items (for reference purposes)
            
        else:
            # Non-free items: must have positive rate
            if unit_rate <= 0:
                errors.append("Rate must be greater than 0 for non-free items")
            
            # MRP validation for non-free items (optional but recommended)
            if unit_mrp > 0 and unit_rate > 0:
                # Check if rate + GST doesn't exceed MRP (basic business rule)
                gst_rate = float(line_data.get('gst_rate', 0))
                discounted_rate = unit_rate * (1 - discount_percent / 100)
                
                # Simple check: if discounted rate exceeds MRP, warn
                if discounted_rate > unit_mrp:
                    difference = discounted_rate - unit_mrp
                    errors.append(
                        f"Purchase rate ( Rs.{discounted_rate:.2f}) exceeds MRP ( Rs.{unit_mrp:.2f}) by  Rs.{difference:.2f}"
                    )
        
        # Medicine ID validation
        medicine_id = line_data.get('medicine_id')
        if not medicine_id:
            errors.append("Medicine selection is required")
        
        # Log validation for debugging free items
        if is_free_item:
            logger.debug(f"Validating free item: Qty={quantity}, Rate={unit_rate}, Errors={len(errors)}")
        
    except (ValueError, TypeError) as e:
        errors.append(f"Invalid data format: {str(e)}")
        logger.error(f"Validation error for line item: {str(e)}")
        
    return errors

def create_purchase_order(
    hospital_id: uuid.UUID,
    po_data: Dict,
    current_user_id: Optional[str] = None,
    branch_id: Optional[uuid.UUID] = None,  # NEW: Optional parameter
    session: Optional[Session] = None,
    skip_validation: bool = False
) -> Dict:
    """
    Create a new purchase order - UPDATED to use branch_service
    """
    # NEW: Use branch service to determine branch
    determined_branch_id = get_branch_for_supplier_operation(
        current_user_id=current_user_id,
        hospital_id=hospital_id,
        specified_branch_id=branch_id
    )
    
    if not determined_branch_id:
        raise ValueError("Could not determine branch for PO creation")
    
    # Add branch to PO data
    po_data['branch_id'] = determined_branch_id

    logger.info(f"Creating purchase order for hospital {hospital_id}, branch {determined_branch_id}")
    logger.debug(f"PO data: {po_data}")
    
    # Server-side validation (existing logic unchanged)
    if not skip_validation:
        logger.debug("Performing server-side validation")
        
        # Validate line items
        line_items = po_data.get('line_items', [])
        if not line_items:
            error_msg = "At least one line item is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Validate each line item
        validation_errors = []
        for idx, item_data in enumerate(line_items):
            line_errors = validate_po_line_item(item_data)
            if line_errors:
                for error in line_errors:
                    validation_errors.append(f"Line {idx + 1}: {error}")
        
        # If validation errors exist, raise exception
        if validation_errors:
            error_msg = "Purchase order validation failed:\n" + "\n".join(validation_errors)
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.debug("Server-side validation passed")
    else:
        logger.debug("Skipping server-side validation (backward compatibility mode)")
    
    # Proceed with PO creation using existing logic
    if session is not None:
        logger.debug("Using provided session")
        return _create_purchase_order(session, hospital_id, po_data, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _create_purchase_order(new_session, hospital_id, po_data, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing purchase order creation: {result.get('po_number')}")
        new_session.commit()
        
        logger.info(f"Successfully created purchase order: {result.get('po_number')}")
        return result

def create_purchase_order_with_validation(
    hospital_id: uuid.UUID,
    po_data: Dict,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Create a new purchase order with mandatory server-side validation - EXISTING FUNCTION (unchanged)
    """
    return create_purchase_order(
        hospital_id=hospital_id,
        po_data=po_data,
        current_user_id=current_user_id,
        session=session,
        skip_validation=False  # Force validation
    )

def create_purchase_order_legacy(
    hospital_id: uuid.UUID,
    po_data: Dict,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Create a new purchase order without validation (legacy mode) - EXISTING FUNCTION (unchanged)
    """
    return create_purchase_order(
        hospital_id=hospital_id,
        po_data=po_data,
        current_user_id=current_user_id,
        session=session,
        skip_validation=True  # Skip validation
    )

def _create_purchase_order(
    session: Session,
    hospital_id: uuid.UUID,
    po_data: Dict,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to create a purchase order within a session - UPDATED with branch support
    """
    try:
        # Validate supplier
        supplier_id = po_data.get('supplier_id')
        if not supplier_id:
            error_msg = "Supplier ID is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Validating supplier {supplier_id}")
        
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Check if supplier is blacklisted
        if supplier.black_listed:
            error_msg = f"Supplier '{supplier.supplier_name}' is blacklisted"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Validate line items
        line_items = po_data.get('line_items', [])
        if not line_items:
            error_msg = "At least one line item is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Processing {len(line_items)} line items")
            
        # Generate PO number
        po_number = _generate_po_number(session, hospital_id)
        logger.info(f"Generated PO number: {po_number}")
        
        # Get hospital state code for interstate calculation
        hospital = session.query(Hospital).filter_by(hospital_id=hospital_id).first()
        hospital_state_code = hospital.state_code if hospital else None
        supplier_state_code = supplier.state_code
        is_interstate = (hospital_state_code != supplier_state_code) if (hospital_state_code and supplier_state_code) else False
        
        logger.debug(f"Interstate transaction: {is_interstate} (Hospital: {hospital_state_code}, Supplier: {supplier_state_code})")
        
        # Create PO header with branch support
        po_header = PurchaseOrderHeader(
            hospital_id=hospital_id,
            branch_id=po_data.get('branch_id'),  # NEW: Add branch_id
            po_number=po_number,
            po_date=po_data.get('po_date', datetime.now(timezone.utc)),
            supplier_id=supplier_id,
            quotation_id=po_data.get('quotation_id'),
            quotation_date=po_data.get('quotation_date'),
            expected_delivery_date=po_data.get('expected_delivery_date'),
            currency_code=po_data.get('currency_code', 'INR'),
            exchange_rate=po_data.get('exchange_rate', 1.0),
            status=po_data.get('status', 'draft'),
            deleted_flag=False
        )
        
        # Add additional fields from po_data
        if po_data.get('delivery_instructions'):
            po_header.delivery_instructions = po_data.get('delivery_instructions')
        if po_data.get('terms_conditions'):
            po_header.terms_conditions = po_data.get('terms_conditions')
        if po_data.get('notes'):
            po_header.notes = po_data.get('notes')
        
        if current_user_id:
            po_header.created_by = current_user_id
            logger.debug(f"Setting created_by to: {current_user_id}")
            
        session.add(po_header)
        session.flush()  # To get the po_id
        
        logger.debug(f"Created PO header with ID: {po_header.po_id}")
        
        # Create PO lines using standardized GST calculation
        po_lines = []
        total_amount = Decimal('0')
        
        for idx, item_data in enumerate(line_items):
            logger.debug(f"Processing line item {idx + 1}")
            
            # Validate medicine
            medicine_id = item_data.get('medicine_id')
            if not medicine_id:
                error_msg = f"Medicine ID is required for line item {idx + 1}"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            medicine = session.query(Medicine).filter_by(
                medicine_id=medicine_id,
                hospital_id=hospital_id
            ).first()
            
            if not medicine:
                error_msg = f"Medicine with ID {medicine_id} not found"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            # Extract line item data using business-friendly terms
            quantity = float(item_data.get('units', 0))  # Purchase quantity
            unit_rate = float(item_data.get('pack_purchase_price', 0))  # Vendor's rate per unit
            unit_mrp = float(item_data.get('pack_mrp', 0))  # MRP per unit
            conversion_factor = float(item_data.get('units_per_pack', 1))  # Sub-units per unit
            discount_percent = float(item_data.get('discount_percent', 0))
            is_free_item = item_data.get('is_free_item', False)
            
            # Get GST rate from medicine master or override from item data
            gst_rate = float(item_data.get('gst_rate', medicine.gst_rate or 0))
            
            logger.debug(f"Line {idx + 1}: Qty={quantity}, Rate={unit_rate}, GST={gst_rate}%, Discount={discount_percent}%, Free={is_free_item}")
            
            # Calculate all amounts using standardized function
            calculations = calculate_gst_values(
                quantity=quantity,
                unit_rate=unit_rate,
                gst_rate=gst_rate,
                discount_percent=discount_percent,
                is_free_item=is_free_item,
                is_interstate=is_interstate,
                conversion_factor=conversion_factor
            )
            
            logger.debug(f"Line {idx + 1} calculations completed")
            
            # Create PO line object with all required fields
            po_line = PurchaseOrderLine(
                hospital_id=hospital_id,
                po_id=po_header.po_id,
                medicine_id=medicine_id,
                medicine_name=medicine.medicine_name,
                pack_mrp=unit_mrp,
                hsn_code=medicine.hsn_code or item_data.get('hsn_code', ''),
                expected_delivery_date=item_data.get('expected_delivery_date', po_header.expected_delivery_date)
            )
            
            # Apply standardized GST calculations to the line object
            apply_gst_calculations_to_line(po_line, calculations)
            
            if current_user_id:
                po_line.created_by = current_user_id
                
            session.add(po_line)
            po_lines.append(po_line)
            
            # Add to total (using the calculated line total)
            total_amount += Decimal(str(po_line.line_total))
            
            logger.debug(f"Line {idx + 1} stored: Rate={po_line.unit_rate}, Discounted={po_line.discounted_rate}, GST={po_line.total_gst}, Total={po_line.line_total}")
            
        # Update header total amount
        po_header.total_amount = total_amount
        logger.info(f"PO total amount: {total_amount}")
        
        session.flush()
        
        # Return the created PO with lines
        result = get_entity_dict(po_header)
        result['line_items'] = [get_entity_dict(line) for line in po_lines]
        result['supplier_name'] = supplier.supplier_name
        
        logger.info(f"Successfully created PO {po_number} with {len(po_lines)} line items")
        
        return result
        
    except IntegrityError as e:
        logger.error(f"Database integrity error creating purchase order: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error creating purchase order: {str(e)}", exc_info=True)
        session.rollback()
        raise

def _generate_po_number(session: Session, hospital_id: uuid.UUID) -> str:
    """
    Generate a sequential purchase order number - EXISTING FUNCTION (unchanged)
    """
    try:
        # Get current financial year
        today = datetime.now()
        if today.month >= 4:  # Financial year starts in April
            fin_year_start = today.year
            fin_year_end = today.year + 1
        else:
            fin_year_start = today.year - 1
            fin_year_end = today.year
            
        fin_year = f"{fin_year_start}-{fin_year_end}"
        logger.debug(f"Generating PO number for financial year: {fin_year}")
        
        # Get the last PO number for this hospital and financial year
        last_po = session.query(PurchaseOrderHeader).filter(
            PurchaseOrderHeader.hospital_id == hospital_id,
            PurchaseOrderHeader.po_number.like(f"PO/{fin_year}/%")
        ).order_by(PurchaseOrderHeader.po_number.desc()).first()
        
        if last_po:
            # Extract the sequence number from the last PO number
            try:
                seq_num = int(last_po.po_number.split('/')[-1])
                new_seq_num = seq_num + 1
                logger.debug(f"Last PO sequence: {seq_num}, new sequence: {new_seq_num}")
            except:
                # If parsing fails, start from 1
                new_seq_num = 1
                logger.warning("Failed to parse last PO number, starting from 1")
        else:
            # No existing PO for this hospital and financial year
            new_seq_num = 1
            logger.debug("No existing PO for this year, starting from 1")
            
        # Format the new PO number
        po_number = f"PO/{fin_year}/{new_seq_num:05d}"
        logger.info(f"Generated PO number: {po_number}")
        
        return po_number
        
    except Exception as e:
        logger.error(f"Error generating PO number: {str(e)}", exc_info=True)
        raise

def update_purchase_order_status(
    po_id: uuid.UUID,
    status: str,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Update the status of a purchase order - EXISTING FUNCTION (unchanged)
    """
    logger.info(f"Updating PO {po_id} status to {status}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _update_purchase_order_status(session, po_id, status, hospital_id, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _update_purchase_order_status(new_session, po_id, status, hospital_id, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing PO status update: {po_id}")
        new_session.commit()
        
        logger.info(f"Successfully updated PO status to: {status}")
        return result

def _update_purchase_order_status(
    session: Session,
    po_id: uuid.UUID,
    status: str,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to update the status of a purchase order within a session - EXISTING FUNCTION (unchanged)
    """
    try:
        # Validate status
        valid_statuses = ['draft', 'approved', 'received', 'cancelled']
        if status not in valid_statuses:
            error_msg = f"Invalid status '{status}'. Must be one of: {', '.join(valid_statuses)}"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Fetching PO {po_id}")
            
        # Get the PO
        po = session.query(PurchaseOrderHeader).filter_by(
            po_id=po_id,
            hospital_id=hospital_id
        ).first()
        
        if not po:
            error_msg = f"Purchase order with ID {po_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Check if status transition is allowed
        current_status = po.status
        logger.debug(f"Current status: {current_status}, new status: {status}")
        
        # Rules for status transitions
        if current_status == 'cancelled':
            error_msg = "Cannot change status of a cancelled purchase order"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        if current_status == 'received' and status != 'cancelled':
            error_msg = "Received purchase orders can only be cancelled"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        if current_status == 'approved' and status == 'draft':
            error_msg = "Cannot change status from approved to draft"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Update the status
        po.status = status
        
        if status == 'approved':
            po.approved_by = current_user_id
            logger.debug(f"Setting approved_by to: {current_user_id}")
            
        if current_user_id:
            po.updated_by = current_user_id
            
        # Update the timestamp
        po.updated_at = datetime.now(timezone.utc)
        
        session.flush()
        
        logger.info(f"PO {po.po_number} status updated to {status}")
        
        # Return the updated PO with lines
        return get_po_with_lines(session, po)
        
    except Exception as e:
        logger.error(f"Error updating purchase order status: {str(e)}", exc_info=True)
        session.rollback()
        raise

def get_po_with_lines(session: Session, po: PurchaseOrderHeader) -> Dict:
    """
    Helper function to get a PO with its line items - EXISTING FUNCTION (unchanged)
    """
    try:
        logger.debug(f"Getting PO {po.po_id} with line items")
        
        # Get the line items
        lines = session.query(PurchaseOrderLine).filter_by(po_id=po.po_id).all()
        
        logger.debug(f"Found {len(lines)} line items")
        
        # Get the supplier
        supplier = session.query(Supplier).filter_by(supplier_id=po.supplier_id).first()
        
        if not supplier:
            logger.warning(f"Supplier {po.supplier_id} not found for PO")
        
        # Convert to dictionary
        result = get_entity_dict(po)
        result['line_items'] = [get_entity_dict(line) for line in lines]
        
        if supplier:
            result['supplier_name'] = supplier.supplier_name
            
        return result
        
    except Exception as e:
        logger.error(f"Error getting PO with lines: {str(e)}", exc_info=True)
        raise

def get_purchase_order_by_id(
    po_id: uuid.UUID,
    hospital_id: uuid.UUID,
    session: Optional[Session] = None
) -> Dict:
    """
    Get a purchase order by ID with its line items - EXISTING FUNCTION (unchanged)
    """
    logger.info(f"Fetching purchase order {po_id} for hospital {hospital_id}")
    
    if session is not None:
        return _get_purchase_order_by_id(session, po_id, hospital_id)
    
    with get_db_session() as new_session:
        return _get_purchase_order_by_id(new_session, po_id, hospital_id)

def _get_purchase_order_by_id(
    session: Session,
    po_id: uuid.UUID,
    hospital_id: uuid.UUID
) -> Dict:
    """
    Internal function to get a purchase order by ID within a session - EXISTING FUNCTION (unchanged)
    """
    try:
        logger.debug(f"Querying purchase order {po_id}")
        
        # Get the PO
        po = session.query(PurchaseOrderHeader).filter_by(
            po_id=po_id,
            hospital_id=hospital_id
        ).first()
        
        if not po:
            error_msg = f"Purchase order with ID {po_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.info(f"Found purchase order: {po.po_number}")
            
        # Return the PO with lines
        return get_po_with_lines(session, po)
        
    except Exception as e:
        logger.error(f"Error getting purchase order: {str(e)}", exc_info=True)
        raise

def search_purchase_orders(
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    po_number: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    branch_id: Optional[uuid.UUID] = None,  # NEW: Added branch filter
    current_user_id: Optional[str] = None,  # NEW: For auto-branch detection
    page: int = 1,
    per_page: int = 20,
    session: Optional[Session] = None
) -> Dict:
    """
    Search purchase orders with filtering and pagination - UPDATED to use branch_service
    """
    logger.info(f"Searching purchase orders for hospital {hospital_id}")
    logger.debug(f"Filters: supplier={supplier_id}, po={po_number}, status={status}, branch={branch_id}")
    logger.debug(f"Date range: {start_date} to {end_date}")
    logger.debug(f"Pagination: page={page}, per_page={per_page}")
    
    # NEW: Use branch service for branch determination
    if not branch_id and current_user_id:
        accessible_branches = get_user_accessible_branches(current_user_id, hospital_id)
        if len(accessible_branches) == 1:
            branch_id = uuid.UUID(accessible_branches[0]['branch_id'])
    
    if session is not None:
        return _search_purchase_orders(
            session, hospital_id, supplier_id, po_number, 
            status, start_date, end_date, branch_id, page, per_page
        )
    
    with get_db_session() as new_session:
        return _search_purchase_orders(
            new_session, hospital_id, supplier_id, po_number, 
            status, start_date, end_date, branch_id, page, per_page
        )

def _search_purchase_orders(
    session: Session,
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    po_number: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    branch_id: Optional[uuid.UUID] = None,  # NEW: Added parameter
    page: int = 1,
    per_page: int = 20
) -> Dict:
    """
    Internal function to search purchase orders within a session - UPDATED with branch support
    """
    try:
        logger.debug("Building purchase order search query with branch support")
        
        # Base query
        query = session.query(PurchaseOrderHeader).filter_by(hospital_id=hospital_id)
        
        # NEW: Add branch filter if specified
        if branch_id:
            logger.debug(f"Filtering by branch: {branch_id}")
            query = query.filter(PurchaseOrderHeader.branch_id == branch_id)
        
        # Apply existing filters (unchanged)
        if supplier_id:
            logger.debug(f"Filtering by supplier: {supplier_id}")
            query = query.filter(PurchaseOrderHeader.supplier_id == supplier_id)
            
        if po_number:
            logger.debug(f"Filtering by PO number: {po_number}")
            query = query.filter(PurchaseOrderHeader.po_number.ilike(f"%{po_number}%"))
            
        if status:
            logger.debug(f"Filtering by status: {status}")
            query = query.filter(PurchaseOrderHeader.status == status)
            
        if start_date:
            logger.debug(f"Filtering by start date: {start_date}")
            query = query.filter(PurchaseOrderHeader.po_date >= start_date)
            
        if end_date:
            logger.debug(f"Filtering by end date: {end_date}")
            query = query.filter(PurchaseOrderHeader.po_date <= end_date)
            
        # Count total for pagination
        total_count = query.count()
        logger.debug(f"Total purchase orders found: {total_count}")
        
        # Apply pagination
        query = query.order_by(desc(PurchaseOrderHeader.po_date))
        query = query.offset((page - 1) * per_page).limit(per_page)
        
        # Execute query
        pos = query.all()
        logger.debug(f"Retrieved {len(pos)} purchase orders for current page")
        
        # Convert to dictionaries
        po_list = []
        
        for po in pos:
            po_dict = get_entity_dict(po)
            
            # Get the supplier name
            supplier = session.query(Supplier).filter_by(supplier_id=po.supplier_id).first()
            if supplier:
                po_dict['supplier_name'] = supplier.supplier_name
                
            # Add line count
            line_count = session.query(PurchaseOrderLine).filter_by(po_id=po.po_id).count()
            po_dict['line_count'] = line_count
            
            po_list.append(po_dict)
        
        # Prepare pagination info
        pagination = {
            'page': page,
            'per_page': per_page,
            'total_count': total_count,
            'total_pages': (total_count + per_page - 1) // per_page
        }
        
        logger.info(f"Search completed: {len(po_list)} purchase orders returned")
        
        return {
            'purchase_orders': po_list,
            'pagination': pagination
        }
        
    except Exception as e:
        logger.error(f"Error searching purchase orders: {str(e)}", exc_info=True)
        raise

def update_purchase_order(
    po_id: uuid.UUID,
    po_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Update an existing purchase order (draft status only) - EXISTING FUNCTION (unchanged)
    """
    logger.info(f"Updating purchase order {po_id}")
    
    if session is not None:
        return _update_purchase_order(session, po_id, po_data, hospital_id, current_user_id)
    
    with get_db_session() as new_session:
        result = _update_purchase_order(new_session, po_id, po_data, hospital_id, current_user_id)
        new_session.commit()
        logger.info(f"Successfully updated PO: {result.get('po_number')}")
        return result
    
# app/services/supplier_service.py - COMPLETE PART 4: Invoice and Search Functions
# REPLACE all invoice and search functions with these

def _update_purchase_order(
    session: Session,
    po_id: uuid.UUID,
    po_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to update purchase order within a session - EXISTING FUNCTION (unchanged)
    """
    try:
        # Get existing PO
        po = session.query(PurchaseOrderHeader).filter_by(
            po_id=po_id,
            hospital_id=hospital_id
        ).first()
        
        if not po:
            raise ValueError(f"Purchase order with ID {po_id} not found")
        
        # Only allow updating draft POs
        if po.status != 'draft':
            raise ValueError(f"Cannot update purchase order in '{po.status}' status. Only draft purchase orders can be edited.")
        
        logger.debug(f"Updating PO {po.po_number} from draft status")
        
        # Safe field updates - only update fields that actually exist on the model
        updatable_fields = [
            'supplier_id', 'po_date', 'expected_delivery_date', 'quotation_id',
            'quotation_date', 'currency_code', 'exchange_rate', 'notes'
        ]
        
        # Add optional fields only if they exist on the model
        optional_fields = ['delivery_instructions', 'terms_conditions']
        for field in optional_fields:
            if hasattr(po, field):
                updatable_fields.append(field)
                logger.debug(f"Model has field: {field}")
            else:
                logger.warning(f"Model missing field: {field} - skipping")
        
        for field in updatable_fields:
            if field in po_data:
                old_value = getattr(po, field, None)
                new_value = po_data[field]
                if old_value != new_value:
                    setattr(po, field, new_value)
                    logger.debug(f"Updated {field}: {old_value} -> {new_value}")
        
        # Update metadata
        po.updated_by = current_user_id
        po.updated_at = datetime.now(timezone.utc)
        
        # Delete existing line items with logging
        existing_lines = session.query(PurchaseOrderLine).filter_by(po_id=po_id).all()
        logger.debug(f"Found {len(existing_lines)} existing line items to delete")
        
        for line in existing_lines:
            logger.debug(f"Deleting line: {line.medicine_name}, Qty: {line.units}, Price: {line.pack_purchase_price}")
        
        deleted_count = session.query(PurchaseOrderLine).filter_by(po_id=po_id).delete()
        logger.debug(f"Deleted {deleted_count} existing line items")
        
        # Recreate line items with fresh GST calculations
        line_items = po_data.get('line_items', [])
        if not line_items:
            raise ValueError("At least one line item is required")
        
        logger.debug(f"Processing {len(line_items)} line items for update")
        
        # Get supplier and hospital for interstate calculation
        supplier = session.query(Supplier).filter_by(supplier_id=po.supplier_id).first()
        if not supplier:
            raise ValueError(f"Supplier with ID {po.supplier_id} not found")
            
        hospital = session.query(Hospital).filter_by(hospital_id=hospital_id).first()
        hospital_state_code = hospital.state_code if hospital else None
        supplier_state_code = supplier.state_code
        is_interstate = (hospital_state_code != supplier_state_code) if (
            hospital_state_code and supplier_state_code
        ) else False
        
        logger.debug(f"Interstate transaction: {is_interstate} (Hospital: {hospital_state_code}, Supplier: {supplier_state_code})")
        
        # Process line items with fresh calculations
        total_amount = Decimal('0')
        
        for idx, item_data in enumerate(line_items):
            logger.debug(f"Processing line item {idx + 1}")
            
            # Validate medicine
            medicine_id = item_data.get('medicine_id')
            if not medicine_id:
                raise ValueError(f"Medicine ID is required for line item {idx + 1}")
                
            medicine = session.query(Medicine).filter_by(
                medicine_id=medicine_id,
                hospital_id=hospital_id
            ).first()
            
            if not medicine:
                raise ValueError(f"Medicine with ID {medicine_id} not found")
            
            # Extract line item data with explicit logging
            quantity = float(item_data.get('units', 0))
            unit_rate = float(item_data.get('pack_purchase_price', 0))
            unit_mrp = float(item_data.get('pack_mrp', 0))
            conversion_factor = float(item_data.get('units_per_pack', 1))
            discount_percent = float(item_data.get('discount_percent', 0))
            is_free_item = item_data.get('is_free_item', False)
            gst_rate = float(item_data.get('gst_rate', medicine.gst_rate or 0))
            
            logger.debug(f"Line {idx + 1} INPUT: Medicine={medicine.medicine_name}, "
                        f"Qty={quantity}, Rate={unit_rate}, GST={gst_rate}%, "
                        f"Discount={discount_percent}%, Free={is_free_item}")
            
            # Calculate using standardized GST function
            calculations = calculate_gst_values(
                quantity=quantity,
                unit_rate=unit_rate,
                gst_rate=gst_rate,
                discount_percent=discount_percent,
                is_free_item=is_free_item,
                is_interstate=is_interstate,
                conversion_factor=conversion_factor
            )
            
            # Create new PO line object
            po_line = PurchaseOrderLine(
                hospital_id=hospital_id,
                po_id=po.po_id,
                medicine_id=medicine_id,
                medicine_name=medicine.medicine_name,
                pack_mrp=unit_mrp,
                hsn_code=medicine.hsn_code or item_data.get('hsn_code', ''),
                expected_delivery_date=item_data.get('expected_delivery_date', po.expected_delivery_date)
            )
            
            # Apply standardized GST calculations
            apply_gst_calculations_to_line(po_line, calculations)
            
            if current_user_id:
                po_line.created_by = current_user_id
                po_line.updated_by = current_user_id
                
            session.add(po_line)
            total_amount += Decimal(str(po_line.line_total))
            
            logger.debug(f"Line {idx + 1} STORED: Qty={po_line.units}, "
                        f"Rate={po_line.pack_purchase_price}, "
                        f"Total={po_line.line_total}")
        
        # Update header total amount
        old_total = po.total_amount
        po.total_amount = total_amount
        logger.info(f"Updated PO total amount: {old_total} -> {total_amount}")
        
        session.flush()
        
        # Return updated PO with lines
        result = get_po_with_lines(session, po)
        
        logger.info(f"Successfully updated PO {po.po_number} with {len(line_items)} line items")
        
        return result
        
    except IntegrityError as e:
        logger.error(f"Database integrity error updating purchase order: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error updating purchase order: {str(e)}", exc_info=True)
        session.rollback()
        raise

# ===================================================================
# Supplier Invoice Management - COMPLETE WITH BRANCH SUPPORT
# ===================================================================

def _create_supplier_invoice(
    session: Session,
    hospital_id: uuid.UUID,
    invoice_data: Dict,
    create_stock_entries: bool = True,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    COMPLETE REDESIGNED: Create supplier invoice with integrated accounting entries
    Includes ALL existing features: PO validation, branch support, interstate calculation, 
    stock entries, GL entries, AP subledger - all in single transaction
    """
    try:
        logger.info(f"Creating supplier invoice with complete feature set for hospital {hospital_id}")
        
        # ===================================================================
        # STEP 1: VALIDATE SUPPLIER (EXISTING FEATURE)
        # ===================================================================
        
        supplier_id = invoice_data.get('supplier_id')
        if not supplier_id:
            error_msg = "Supplier ID is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Validating supplier {supplier_id}")
        
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Check if supplier is blacklisted
        if supplier.black_listed:
            error_msg = f"Supplier '{supplier.supplier_name}' is blacklisted"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # ===================================================================
        # STEP 2: VALIDATE LINE ITEMS (EXISTING FEATURE)
        # ===================================================================
            
        line_items = invoice_data.get('line_items', [])
        if not line_items:
            error_msg = "At least one line item is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Processing {len(line_items)} line items")
        
        # ===================================================================
        # STEP 3: GET AND VALIDATE PO IF PROVIDED (MISSING FEATURE - RESTORED)
        # ===================================================================
        
        po = None
        po_id = invoice_data.get('po_id')
        if po_id:
            logger.debug(f"Validating purchase order {po_id}")
            
            po = session.query(PurchaseOrderHeader).filter_by(
                po_id=po_id,
                hospital_id=hospital_id
            ).first()
            
            if not po:
                error_msg = f"Purchase order with ID {po_id} not found"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            # Check if PO is cancelled
            if po.status == 'cancelled':
                error_msg = "Cannot create invoice for a cancelled purchase order"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            # CRITICAL: Check if supplier matches PO supplier
            if str(po.supplier_id) != str(supplier_id):
                error_msg = "Invoice supplier does not match purchase order supplier"
                logger.error(f"Supplier mismatch: PO supplier={po.supplier_id}, Invoice supplier={supplier_id}")
                raise ValueError(error_msg)
                
            logger.info(f"PO validation successful: {po.po_number}")
        
        # ===================================================================
        # STEP 4: CREATE INVOICE HEADER (EXISTING FEATURE WITH ENHANCEMENTS)
        # ===================================================================
        
        logger.debug("Creating invoice header")
        
        invoice = SupplierInvoice(
            hospital_id=hospital_id,
            branch_id=invoice_data.get('branch_id'),
            po_id=po_id,  # Link to PO if provided
            supplier_id=supplier_id,
            supplier_invoice_number=invoice_data.get('supplier_invoice_number'),
            invoice_date=invoice_data.get('invoice_date', datetime.now(timezone.utc)),
            supplier_gstin=invoice_data.get('supplier_gstin', supplier.gst_registration_number),
            place_of_supply=invoice_data.get('place_of_supply', supplier.state_code),
            notes=invoice_data.get('notes'),
            reverse_charge=invoice_data.get('reverse_charge', False),
            currency_code=invoice_data.get('currency_code', 'INR'),
            exchange_rate=invoice_data.get('exchange_rate', 1.0),
            cgst_amount=Decimal('0'),
            sgst_amount=Decimal('0'),
            igst_amount=Decimal('0'),
            total_gst_amount=Decimal('0'),
            total_amount=Decimal('0'),
            itc_eligible=invoice_data.get('itc_eligible', True),
            payment_status='unpaid',
            due_date=invoice_data.get('due_date'),
            created_by=current_user_id
        )
        
        session.add(invoice)
        session.flush()  # Get invoice_id
        
        logger.debug(f"Created invoice header with ID: {invoice.invoice_id}")
        
        # ===================================================================
        # STEP 5: DETERMINE INTERSTATE TRANSACTION (EXISTING FEATURE)
        # ===================================================================
        
        # Get hospital state code for interstate calculation
        hospital = session.query(Hospital).filter_by(hospital_id=hospital_id).first()
        hospital_state_code = hospital.state_code if hospital else None
        supplier_state_code = invoice_data.get('place_of_supply', supplier.state_code)
        is_interstate = (hospital_state_code != supplier_state_code) if (hospital_state_code and supplier_state_code) else False
        
        logger.debug(f"Interstate transaction: {is_interstate} (Hospital: {hospital_state_code}, Supplier: {supplier_state_code})")
        
        # ===================================================================
        # STEP 6: CREATE INVOICE LINES WITH GST CALCULATIONS (EXISTING FEATURE)
        # ===================================================================
        
        invoice_lines = []
        total_amount = Decimal('0')
        total_cgst = Decimal('0')
        total_sgst = Decimal('0')
        total_igst = Decimal('0')
        
        for idx, item_data in enumerate(line_items):
            logger.debug(f"Processing line item {idx + 1}")
            
            # Validate medicine
            medicine_id = item_data.get('medicine_id')
            if not medicine_id:
                error_msg = f"Medicine ID is required for line item {idx + 1}"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            medicine = session.query(Medicine).filter_by(
                medicine_id=medicine_id,
                hospital_id=hospital_id
            ).first()
            
            if not medicine:
                error_msg = f"Medicine with ID {medicine_id} not found"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            # Extract line item data with proper type conversion
            quantity = float(item_data.get('units', 0))
            unit_rate = float(item_data.get('pack_purchase_price', 0))
            unit_mrp = float(item_data.get('pack_mrp', 0))
            conversion_factor = float(item_data.get('units_per_pack', 1))
            discount_percent = float(item_data.get('discount_percent', 0))
            is_free_item = item_data.get('is_free_item', False)
            gst_rate = float(item_data.get('gst_rate', medicine.gst_rate or 0))
            
            logger.debug(f"Line {idx + 1}: Medicine={medicine.medicine_name}, "
                        f"Qty={quantity}, Rate={unit_rate}, GST={gst_rate}%, "
                        f"Discount={discount_percent}%, Free={is_free_item}")
            
            # Calculate using standardized GST function (EXISTING FEATURE)
            calculations = calculate_gst_values(
                quantity=quantity,
                unit_rate=unit_rate,
                gst_rate=gst_rate,
                discount_percent=discount_percent,
                is_free_item=is_free_item,
                is_interstate=is_interstate,
                conversion_factor=conversion_factor
            )
            
            logger.debug(f"Line {idx + 1} calculations completed")
            
            # Create invoice line object with all required fields
            invoice_line = SupplierInvoiceLine(
                hospital_id=hospital_id,
                invoice_id=invoice.invoice_id,
                medicine_id=medicine_id,
                medicine_name=medicine.medicine_name,
                pack_mrp=unit_mrp,
                hsn_code=medicine.hsn_code or item_data.get('hsn_code'),
                batch_number=item_data.get('batch_number'),
                manufacturing_date=item_data.get('manufacturing_date'),
                expiry_date=item_data.get('expiry_date'),
                itc_eligible=item_data.get('itc_eligible', invoice.itc_eligible),
                created_by=current_user_id
            )
            
            # Apply standardized GST calculations to the line object (EXISTING FEATURE)
            apply_gst_calculations_to_invoice_line(invoice_line, calculations)
            
            session.add(invoice_line)
            invoice_lines.append(invoice_line)
            
            # Add to totals (using the calculated line total)
            total_amount += Decimal(str(invoice_line.line_total))
            total_cgst += Decimal(str(invoice_line.cgst))
            total_sgst += Decimal(str(invoice_line.sgst))
            total_igst += Decimal(str(invoice_line.igst))
            
            logger.debug(f"Line {idx + 1} stored: Rate={invoice_line.pack_purchase_price}, "
                        f"Taxable={invoice_line.taxable_amount}, GST={invoice_line.total_gst}, "
                        f"Total={invoice_line.line_total}")
        
        # Update invoice totals
        invoice.cgst_amount = total_cgst
        invoice.sgst_amount = total_sgst
        invoice.igst_amount = total_igst
        invoice.total_gst_amount = total_cgst + total_sgst + total_igst
        invoice.total_amount = total_amount
        
        logger.info(f"Invoice total amount: {total_amount}")
        
        session.flush()  # Ensure invoice and lines are saved before GL entries
        
        # ===================================================================
        # STEP 7: CREATE GL TRANSACTION AND ENTRIES (INTEGRATED FROM HELPER)
        # ===================================================================
        
        logger.info(f"Creating GL entries for invoice {invoice.supplier_invoice_number}")
        
        # Generate posting reference with proper truncation (FROM ENHANCED HELPER)
        doc_prefix = 'INV'
        doc_number = invoice.supplier_invoice_number
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        
        # CRITICAL FIX: Ensure posting reference fits in VARCHAR(50) database field
        base_reference = f"ENH-{doc_prefix}-{doc_number}-{timestamp}"
        
        # Truncate to 50 characters if too long (database constraint)
        if len(base_reference) > 50:
            # Keep the timestamp (14 chars) and truncate the middle part
            available_for_doc = 50 - len(f"ENH-{doc_prefix}--{timestamp}")
            if available_for_doc > 0:
                truncated_doc = doc_number[:available_for_doc] if doc_number else ""
                posting_reference = f"ENH-{doc_prefix}-{truncated_doc}-{timestamp}"
            else:
                posting_reference = f"ENH-{doc_prefix}-{timestamp}"
            
            posting_reference = posting_reference[:50]
            logger.warning(f"Posting reference truncated to fit 50 char limit: {posting_reference}")
        else:
            posting_reference = base_reference
        
        # Create GL Transaction (FROM ENHANCED HELPER)
        gl_transaction = GLTransaction(
            hospital_id=hospital_id,
            transaction_date=invoice.invoice_date,
            transaction_type='SUPPLIER_INVOICE',
            reference_id=str(invoice.invoice_id),
            description=f"Supplier Invoice - {invoice.supplier_invoice_number}",
            source_document_type='SUPPLIER_INVOICE',
            source_document_id=invoice.invoice_id,
            total_debit=total_amount,
            total_credit=total_amount,
            created_by=current_user_id
        )
        
        session.add(gl_transaction)
        session.flush()  # Get transaction_id
        
        # Get account mappings (FROM ENHANCED HELPER)
        from app.services.posting_config_service import get_posting_config
        config = get_posting_config(str(hospital_id))
        
        ap_account_no = config.get('DEFAULT_AP_ACCOUNT', '2100')
        inventory_account_no = config.get('DEFAULT_INVENTORY_ACCOUNT', '1410')
        cgst_account_no = config.get('CGST_RECEIVABLE_ACCOUNT', '1710')
        sgst_account_no = config.get('SGST_RECEIVABLE_ACCOUNT', '1720')
        igst_account_no = config.get('IGST_RECEIVABLE_ACCOUNT', '1730')
        
        # Get account records and validate they exist
        accounts = {}
        account_nos = [ap_account_no, inventory_account_no, cgst_account_no, sgst_account_no, igst_account_no]
        
        for acc_no in account_nos:
            account = session.query(ChartOfAccounts).filter_by(
                hospital_id=hospital_id,
                gl_account_no=acc_no,
                is_active=True
            ).first()
            if not account:
                raise ValueError(f"Account {acc_no} not found in chart of accounts for hospital {hospital_id}")
            accounts[acc_no] = account
        
        # Calculate taxable amount (total - GST)
        taxable_amount = total_amount - total_cgst - total_sgst - total_igst
        
        gl_entries = []
        
        # 1. Inventory/Expense Debit (FROM ENHANCED HELPER)
        if taxable_amount > 0:
            inventory_entry = GLEntry(
                hospital_id=hospital_id,
                transaction_id=gl_transaction.transaction_id,
                account_id=accounts[inventory_account_no].account_id,
                entry_date=invoice.invoice_date,
                description=f"Inventory - {invoice.supplier_invoice_number}",
                debit_amount=taxable_amount,
                credit_amount=Decimal('0'),
                source_document_type='SUPPLIER_INVOICE',
                source_document_id=invoice.invoice_id,
                posting_reference=posting_reference,
                created_by=current_user_id
            )
            session.add(inventory_entry)
            gl_entries.append(inventory_entry)
            logger.debug(f"Created inventory entry: {inventory_account_no} - Debit: {taxable_amount}")
        
        # 2. CGST Debit (FROM ENHANCED HELPER)
        if total_cgst > 0:
            cgst_entry = GLEntry(
                hospital_id=hospital_id,
                transaction_id=gl_transaction.transaction_id,
                account_id=accounts[cgst_account_no].account_id,
                entry_date=invoice.invoice_date,
                description=f"CGST - {invoice.supplier_invoice_number}",
                debit_amount=total_cgst,
                credit_amount=Decimal('0'),
                source_document_type='SUPPLIER_INVOICE',
                source_document_id=invoice.invoice_id,
                posting_reference=posting_reference,
                created_by=current_user_id
            )
            session.add(cgst_entry)
            gl_entries.append(cgst_entry)
            logger.debug(f"Created CGST entry: {cgst_account_no} - Debit: {total_cgst}")
        
        # 3. SGST Debit (FROM ENHANCED HELPER)
        if total_sgst > 0:
            sgst_entry = GLEntry(
                hospital_id=hospital_id,
                transaction_id=gl_transaction.transaction_id,
                account_id=accounts[sgst_account_no].account_id,
                entry_date=invoice.invoice_date,
                description=f"SGST - {invoice.supplier_invoice_number}",
                debit_amount=total_sgst,
                credit_amount=Decimal('0'),
                source_document_type='SUPPLIER_INVOICE',
                source_document_id=invoice.invoice_id,
                posting_reference=posting_reference,
                created_by=current_user_id
            )
            session.add(sgst_entry)
            gl_entries.append(sgst_entry)
            logger.debug(f"Created SGST entry: {sgst_account_no} - Debit: {total_sgst}")
        
        # 4. IGST Debit (FROM ENHANCED HELPER)
        if total_igst > 0:
            igst_entry = GLEntry(
                hospital_id=hospital_id,
                transaction_id=gl_transaction.transaction_id,
                account_id=accounts[igst_account_no].account_id,
                entry_date=invoice.invoice_date,
                description=f"IGST - {invoice.supplier_invoice_number}",
                debit_amount=total_igst,
                credit_amount=Decimal('0'),
                source_document_type='SUPPLIER_INVOICE',
                source_document_id=invoice.invoice_id,
                posting_reference=posting_reference,
                created_by=current_user_id
            )
            session.add(igst_entry)
            gl_entries.append(igst_entry)
            logger.debug(f"Created IGST entry: {igst_account_no} - Debit: {total_igst}")
        
        # 5. AP Credit (FROM ENHANCED HELPER)
        ap_entry = GLEntry(
            hospital_id=hospital_id,
            transaction_id=gl_transaction.transaction_id,
            account_id=accounts[ap_account_no].account_id,
            entry_date=invoice.invoice_date,
            description=f"Accounts Payable - {invoice.supplier_invoice_number}",
            debit_amount=Decimal('0'),
            credit_amount=total_amount,
            source_document_type='SUPPLIER_INVOICE',
            source_document_id=invoice.invoice_id,
            posting_reference=posting_reference,
            created_by=current_user_id
        )
        session.add(ap_entry)
        gl_entries.append(ap_entry)
        logger.debug(f"Created AP entry: {ap_account_no} - Credit: {total_amount}")
        
        # ===================================================================
        # STEP 8: CREATE AP SUBLEDGER ENTRY (FROM ENHANCED HELPER)
        # ===================================================================
        
        logger.info(f"Creating AP subledger entry for invoice {invoice.supplier_invoice_number}")
        
        # Calculate current AP balance for this supplier (FROM ENHANCED HELPER)
        current_balance = session.query(func.coalesce(func.sum(
            APSubledger.credit_amount - APSubledger.debit_amount
        ), 0)).filter_by(
            hospital_id=hospital_id,
            supplier_id=supplier_id
        ).scalar() or Decimal('0')
        
        new_balance = current_balance + total_amount
        
        ap_subledger_entry = APSubledger(
            hospital_id=hospital_id,
            branch_id=invoice.branch_id,
            transaction_date=invoice.invoice_date,
            entry_type='invoice',
            reference_id=invoice.invoice_id,
            reference_type='invoice',
            reference_number=invoice.supplier_invoice_number,
            supplier_id=supplier_id,
            debit_amount=Decimal('0'),
            credit_amount=total_amount,
            current_balance=new_balance,
            gl_transaction_id=gl_transaction.transaction_id,
            created_by=current_user_id
        )
        
        session.add(ap_subledger_entry)
        logger.debug(f"Created AP subledger entry: Credit={total_amount}, New Balance={new_balance}")
        
        # ===================================================================
        # STEP 9: MARK INVOICE AS POSTED (FROM ENHANCED HELPER)
        # ===================================================================
        
        invoice.gl_posted = True
        invoice.inventory_posted = True
        invoice.posting_date = datetime.now(timezone.utc)
        invoice.posting_reference = posting_reference
        
        # ===================================================================
        # STEP 10: CREATE STOCK ENTRIES IF REQUESTED (EXISTING FEATURE)
        # ===================================================================
        
        if create_stock_entries:
            logger.debug("Creating stock entries")
            try:
                stock_result = record_stock_from_supplier_invoice(
                    hospital_id=hospital_id,
                    invoice_id=invoice.invoice_id,
                    current_user_id=current_user_id,
                    session=session
                )
                logger.info("Stock entries created successfully")
            except Exception as e:
                logger.error(f"Error creating stock entries: {str(e)}")
                # In integrated approach, stock failure should fail entire transaction
                raise
        
        session.flush()  # Ensure all entries are saved
        
        # ===================================================================
        # STEP 11: RETURN COMPLETE RESULT (EXISTING FEATURE + ENHANCEMENTS)
        # ===================================================================
        
        result = get_entity_dict(invoice)
        result['line_items'] = [get_entity_dict(line) for line in invoice_lines]
        result['supplier_name'] = supplier.supplier_name
        
        # Enhanced return data with GL posting information
        result['gl_posted'] = True
        result['gl_transaction_id'] = str(gl_transaction.transaction_id)
        result['posting_reference'] = posting_reference
        result['gl_entries_created'] = len(gl_entries)
        result['ap_subledger_created'] = True
        result['stock_entries_created'] = create_stock_entries
        
        # Include PO information if invoice was created from PO
        if po:
            result['po_number'] = po.po_number
            result['po_id'] = str(po.po_id)
            result['created_from_po'] = True
        else:
            result['created_from_po'] = False
        
        # Include accounting summary
        result['accounting_summary'] = {
            'total_amount': float(total_amount),
            'taxable_amount': float(taxable_amount),
            'cgst_amount': float(total_cgst),
            'sgst_amount': float(total_sgst),
            'igst_amount': float(total_igst),
            'is_interstate': is_interstate
        }
        
        logger.info(f"Successfully created invoice {invoice.supplier_invoice_number} with complete accounting integration")
        logger.info(f"GL Transaction: {gl_transaction.transaction_id}, Posting Reference: {posting_reference}")
        
        return result
        
    except Exception as e:
        logger.error(f"Error creating supplier invoice with complete feature set: {str(e)}")
        session.rollback()
        raise


def create_supplier_invoice(
    hospital_id: uuid.UUID,
    invoice_data: Dict,
    branch_id: Optional[uuid.UUID] = None,
    create_stock_entries: bool = True,  # KEPT - still optional
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    SIMPLIFIED: Create supplier invoice with mandatory accounting integration
    Includes ALL features: PO validation, branch support, interstate calculation,
    stock entries, GL entries, AP subledger
    
    Removed create_gl_entries parameter - accounting is always created
    """
    # NEW: Use branch service to determine branch (EXISTING FEATURE)
    determined_branch_id = get_branch_for_supplier_operation(
        current_user_id=current_user_id,
        hospital_id=hospital_id,
        specified_branch_id=branch_id
    )
    
    if not determined_branch_id:
        raise ValueError("Could not determine branch for invoice creation")
    
    # Add branch to invoice data
    invoice_data['branch_id'] = determined_branch_id
    
    logger.info(f"Creating supplier invoice for hospital {hospital_id}, branch {determined_branch_id}")
    
    if session is not None:
        return _create_supplier_invoice(session, hospital_id, invoice_data, create_stock_entries, current_user_id)
    
    with get_db_session() as new_session:
        result = _create_supplier_invoice(new_session, hospital_id, invoice_data, create_stock_entries, current_user_id)
        new_session.commit()
        logger.info(f"Successfully created supplier invoice: {result.get('supplier_invoice_number')}")
        return result

def update_supplier_invoice(
    invoice_id: uuid.UUID,
    invoice_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Update an existing supplier invoice (unpaid invoices only) - EXISTING FUNCTION (unchanged)
    """
    logger.info(f"Updating supplier invoice {invoice_id}")
    
    if session is not None:
        return _update_supplier_invoice(session, invoice_id, invoice_data, hospital_id, current_user_id)
    
    with get_db_session() as new_session:
        result = _update_supplier_invoice(new_session, invoice_id, invoice_data, hospital_id, current_user_id)
        new_session.commit()
        logger.info(f"Successfully updated invoice: {result.get('supplier_invoice_number')}")
        return result

def _update_supplier_invoice(
    session: Session,
    invoice_id: uuid.UUID,
    invoice_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    ENHANCED: Update supplier invoice with proper accounting and inventory reversal/recreation
    This function now properly handles GL entries, AP subledger, and inventory changes
    Following the same patterns as create and cancel functions - all inline
    """
    try:
        logger.info(f"=== ENHANCED SUPPLIER INVOICE EDIT START ===")
        logger.info(f"Updating invoice {invoice_id} with accounting and inventory corrections")
        
        # ===================================================================
        # STEP 1: VALIDATE INVOICE AND ELIGIBILITY (EXISTING LOGIC)
        # ===================================================================
        
        # Get existing invoice
        invoice = session.query(SupplierInvoice).filter_by(
            invoice_id=invoice_id,
            hospital_id=hospital_id
        ).first()
        
        if not invoice:
            raise ValueError(f"Supplier invoice with ID {invoice_id} not found")
        
        # Validate invoice can be edited
        if invoice.payment_status == 'cancelled':
            raise ValueError(f"Cannot update cancelled invoice {invoice.supplier_invoice_number}")
        
        if invoice.payment_status == 'paid':
            raise ValueError(f"Cannot update paid invoice {invoice.supplier_invoice_number}. Create a credit note instead.")
        
        # Check for existing payments
        existing_payments = session.query(SupplierPayment).filter_by(invoice_id=invoice_id).count()
        if existing_payments > 0:
            raise ValueError(f"Cannot update invoice {invoice.supplier_invoice_number} with recorded payments")
        
        logger.info(f"Original invoice {invoice.supplier_invoice_number}: Amount={invoice.total_amount}")
        
        # ===================================================================
        # STEP 2: REVERSE ORIGINAL ACCOUNTING ENTRIES (FOLLOWING CANCEL PATTERN)
        # ===================================================================
        
        if invoice.gl_posted:
            logger.info("Reversing original accounting entries")
            
            # Find original GL transaction
            original_gl_transaction = session.query(GLTransaction).filter_by(
                reference_id=str(invoice_id),
                transaction_type='SUPPLIER_INVOICE',
                hospital_id=hospital_id,
                is_reversal=False
            ).first()
            
            if original_gl_transaction:
                logger.info(f"Found original GL transaction: {original_gl_transaction.transaction_id}")
                
                # Create reversal timestamp and reference
                reversal_timestamp = datetime.now(timezone.utc)
                reversal_reference = f"EDIT-REV-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(invoice_id)[:8]}"
                
                # Create reversal GL transaction
                reversal_gl_transaction = GLTransaction(
                    hospital_id=hospital_id,
                    transaction_date=reversal_timestamp,
                    transaction_type='SUPPLIER_INV_EDIT_REV',  # Shortened for field limits
                    reference_id=str(invoice_id),
                    description=f"EDIT REVERSAL: {original_gl_transaction.description}",
                    source_document_type='SUPPLIER_INVOICE',  # Use standard type
                    source_document_id=invoice_id,
                    total_debit=original_gl_transaction.total_credit,  # Swap amounts
                    total_credit=original_gl_transaction.total_debit,  # Swap amounts
                    is_reversal=True,
                    original_entry_id=original_gl_transaction.transaction_id,  # FIXED: Populate original transaction ID
                    created_by=current_user_id
                )
                
                session.add(reversal_gl_transaction)
                session.flush()  # Get transaction ID
                
                # Reverse GL entries
                original_gl_entries = session.query(GLEntry).filter_by(
                    transaction_id=original_gl_transaction.transaction_id
                ).all()
                
                for original_entry in original_gl_entries:
                    reversal_entry = GLEntry(
                        hospital_id=hospital_id,
                        transaction_id=reversal_gl_transaction.transaction_id,
                        account_id=original_entry.account_id,
                        entry_date=reversal_timestamp,
                        description=f"EDIT REV: {original_entry.description}",
                        debit_amount=original_entry.credit_amount,  # Swap amounts
                        credit_amount=original_entry.debit_amount,  # Swap amounts
                        source_document_type='SUPPLIER_INVOICE',  # Use standard type
                        source_document_id=invoice_id,
                        posting_reference=reversal_reference,
                        is_reversal=True,
                        original_entry_id=original_entry.entry_id,  # FIXED: Populate original entry ID
                        created_by=current_user_id
                    )
                    session.add(reversal_entry)
                
                logger.info(f"Created {len(original_gl_entries)} GL entry reversals")
                
                # Reverse AP subledger entry
                original_ap_entry = session.query(APSubledger).filter_by(
                    reference_id=str(invoice_id),
                    reference_type='invoice',
                    hospital_id=hospital_id
                ).first()
                
                if original_ap_entry:
                    # Get current supplier balance
                    current_balance = session.query(func.sum(APSubledger.credit_amount) - func.sum(APSubledger.debit_amount)).filter_by(
                        hospital_id=hospital_id,
                        supplier_id=original_ap_entry.supplier_id
                    ).scalar() or Decimal('0')
                    
                    # Create reversal AP entry
                    reversal_ap_entry = APSubledger(
                        hospital_id=hospital_id,
                        branch_id=getattr(original_ap_entry, 'branch_id', None) or getattr(invoice, 'branch_id', None),
                        transaction_date=reversal_timestamp,
                        entry_type='reversal',  # Kept short for existing field
                        reference_id=str(invoice_id),
                        reference_type='invoice_reversal',  # FIXED: Updated text for clarity
                        reference_number=invoice.supplier_invoice_number,
                        supplier_id=original_ap_entry.supplier_id,
                        debit_amount=original_ap_entry.credit_amount,  # Reverse the credit
                        credit_amount=Decimal('0'),
                        current_balance=current_balance - original_ap_entry.credit_amount,
                        gl_transaction_id=reversal_gl_transaction.transaction_id,
                        created_by=current_user_id
                    )
                    session.add(reversal_ap_entry)
                    logger.info(f"Created AP subledger reversal entry")
            
        
        # ===================================================================
        # STEP 3: REVERSE ORIGINAL INVENTORY ENTRIES (SAFE VERSION)
        # ===================================================================

        if invoice.inventory_posted:
            logger.info("Reversing original inventory entries")
            
            # ✅ SAFE: Find ONLY inventory entries for THIS SPECIFIC INVOICE
            stock_entries = session.query(Inventory).filter_by(
                hospital_id=hospital_id,
                distributor_invoice_no=invoice.supplier_invoice_number,
                stock_type='Purchase'
            ).all()
            
            logger.info(f"Found {len(stock_entries)} stock entries for invoice {invoice.supplier_invoice_number}")
            
            if stock_entries:
                reversal_timestamp = datetime.now(timezone.utc)
                
                # Log what we're reversing for audit trail
                for i, stock_entry in enumerate(stock_entries, 1):
                    logger.info(f"Reversing stock entry {i}: Medicine={stock_entry.medicine_name}, "
                            f"Batch={stock_entry.batch}, Qty={stock_entry.units}")
                
                for stock_entry in stock_entries:
                    # Use exact same reversal logic as cancel function
                    reversal_stock_entry = Inventory(
                        hospital_id=hospital_id,
                        branch_id=getattr(stock_entry, 'branch_id', None) or getattr(invoice, 'branch_id', None),
                        medicine_id=stock_entry.medicine_id,
                        medicine_name=getattr(stock_entry, 'medicine_name', 'Unknown Medicine'),
                        medicine_category=getattr(stock_entry, 'medicine_category', 'Unknown'),
                        batch=stock_entry.batch,
                        expiry=stock_entry.expiry,
                        pack_purchase_price=getattr(stock_entry, 'pack_purchase_price', 0),
                        pack_mrp=getattr(stock_entry, 'pack_mrp', 0),
                        units_per_pack=getattr(stock_entry, 'units_per_pack', 1),
                        unit_price=getattr(stock_entry, 'unit_price', 0),
                        sale_price=getattr(stock_entry, 'sale_price', 0),
                        units=-int(stock_entry.units),  # Negative quantity for reversal
                        percent_discount=getattr(stock_entry, 'percent_discount', 0),
                        cgst=getattr(stock_entry, 'cgst', 0),
                        sgst=getattr(stock_entry, 'sgst', 0),
                        igst=getattr(stock_entry, 'igst', 0),
                        total_gst=getattr(stock_entry, 'total_gst', 0),
                        stock_type='Purchase_Edit_Reversal',
                        distributor_invoice_no=f"REV-{stock_entry.distributor_invoice_no}",
                        transaction_date=reversal_timestamp,
                        reason=f"Edit reversal: {invoice.supplier_invoice_number}",
                        current_stock=stock_entry.current_stock - int(stock_entry.units),
                        created_by=current_user_id
                    )
                    session.add(reversal_stock_entry)
                
                logger.info(f"Created {len(stock_entries)} precise stock entry reversals")
            else:
                logger.warning(f"No stock entries found for invoice {invoice.supplier_invoice_number}")
        
        # ===================================================================
        # STEP 4: UPDATE INVOICE HEADER (EXISTING LOGIC)
        # ===================================================================
        
        logger.info("Updating invoice header fields")
        
        # Update header fields safely
        updatable_fields = [
            'supplier_id', 'supplier_invoice_number', 'invoice_date', 'po_id',
            'supplier_gstin', 'place_of_supply', 'due_date', 'reverse_charge',
            'itc_eligible', 'currency_code', 'exchange_rate', 'notes'
        ]
        
        for field in updatable_fields:
            if field in invoice_data:
                old_value = getattr(invoice, field, None)
                new_value = invoice_data[field]
                if old_value != new_value:
                    setattr(invoice, field, new_value)
                    logger.debug(f"Updated {field}: {old_value} -> {new_value}")
        
        # Update metadata
        invoice.updated_by = current_user_id
        invoice.updated_at = datetime.now(timezone.utc)
        
        # ===================================================================
        # STEP 5: UPDATE LINE ITEMS (EXISTING LOGIC WITH GST CALCULATIONS)
        # ===================================================================
        
        logger.info("Recreating line items")
        
        # Delete existing line items
        existing_lines = session.query(SupplierInvoiceLine).filter_by(invoice_id=invoice_id).all()
        logger.debug(f"Found {len(existing_lines)} existing line items to delete")
        
        deleted_count = session.query(SupplierInvoiceLine).filter_by(invoice_id=invoice_id).delete()
        logger.debug(f"Deleted {deleted_count} existing line items")
        
        # Recreate line items with fresh GST calculations
        line_items = invoice_data.get('line_items', [])
        if not line_items:
            raise ValueError("At least one line item is required")
        
        logger.debug(f"Processing {len(line_items)} line items for update")
        
        # Get supplier and hospital for interstate calculation (following create pattern)
        supplier = session.query(Supplier).filter_by(
            supplier_id=invoice.supplier_id,
            hospital_id=hospital_id
        ).first()
        if not supplier:
            raise ValueError(f"Supplier with ID {invoice.supplier_id} not found")
            
        hospital = session.query(Hospital).filter_by(hospital_id=hospital_id).first()
        hospital_state_code = hospital.state_code if hospital else None
        supplier_state_code = invoice_data.get('place_of_supply') or supplier.state_code
        is_interstate = (hospital_state_code != supplier_state_code) if (hospital_state_code and supplier_state_code) else False
        
        # Initialize totals for new amounts
        total_amount = Decimal('0')
        total_cgst = Decimal('0')
        total_sgst = Decimal('0')
        total_igst = Decimal('0')
        taxable_amount = Decimal('0')
        invoice_lines = []
        
        # Process each line item (following create pattern exactly)
        for idx, item_data in enumerate(line_items):
            # Validate and get medicine
            medicine_id = item_data.get('medicine_id')
            if not medicine_id:
                raise ValueError(f"Line {idx + 1}: Medicine ID is required")
            
            medicine = session.query(Medicine).filter_by(
                medicine_id=medicine_id,
                hospital_id=hospital_id
            ).first()
            if not medicine:
                raise ValueError(f"Line {idx + 1}: Medicine with ID {medicine_id} not found")
            
            # Validate line item data
            quantity = float(item_data.get('quantity') or item_data.get('units', 0))
            if not quantity or quantity <= 0:
                raise ValueError(f"Line {idx + 1}: Valid quantity is required")
            
            # Extract free item flag first
            is_free_item = item_data.get('is_free_item', False)
            pack_purchase_price = float(item_data.get('pack_purchase_price', 0))

            # Validate based on free item status
            if is_free_item:
                # Free items: price should be 0 or force it to 0
                if pack_purchase_price != 0:
                    logger.info(f"Line {idx + 1}: Adjusting free item price from {pack_purchase_price} to 0")
                    pack_purchase_price = 0.0
                    item_data['pack_purchase_price'] = 0.0
            else:
                # Regular items: must have positive price
                if pack_purchase_price <= 0:
                    raise ValueError(f"Line {idx + 1}: Valid purchase price is required for non-free items")
            
            # Extract other line item data (following create pattern)
            unit_mrp = float(item_data.get('pack_mrp', 0))
            conversion_factor = float(item_data.get('units_per_pack', 1))
            discount_percent = float(item_data.get('discount_percent', 0))
            is_free_item = item_data.get('is_free_item', False)
            gst_rate = float(item_data.get('gst_rate', medicine.gst_rate or 0))
            
            logger.debug(f"Line {idx + 1}: Medicine={medicine.medicine_name}, "
                        f"Qty={quantity}, Rate={pack_purchase_price}, GST={gst_rate}%, "
                        f"Discount={discount_percent}%, Free={is_free_item}")
            
            # Calculate using standardized GST function (EXISTING FUNCTION)
            calculations = calculate_gst_values(
                quantity=quantity,
                unit_rate=pack_purchase_price,
                gst_rate=gst_rate,
                discount_percent=discount_percent,
                is_free_item=is_free_item,
                is_interstate=is_interstate,
                conversion_factor=conversion_factor
            )
            
            logger.debug(f"Line {idx + 1} calculations completed")
            
            # Create new line item (following create pattern)
            invoice_line = SupplierInvoiceLine(
                hospital_id=hospital_id,
                invoice_id=invoice.invoice_id,
                medicine_id=medicine_id,
                medicine_name=medicine.medicine_name,
                pack_mrp=unit_mrp,
                hsn_code=medicine.hsn_code or item_data.get('hsn_code', ''),
                batch_number=item_data.get('batch_number', ''),
                manufacturing_date=item_data.get('manufacturing_date'),
                expiry_date=item_data.get('expiry_date'),
                itc_eligible=item_data.get('itc_eligible', invoice.itc_eligible),
                created_by=current_user_id,
                updated_by=current_user_id
            )
            
            # Apply standardized GST calculations (EXISTING FUNCTION)
            apply_gst_calculations_to_invoice_line(invoice_line, calculations)
            
            session.add(invoice_line)
            invoice_lines.append(invoice_line)
            
            # Add to totals
            total_amount += Decimal(str(invoice_line.line_total))
            total_cgst += Decimal(str(invoice_line.cgst))
            total_sgst += Decimal(str(invoice_line.sgst))
            total_igst += Decimal(str(invoice_line.igst))
            taxable_amount += Decimal(str(invoice_line.taxable_amount))
            
            logger.debug(f"Line {idx + 1} STORED: Qty={invoice_line.units}, "
                        f"Rate={invoice_line.pack_purchase_price}, "
                        f"Batch={invoice_line.batch_number}, "
                        f"Total={invoice_line.line_total}")
        
        # Update invoice totals
        old_total = invoice.total_amount
        invoice.cgst_amount = total_cgst
        invoice.sgst_amount = total_sgst
        invoice.igst_amount = total_igst
        invoice.total_gst_amount = total_cgst + total_sgst + total_igst
        invoice.total_amount = total_amount
        
        logger.info(f"Updated invoice total amount: {old_total} -> {total_amount}")
        
        session.flush()  # Ensure all entries are saved
        
        # ===================================================================
        # STEP 6: CREATE NEW ACCOUNTING ENTRIES (FOLLOWING CREATE PATTERN)
        # ===================================================================
        
        logger.info("Creating new accounting entries for updated amounts")
        
        # Create new posting reference (following create pattern)
        doc_prefix = 'INV'
        doc_number = invoice.supplier_invoice_number
        timestamp = datetime.now().strftime('%Y%m%d%H%M%S')
        
        # Ensure posting reference fits in VARCHAR(50) database field
        base_reference = f"EDIT-{doc_prefix}-{doc_number}-{timestamp}"
        
        if len(base_reference) > 50:
            # Keep the timestamp and truncate the middle part
            available_for_doc = 50 - len(f"EDIT-{doc_prefix}--{timestamp}")
            if available_for_doc > 0:
                truncated_doc = doc_number[:available_for_doc] if doc_number else ""
                new_posting_reference = f"EDIT-{doc_prefix}-{truncated_doc}-{timestamp}"
            else:
                new_posting_reference = f"EDIT-{doc_prefix}-{timestamp}"
            
            if len(new_posting_reference) > 50:
                new_posting_reference = new_posting_reference[:50]
            logger.warning(f"Posting reference truncated to fit 50 char limit: {new_posting_reference}")
        else:
            new_posting_reference = base_reference
        
        # Create new GL Transaction (following create pattern)
        new_gl_transaction = GLTransaction(
            hospital_id=hospital_id,
            transaction_date=invoice.invoice_date,
            transaction_type='SUPPLIER_INVOICE',
            reference_id=str(invoice_id),
            description=f"Supplier Invoice - {invoice.supplier_invoice_number} (Edited)",
            source_document_type='SUPPLIER_INVOICE',
            source_document_id=invoice_id,
            total_debit=total_amount,
            total_credit=total_amount,
            created_by=current_user_id
        )
        
        session.add(new_gl_transaction)
        session.flush()  # Get transaction ID
        
        # Get account mappings (following create pattern)
        from app.services.posting_config_service import get_posting_config
        config = get_posting_config(str(hospital_id))
        
        ap_account_no = config.get('DEFAULT_AP_ACCOUNT', '2100')
        inventory_account_no = config.get('DEFAULT_INVENTORY_ACCOUNT', '1410')
        cgst_account_no = config.get('CGST_RECEIVABLE_ACCOUNT', '1710')
        sgst_account_no = config.get('SGST_RECEIVABLE_ACCOUNT', '1720')
        igst_account_no = config.get('IGST_RECEIVABLE_ACCOUNT', '1730')
        
        # Get account records and validate they exist
        accounts = {}
        account_nos = [ap_account_no, inventory_account_no, cgst_account_no, sgst_account_no, igst_account_no]
        
        for acc_no in account_nos:
            account = session.query(ChartOfAccounts).filter_by(
                hospital_id=hospital_id,
                gl_account_no=acc_no,
                is_active=True
            ).first()
            if not account:
                raise ValueError(f"Account {acc_no} not found in chart of accounts")
            accounts[acc_no] = account
        
        # Create new GL entries (following create pattern exactly)
        gl_entries = []
        
        # 1. Inventory/Expense Debit
        if taxable_amount > 0:
            inventory_entry = GLEntry(
                hospital_id=hospital_id,
                transaction_id=new_gl_transaction.transaction_id,
                account_id=accounts[inventory_account_no].account_id,
                entry_date=invoice.invoice_date,
                description=f"Inventory - {invoice.supplier_invoice_number} (Edited)",
                debit_amount=taxable_amount,
                credit_amount=Decimal('0'),
                source_document_type='SUPPLIER_INVOICE',
                source_document_id=invoice_id,
                posting_reference=new_posting_reference,
                created_by=current_user_id
            )
            session.add(inventory_entry)
            gl_entries.append(inventory_entry)
            logger.debug(f"Created inventory entry: {inventory_account_no} - Debit: {taxable_amount}")
        
        # 2. CGST Debit
        if total_cgst > 0:
            cgst_entry = GLEntry(
                hospital_id=hospital_id,
                transaction_id=new_gl_transaction.transaction_id,
                account_id=accounts[cgst_account_no].account_id,
                entry_date=invoice.invoice_date,
                description=f"CGST - {invoice.supplier_invoice_number} (Edited)",
                debit_amount=total_cgst,
                credit_amount=Decimal('0'),
                source_document_type='SUPPLIER_INVOICE',
                source_document_id=invoice_id,
                posting_reference=new_posting_reference,
                created_by=current_user_id
            )
            session.add(cgst_entry)
            gl_entries.append(cgst_entry)
            logger.debug(f"Created CGST entry: {cgst_account_no} - Debit: {total_cgst}")
        
        # 3. SGST Debit
        if total_sgst > 0:
            sgst_entry = GLEntry(
                hospital_id=hospital_id,
                transaction_id=new_gl_transaction.transaction_id,
                account_id=accounts[sgst_account_no].account_id,
                entry_date=invoice.invoice_date,
                description=f"SGST - {invoice.supplier_invoice_number} (Edited)",
                debit_amount=total_sgst,
                credit_amount=Decimal('0'),
                source_document_type='SUPPLIER_INVOICE',
                source_document_id=invoice_id,
                posting_reference=new_posting_reference,
                created_by=current_user_id
            )
            session.add(sgst_entry)
            gl_entries.append(sgst_entry)
            logger.debug(f"Created SGST entry: {sgst_account_no} - Debit: {total_sgst}")
        
        # 4. IGST Debit
        if total_igst > 0:
            igst_entry = GLEntry(
                hospital_id=hospital_id,
                transaction_id=new_gl_transaction.transaction_id,
                account_id=accounts[igst_account_no].account_id,
                entry_date=invoice.invoice_date,
                description=f"IGST - {invoice.supplier_invoice_number} (Edited)",
                debit_amount=total_igst,
                credit_amount=Decimal('0'),
                source_document_type='SUPPLIER_INVOICE',
                source_document_id=invoice_id,
                posting_reference=new_posting_reference,
                created_by=current_user_id
            )
            session.add(igst_entry)
            gl_entries.append(igst_entry)
            logger.debug(f"Created IGST entry: {igst_account_no} - Debit: {total_igst}")
        
        # 5. Accounts Payable Credit
        ap_entry = GLEntry(
            hospital_id=hospital_id,
            transaction_id=new_gl_transaction.transaction_id,
            account_id=accounts[ap_account_no].account_id,
            entry_date=invoice.invoice_date,
            description=f"Accounts Payable - {invoice.supplier_invoice_number} (Edited)",
            debit_amount=Decimal('0'),
            credit_amount=total_amount,
            source_document_type='SUPPLIER_INVOICE',
            source_document_id=invoice_id,
            posting_reference=new_posting_reference,
            created_by=current_user_id
        )
        session.add(ap_entry)
        gl_entries.append(ap_entry)
        logger.debug(f"Created AP entry: {ap_account_no} - Credit: {total_amount}")
        
        # Create new AP subledger entry (following create pattern)
        current_supplier_balance = session.query(func.sum(APSubledger.credit_amount) - func.sum(APSubledger.debit_amount)).filter_by(
            hospital_id=hospital_id,
            supplier_id=invoice.supplier_id
        ).scalar() or Decimal('0')
        
        new_balance = current_supplier_balance + total_amount
        
        new_ap_subledger_entry = APSubledger(
            hospital_id=hospital_id,
            branch_id=getattr(invoice, 'branch_id', None),
            transaction_date=invoice.invoice_date,
            entry_type='invoice',
            reference_id=str(invoice_id),
            reference_type='invoice',
            reference_number=invoice.supplier_invoice_number,
            supplier_id=invoice.supplier_id,
            debit_amount=Decimal('0'),
            credit_amount=total_amount,
            current_balance=new_balance,
            gl_transaction_id=new_gl_transaction.transaction_id,
            created_by=current_user_id
        )
        session.add(new_ap_subledger_entry)
        logger.debug(f"Created AP subledger entry: Credit={total_amount}, New Balance={new_balance}")
        
        # ===================================================================
        # STEP 7: CREATE NEW INVENTORY ENTRIES (FOLLOWING CREATE PATTERN)
        # ===================================================================
        
        logger.info("Creating new inventory entries for updated quantities")
        
        # Create stock entries directly (following create pattern with complete fields)
        for line in invoice_lines:
            if line.units > 0:  # Only create stock for positive quantities
                # Get medicine details for additional fields
                medicine = session.query(Medicine).filter_by(
                    medicine_id=line.medicine_id,
                    hospital_id=hospital_id
                ).first()
                
                stock_entry = Inventory(
                    hospital_id=hospital_id,
                    branch_id=getattr(invoice, 'branch_id', None),
                    medicine_id=line.medicine_id,
                    medicine_name=line.medicine_name,
                    medicine_category=medicine.category.name if medicine and medicine.category else 'Unknown',  # FIXED: Populate category
                    batch=line.batch_number or 'BATCH-' + datetime.now().strftime('%Y%m%d'),
                    expiry=line.expiry_date,
                    pack_purchase_price=line.pack_purchase_price,  # FIXED: Populate purchase price
                    pack_mrp=line.pack_mrp,  # FIXED: Populate MRP
                    units_per_pack=line.units_per_pack,  # FIXED: Populate units per pack
                    unit_price=line.unit_price,  # FIXED: Populate unit price
                    sale_price=line.pack_mrp,  # FIXED: Use MRP as sale price (or could be different logic)
                    units=int(line.units),
                    stock_type='Purchase',
                    distributor_invoice_no=invoice.supplier_invoice_number,
                    transaction_date=invoice.invoice_date,
                    reason=f"Stock from edited supplier invoice {invoice.supplier_invoice_number}",
                    current_stock=int(line.units),  # This would need proper calculation in real system
                    created_by=current_user_id
                )
                session.add(stock_entry)
                logger.debug(f"Created stock entry: {line.medicine_name}, Qty: {line.units}, Batch: {line.batch_number or 'DEFAULT'}")
        
        logger.info("New stock entries created successfully")
        
        # ===================================================================
        # STEP 8: UPDATE POSTING FLAGS
        # ===================================================================
        
        # Mark invoice as properly posted with new entries
        invoice.gl_posted = True
        invoice.inventory_posted = True
        invoice.posting_date = datetime.now(timezone.utc)
        invoice.posting_reference = new_posting_reference
        
        session.flush()
        
        # ===================================================================
        # STEP 9: RETURN ENHANCED RESULT
        # ===================================================================
        
        # Verify the update was successful
        verification_lines = session.query(SupplierInvoiceLine).filter_by(invoice_id=invoice_id).all()
        logger.debug(f"Verification: Created {len(verification_lines)} new line items")
        
        # Return updated invoice with comprehensive information
        result = get_entity_dict(invoice)
        result['line_items'] = [get_entity_dict(line) for line in verification_lines]
        result['supplier_name'] = supplier.supplier_name
        
        # Enhanced return data with accounting information
        result.update({
            'edit_completed': True,
            'accounting_properly_updated': True,
            'inventory_properly_updated': True,
            'original_entries_reversed': True,
            'new_entries_created': True,
            'gl_posted': True,
            'gl_transaction_id': str(new_gl_transaction.transaction_id),
            'posting_reference': new_posting_reference,
            'gl_entries_created': len(gl_entries),
            'ap_subledger_created': True,
            'stock_entries_created': True,
            'accounting_summary': {
                'total_amount': float(total_amount),
                'taxable_amount': float(taxable_amount),
                'cgst_amount': float(total_cgst),
                'sgst_amount': float(total_sgst),
                'igst_amount': float(total_igst),
                'is_interstate': is_interstate
            }
        })
        
        logger.info(f"✅ Successfully updated invoice {invoice.supplier_invoice_number} with complete accounting integration")
        logger.info(f"=== ENHANCED SUPPLIER INVOICE EDIT COMPLETE ===")
        
        return result
        
    except Exception as e:
        logger.error(f"Error updating supplier invoice with accounting integration: {str(e)}")
        session.rollback()
        raise

def get_supplier_invoice_by_id(
    invoice_id: uuid.UUID,
    hospital_id: uuid.UUID,
    include_payments: bool = True,
    session: Optional[Session] = None
) -> Dict:
    """
    UNCHANGED SIGNATURE: Get supplier invoice with authoritative calculations
    Same function signature, just improved implementation
    """
    logger.info(f"Fetching supplier invoice {invoice_id} with authoritative calculations")
    
    if session is not None:
        return _get_supplier_invoice_by_id(session, invoice_id, hospital_id, include_payments)
    
    with get_db_session() as new_session:
        return _get_supplier_invoice_by_id(new_session, invoice_id, hospital_id, include_payments)



def _get_supplier_invoice_by_id(
    session: Session,
    invoice_id: uuid.UUID,
    hospital_id: uuid.UUID,
    include_payments: bool = True
) -> Dict:
    """
    FINAL CORRECTED: Get supplier invoice with authoritative calculations
    All field names verified against transaction.py model
    """
    try:
        logger.info(f"Fetching supplier invoice {invoice_id} with authoritative calculations")
        
        # Get invoice entity
        invoice = session.query(SupplierInvoice).filter_by(
            invoice_id=invoice_id,  # ✅ CORRECT: invoice_id is primary key
            hospital_id=hospital_id
        ).first()
        
        if not invoice:
            return None
        
        # Convert to dict
        result = get_entity_dict(invoice)
        
        # CORRECTED: Add line items using correct primary key
        invoice_lines = session.query(SupplierInvoiceLine).filter_by(
            invoice_id=invoice_id
        ).order_by(SupplierInvoiceLine.line_id).all()  # ✅ CORRECT: line_id is primary key
        
        result['line_items'] = [get_entity_dict(line) for line in invoice_lines]
        
        # Add supplier information
        supplier = session.query(Supplier).filter_by(
            supplier_id=invoice.supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if supplier:
            result['supplier_name'] = supplier.supplier_name
        
        # AUTHORITATIVE PAYMENT CALCULATIONS
        if include_payments:
            # Get payment records
            payments = session.query(SupplierPayment).filter_by(
                invoice_id=invoice_id
            ).order_by(SupplierPayment.payment_date.desc()).all()
            
            result['payments'] = [get_entity_dict(payment) for payment in payments]
            
            # APPLY AUTHORITATIVE CALCULATIONS
            auth_totals = calculate_invoice_payment_totals(invoice_id, session)
            result.update(auth_totals)
            
            logger.info(f"Applied authoritative calculations to {invoice.supplier_invoice_number}: "
                       f"Net Payment={auth_totals['net_payment']}, Balance={auth_totals['balance_due']}")
        
        logger.info(f"Successfully retrieved invoice {invoice.supplier_invoice_number} using authoritative calculations")
        return result
        
    except Exception as e:
        logger.error(f"Error getting supplier invoice with authoritative calculations: {str(e)}", exc_info=True)
        raise


# commented on 16.06.25 The code works for credit note feature.  However we are now rationalizing totals Hence new code is brought in.
# def _get_supplier_invoice_by_id(
#     session: Session,
#     invoice_id: uuid.UUID,
#     hospital_id: uuid.UUID,
#     include_payments: bool = True
# ) -> Dict:
#     """
#     Internal function to get a supplier invoice by ID within a session
#     UPDATED: Uses stored GST values only, no recalculation
#     """
#     try:
#         logger.debug(f"Querying supplier invoice {invoice_id}")
        
#         # Get the invoice
#         invoice = session.query(SupplierInvoice).filter_by(
#             invoice_id=invoice_id,
#             hospital_id=hospital_id
#         ).first()
        
#         if not invoice:
#             error_msg = f"Supplier invoice with ID {invoice_id} not found"
#             logger.error(error_msg)
#             raise ValueError(error_msg)
            
#         logger.debug("Fetching invoice line items")
            
#         # Get the line items
#         lines = session.query(SupplierInvoiceLine).filter_by(
#             invoice_id=invoice_id
#         ).all()
        
#         logger.debug(f"Found {len(lines)} line items")
        
#         # Get the supplier
#         supplier = session.query(Supplier).filter_by(
#             supplier_id=invoice.supplier_id
#         ).first()
        
#         if not supplier:
#             logger.warning(f"Supplier {invoice.supplier_id} not found for invoice")
        
#         # Convert to dictionary
#         result = get_entity_dict(invoice)
#         result['line_items'] = [get_entity_dict(line) for line in lines]
        
#         if supplier:
#             result['supplier_name'] = supplier.supplier_name
#             result['supplier_address'] = supplier.supplier_address
#             result['contact_info'] = supplier.contact_info

#         if invoice.branch_id:
#             try:
#                 branch = session.query(Branch).filter_by(
#                     branch_id=invoice.branch_id,
#                     hospital_id=hospital_id
#                 ).first()
                
#                 if branch:
#                     result['branch_name'] = branch.name
#                     logger.debug(f"Added branch_name '{branch.name}' to invoice {invoice.supplier_invoice_number}")
#                 else:
#                     result['branch_name'] = 'Unknown Branch'
#                     logger.warning(f"Branch {invoice.branch_id} not found for invoice {invoice.supplier_invoice_number}")
                    
#             except Exception as branch_error:
#                 logger.warning(f"Error fetching branch name for invoice {invoice.supplier_invoice_number}: {str(branch_error)}")
#                 result['branch_name'] = 'Unknown Branch'
#         else:
#             result['branch_name'] = 'No Branch'
#             logger.debug(f"No branch_id for invoice {invoice.supplier_invoice_number}")

#         # Use stored GST values only (no recalculation in view)
#         subtotal = 0
#         total_gst = 0
        
#         for line in result['line_items']:
#             # Use stored values directly from database
#             line_subtotal = float(line.get('taxable_amount', 0))
#             line_gst = float(line.get('total_gst', 0))
#             line_total = float(line.get('line_total', 0))
            
#             # Add display values using stored data
#             line['subtotal'] = line_subtotal
#             line['gst_amount'] = line_gst
#             line['total_with_gst'] = line_total
            
#             # Update running totals
#             subtotal += line_subtotal
#             total_gst += line_gst
        
#         # Add totals to invoice dictionary using stored values
#         result['calculated_subtotal'] = subtotal
#         result['calculated_total_gst'] = total_gst
#         result['calculated_total_with_gst'] = subtotal + total_gst
        
#         # Log verification of stored vs calculated totals
#         stored_total = float(result.get('total_amount', 0))
#         calculated_total = subtotal + total_gst
        
#         logger.debug(f"Invoice {invoice.supplier_invoice_number}: "
#                     f"Stored total={stored_total}, Calculated total={calculated_total}")
        
#         if abs(stored_total - calculated_total) > 0.01:
#             logger.warning(f"Invoice total mismatch - using stored total as authoritative")
            
#         # Include payments if requested - ENHANCED with payment calculations
#         if include_payments:
#             logger.debug("Fetching payment details")
            
#             payments = session.query(SupplierPayment).filter_by(
#                 invoice_id=invoice_id
#             ).all()
            
#             logger.debug(f"Found {len(payments)} payments")
            
#             result['payments'] = [get_entity_dict(payment) for payment in payments]
            
#             # CRITICAL FIX: Calculate actual payment totals from approved payments
#             payment_total = session.query(func.sum(SupplierPayment.amount)).filter_by(
#                 invoice_id=invoice_id,
#                 workflow_status='approved'  # Only count approved payments
#             ).scalar() or 0
            
#             payment_total = float(payment_total)
#             total_amount = float(result.get('total_amount', 0))
#             balance_due = total_amount - payment_total
            
#             # Ensure balance_due is not negative
#             if balance_due < 0:
#                 balance_due = 0
            
#             # CRITICAL FIX: Update payment calculations and status
#             result['payment_total'] = payment_total
#             result['balance_due'] = balance_due
            
#             # CRITICAL FIX: Recalculate payment status based on actual payments
#             # BUT preserve cancelled status
#             if result.get('payment_status') != 'cancelled':
#                 if payment_total >= total_amount:
#                     result['payment_status'] = 'paid'
#                 elif payment_total > 0:
#                     result['payment_status'] = 'partial'
#                 else:
#                     result['payment_status'] = 'unpaid'
            
#             logger.debug(f"Invoice {result.get('supplier_invoice_number')}: Total={total_amount}, Paid={payment_total}, Balance={balance_due}, Status={result['payment_status']}")
            
#         logger.info(f"Successfully retrieved invoice {invoice.supplier_invoice_number} using stored values")
            
#         return result
        
#     except Exception as e:
#         logger.error(f"Error getting supplier invoice: {str(e)}", exc_info=True)
#         raise

def search_supplier_invoices(
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    invoice_number: Optional[str] = None,
    po_id: Optional[uuid.UUID] = None,
    payment_status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    branch_id: Optional[uuid.UUID] = None,
    current_user_id: Optional[str] = None,
    page: int = 1,
    per_page: int = 20,
    session: Optional[Session] = None
) -> Dict:
    """
    UNCHANGED SIGNATURE: Search invoices with authoritative calculations
    Same function signature, just improved implementation
    """
    logger.info(f"Searching supplier invoices with authoritative calculations")
    
    # Branch determination logic (unchanged)
    if not branch_id and current_user_id:
        accessible_branches = get_user_accessible_branches(current_user_id, hospital_id)
        if len(accessible_branches) == 1:
            branch_id = uuid.UUID(accessible_branches[0]['branch_id'])
    
    if session is not None:
        return _search_supplier_invoices(
            session, hospital_id, supplier_id, invoice_number, po_id, 
            payment_status, start_date, end_date, branch_id, page, per_page
        )
    
    with get_db_session(read_only=True) as new_session:
        return _search_supplier_invoices(
            new_session, hospital_id, supplier_id, invoice_number, po_id,
            payment_status, start_date, end_date, branch_id, page, per_page
        )

def _search_supplier_invoices(
    session: Session,
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    invoice_number: Optional[str] = None,
    po_id: Optional[uuid.UUID] = None,
    payment_status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    branch_id: Optional[uuid.UUID] = None,
    page: int = 1,
    per_page: int = 20
) -> Dict:
    """
    FINAL CORRECTED: Invoice search with authoritative calculations
    All field names verified against transaction.py model
    """
    try:
        logger.debug("Building invoice search query with authoritative calculations")
        
        # Build query
        query = session.query(SupplierInvoice).filter(
            SupplierInvoice.hospital_id == hospital_id
        )
        
        # Apply filters - all field names verified
        if supplier_id:
            query = query.filter(SupplierInvoice.supplier_id == supplier_id)
        if invoice_number:
            query = query.filter(SupplierInvoice.supplier_invoice_number.ilike(f'%{invoice_number}%'))  # ✅ CORRECT
        if po_id:
            query = query.filter(SupplierInvoice.po_id == po_id)
        if payment_status:
            query = query.filter(SupplierInvoice.payment_status == payment_status)  # ✅ CORRECT
        if start_date:
            query = query.filter(SupplierInvoice.invoice_date >= start_date)  # ✅ CORRECT
        if end_date:
            query = query.filter(SupplierInvoice.invoice_date <= end_date)  # ✅ CORRECT
        if branch_id:
            query = query.filter(SupplierInvoice.branch_id == branch_id)  # ✅ CORRECT
        
        query = query.order_by(desc(SupplierInvoice.invoice_date))  # ✅ CORRECT
        
        # Get total and paginated results
        total = query.count()
        offset = (page - 1) * per_page
        invoices = query.offset(offset).limit(per_page).all()
        
        logger.debug(f"Retrieved {len(invoices)} invoices for authoritative calculation")
        
        # APPLY AUTHORITATIVE CALCULATIONS TO ALL INVOICES
        invoice_list = []
        for invoice in invoices:
            invoice_dict = get_entity_dict(invoice)
            
            # Add supplier name
            supplier = session.query(Supplier).filter_by(
                supplier_id=invoice.supplier_id  # ✅ CORRECT
            ).first()
            invoice_dict['supplier_name'] = supplier.supplier_name if supplier else 'Unknown Supplier'
            
            # APPLY AUTHORITATIVE CALCULATION
            auth_totals = calculate_invoice_payment_totals(invoice.invoice_id, session)  # ✅ CORRECT: invoice_id
            invoice_dict.update(auth_totals)
            
            invoice_list.append(invoice_dict)
        
        logger.debug(f"Applied authoritative calculations to {len(invoice_list)} invoices")
        
        return {
            'invoices': invoice_list,
            'total': total,
            'page': page,
            'per_page': per_page,
            'pages': (total + per_page - 1) // per_page
        }
        
    except Exception as e:
        logger.error(f"Error in authoritative invoice search: {str(e)}")
        raise


# commented on 16.06.25 The code works for credit note feature.  However we are now rationalizing totals Hence new code is brought in.
# def _search_supplier_invoices(
#     session: Session,
#     hospital_id: uuid.UUID,
#     supplier_id: Optional[uuid.UUID] = None,
#     invoice_number: Optional[str] = None,
#     po_id: Optional[uuid.UUID] = None,
#     payment_status: Optional[str] = None,
#     start_date: Optional[datetime] = None,
#     end_date: Optional[datetime] = None,
#     branch_id: Optional[uuid.UUID] = None,
#     page: int = 1,
#     per_page: int = 20
# ) -> Dict:
#     """
#     Internal invoice search function - FIXED payment total calculation
#     """
#     try:
#         logger.debug("Building invoice search query")
        
#         # Base query
#         query = session.query(SupplierInvoice).filter(
#             SupplierInvoice.hospital_id == hospital_id
#         )
        
#         # Apply filters
#         if supplier_id:
#             query = query.filter(SupplierInvoice.supplier_id == supplier_id)
            
#         if invoice_number:
#             query = query.filter(SupplierInvoice.supplier_invoice_number.ilike(f'%{invoice_number}%'))
            
#         if po_id:
#             query = query.filter(SupplierInvoice.po_id == po_id)
            
#         if payment_status:
#             query = query.filter(SupplierInvoice.payment_status == payment_status)
            
#         if start_date:
#             query = query.filter(SupplierInvoice.invoice_date >= start_date)
            
#         if end_date:
#             query = query.filter(SupplierInvoice.invoice_date <= end_date)
            
#         if branch_id:
#             query = query.filter(SupplierInvoice.branch_id == branch_id)
        
#         # Order by invoice date (newest first)
#         query = query.order_by(desc(SupplierInvoice.invoice_date))
        
#         # Get total count
#         total = query.count()
#         logger.debug(f"Total invoices matching criteria: {total}")
        
#         # Apply pagination
#         offset = (page - 1) * per_page
#         invoices = query.offset(offset).limit(per_page).all()
        
#         logger.debug(f"Retrieved {len(invoices)} invoices for current page")
        
#         # CRITICAL FIX: Calculate payment totals correctly for each invoice
#         invoice_list = []
#         for invoice in invoices:
#             invoice_dict = get_entity_dict(invoice)
            
#             # Add supplier name if available
#             supplier = session.query(Supplier).filter_by(
#                 supplier_id=invoice.supplier_id
#             ).first()
#             if supplier:
#                 invoice_dict['supplier_name'] = supplier.supplier_name
#             else:
#                 invoice_dict['supplier_name'] = 'Unknown Supplier'
            
#             # FIXED: Separate positive payments and credit adjustments
#             positive_payments = session.query(func.sum(SupplierPayment.amount)).filter_by(
#                 invoice_id=invoice.invoice_id,
#                 workflow_status='approved'
#             ).filter(SupplierPayment.amount > 0).scalar() or 0
            
#             negative_payments = session.query(func.sum(SupplierPayment.amount)).filter_by(
#                 invoice_id=invoice.invoice_id,
#                 workflow_status='approved'
#             ).filter(SupplierPayment.amount < 0).scalar() or 0
            
#             # Calculate net effective payment
#             credit_adjustments = abs(float(negative_payments))
#             positive_payments = float(positive_payments)
#             effective_payment = positive_payments - credit_adjustments
            
#             total_amount = float(invoice.total_amount or 0)
#             balance_due = total_amount - effective_payment
            
#             # Ensure balance_due is not negative
#             if balance_due < 0:
#                 balance_due = 0
            
#             # FIXED: Update payment calculations with NET values
#             invoice_dict['payment_total'] = effective_payment  # NET payment after credits
#             invoice_dict['balance_due'] = balance_due
#             invoice_dict['positive_payments_total'] = positive_payments  # Gross payments
#             invoice_dict['credit_adjustments_total'] = credit_adjustments  # Credit notes
            
#             # FIXED: Recalculate payment status based on NET effective payments
#             if invoice_dict.get('payment_status') != 'cancelled':
#                 if effective_payment >= total_amount:
#                     invoice_dict['payment_status'] = 'paid'
#                 elif effective_payment > 0:
#                     invoice_dict['payment_status'] = 'partial'
#                 else:
#                     invoice_dict['payment_status'] = 'unpaid'
            
#             # Add credit note indicators
#             invoice_dict['has_credit_notes'] = credit_adjustments > 0
            
#             logger.debug(f"Invoice {invoice.supplier_invoice_number}: "
#                         f"Total={total_amount}, "
#                         f"Positive Payments={positive_payments}, "
#                         f"Credits={credit_adjustments}, "
#                         f"NET Effective={effective_payment}, "
#                         f"Balance={balance_due}, "
#                         f"Status={invoice_dict['payment_status']}")
            
#             invoice_list.append(invoice_dict)
        
#         logger.info(f"Search completed: {len(invoice_list)} invoices returned with correct payment calculations")
        
#         return {
#             'invoices': invoice_list,
#             'total': total,
#             'page': page,
#             'per_page': per_page,
#             'has_more': total > page * per_page,
#             'pagination': {
#                 'page': page,
#                 'per_page': per_page,
#                 'total_count': total,
#                 'total_pages': (total + per_page - 1) // per_page
#             }
#         }
        
#     except Exception as e:
#         logger.error(f"Error searching supplier invoices: {str(e)}", exc_info=True)
#         return {
#             'invoices': [],
#             'total': 0,
#             'page': page,
#             'per_page': per_page,
#             'has_more': False,
#             'pagination': {
#                 'page': page,
#                 'per_page': per_page,
#                 'total_count': 0,
#                 'total_pages': 0
#             }
#         }

def cancel_supplier_invoice(
    invoice_id: uuid.UUID,
    hospital_id: uuid.UUID,
    cancellation_reason: str,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Cancel a supplier invoice - EXISTING FUNCTION (unchanged)
    """
    logger.info(f"Cancelling supplier invoice {invoice_id} for hospital {hospital_id}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _cancel_supplier_invoice(session, invoice_id, hospital_id, cancellation_reason, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _cancel_supplier_invoice(new_session, invoice_id, hospital_id, cancellation_reason, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing supplier invoice cancellation: {invoice_id}")
        new_session.commit()
        
        logger.info(f"Successfully cancelled supplier invoice")
        return result

def _cancel_supplier_invoice(session, invoice_id, hospital_id, cancellation_reason, current_user_id):
    """
    CORRECTED VERSION: Cancel supplier invoice with proper accounting reversal
    
    This function creates actual reversal entries instead of just marking as reversed.
    Considers all tables: GLEntry, GLTransaction, APSubledger, Inventory, SupplierInvoice, SupplierPayment
    
    Args:
        session: Database session
        invoice_id: UUID of invoice to cancel
        hospital_id: UUID of hospital
        cancellation_reason: Reason for cancellation
        current_user_id: User performing cancellation
    
    Returns:
        Dict with cancellation details
    """
    from app.models.transaction import (
        SupplierInvoice, SupplierInvoiceLine, SupplierPayment, 
        GLTransaction, GLEntry, APSubledger, Inventory, PaymentDocument
    )
    from decimal import Decimal
    from sqlalchemy import func
    
    try:
        # ===================================================================
        # STEP 1: VALIDATE INVOICE AND CHECK CONSTRAINTS
        # ===================================================================
        
        invoice = session.query(SupplierInvoice).filter_by(
            invoice_id=invoice_id,
            hospital_id=hospital_id
        ).first()
        
        if not invoice:
            raise ValueError(f"Invoice {invoice_id} not found")
        
        if invoice.is_reversed:
            raise ValueError("Invoice is already cancelled/reversed")
            
        if invoice.payment_status == 'cancelled':
            raise ValueError("Invoice is already cancelled")

        logger.info(f"[INFO] DEBUG: Invoice details:")
        logger.info(f"   - Invoice ID: {invoice_id}")
        logger.info(f"   - Invoice Number: {invoice.supplier_invoice_number}")
        logger.info(f"   - Hospital ID: {hospital_id}")
        logger.info(f"   - GL Posted: {invoice.gl_posted}")
        logger.info(f"   - Inventory Posted: {invoice.inventory_posted}")
        logger.info(f"   - Current Status: {invoice.payment_status}")
        
        # Check for approved payments - prevent cancellation if payments exist
        approved_payments = session.query(SupplierPayment).filter_by(
            invoice_id=invoice_id,
            workflow_status='approved'
        ).all()
        
        if approved_payments:
            raise ValueError(
                f"Cannot cancel invoice with approved payments. "
                f"Found {len(approved_payments)} approved payment(s). "
                f"Please reverse payments first."
            )

        logger.info(f"Starting PROPER accounting reversal for invoice {invoice.supplier_invoice_number}")
        
        # Create reversal reference and timestamp
        reversal_reference = f"CANCEL-INV-{datetime.now().strftime('%Y%m%d%H%M%S')}-{str(invoice_id)[:8]}"
        reversal_timestamp = datetime.now(timezone.utc)
        
        # ===================================================================
        # STEP 2: REVERSE GL TRANSACTIONS AND ENTRIES (PROPER ACCOUNTING)
        # ===================================================================
        
        # Find original GL transaction for this invoice
        original_gl_transaction = session.query(GLTransaction).filter_by(
            reference_id=str(invoice_id),
            transaction_type='SUPPLIER_INVOICE',
            hospital_id=hospital_id,
            is_reversal=False  # Only get original transaction, not previous reversals
        ).first()
        
        if original_gl_transaction:
            logger.info(f"Found original GL transaction: {original_gl_transaction.transaction_id}")
            
            # Create reversal GL transaction
            reversal_gl_transaction = GLTransaction(
                hospital_id=hospital_id,
                transaction_date=reversal_timestamp,
                transaction_type='SUPPLIER_INVOICE_REVERSAL',
                reference_id=str(invoice_id),
                description=f"REVERSAL: {original_gl_transaction.description}",
                source_document_type='SUPPLIER_INVOICE_CANCELLATION',
                source_document_id=invoice_id,
                total_debit=original_gl_transaction.total_credit,  # Swap amounts
                total_credit=original_gl_transaction.total_debit,  # Swap amounts
                is_reversal=True,  # Mark as reversal
                original_entry_id=original_gl_transaction.transaction_id,  # Link to original
                created_by=current_user_id
            )
            
            session.add(reversal_gl_transaction)
            session.flush()  # Get transaction_id
            
            logger.info(f"Created reversal GL transaction: {reversal_gl_transaction.transaction_id}")
            
            # Get original GL entries
            original_gl_entries = session.query(GLEntry).filter_by(
                transaction_id=original_gl_transaction.transaction_id,
                is_reversal=False  # Only original entries
            ).all()
            
            logger.info(f"Found {len(original_gl_entries)} original GL entries to reverse")
            
            # Create reversal GL entries with OPPOSITE amounts
            gl_reversal_count = 0
            for original_entry in original_gl_entries:
                reversal_entry = GLEntry(
                    hospital_id=hospital_id,
                    transaction_id=reversal_gl_transaction.transaction_id,
                    account_id=original_entry.account_id,  # Same account
                    entry_date=reversal_timestamp,
                    description=f"REVERSAL: {original_entry.description}",
                    debit_amount=original_entry.credit_amount,  # SWAP: Credit becomes Debit
                    credit_amount=original_entry.debit_amount,  # SWAP: Debit becomes Credit
                    source_document_type='SUPPLIER_INVOICE_CANCELLATION',
                    source_document_id=invoice_id,
                    posting_reference=reversal_reference[:100],  # Truncate to fit column limit
                    is_reversal=True,  # Mark as reversal
                    original_entry_id=original_entry.entry_id,  # Link to original entry
                    created_by=current_user_id
                )
                
                session.add(reversal_entry)
                gl_reversal_count += 1
                
                logger.debug(f"Created GL reversal entry - Account: {original_entry.account_id}, "
                           f"Original Dr:{original_entry.debit_amount} Cr:{original_entry.credit_amount}, "
                           f"Reversal Dr:{reversal_entry.debit_amount} Cr:{reversal_entry.credit_amount}")
            
            logger.info(f"✅ RESULT: Created {gl_reversal_count} GL reversal entries")
            
        else:
            logger.warning(f"❌ No original GL transaction found for invoice {invoice_id}")

        # ===================================================================
        # STEP 3: REVERSE AP SUBLEDGER ENTRIES (PROPER ACCOUNTING)
        # ===================================================================
        
        # Find original AP entry
        original_ap_entry = session.query(APSubledger).filter_by(
            reference_id=str(invoice_id),
            reference_type='invoice',
            hospital_id=hospital_id
        ).first()
        
        if original_ap_entry:
            logger.info(f"Found original AP entry: {original_ap_entry.entry_id}")
            
            # Create reversal AP entry with OPPOSITE amounts
            reversal_ap_entry = APSubledger(
                hospital_id=hospital_id,
                branch_id=original_ap_entry.branch_id,
                transaction_date=reversal_timestamp,
                entry_type='reversal',  # Mark as reversal
                reference_id=str(invoice_id),
                reference_type='invoice_reversal',
                reference_number=f"REV-{original_ap_entry.reference_number}",
                supplier_id=original_ap_entry.supplier_id,
                debit_amount=original_ap_entry.credit_amount,  # SWAP: Credit becomes Debit
                credit_amount=original_ap_entry.debit_amount,  # SWAP: Debit becomes Credit
                current_balance=None,  # Will be calculated below
                gl_transaction_id=reversal_gl_transaction.transaction_id if 'reversal_gl_transaction' in locals() else None,
                created_by=current_user_id
            )
            
            # Calculate new supplier balance after reversal
            current_supplier_balance = session.query(func.coalesce(func.sum(
                APSubledger.credit_amount - APSubledger.debit_amount
            ), 0)).filter_by(
                hospital_id=hospital_id,
                supplier_id=original_ap_entry.supplier_id
            ).scalar() or Decimal('0')
            
            # Update balance to reflect the reversal
            new_balance = current_supplier_balance - original_ap_entry.credit_amount
            reversal_ap_entry.current_balance = new_balance
            
            session.add(reversal_ap_entry)
            
            logger.info(f"✅ RESULT: Created AP reversal entry, new supplier balance: {new_balance}")
            
        else:
            logger.warning(f"❌ No original AP entry found for invoice {invoice_id}")


        # ===================================================================
        # STEP 4: REVERSE INVENTORY/STOCK ENTRIES IF POSTED (SAFE VERSION)
        # ===================================================================

        if invoice.inventory_posted:
            logger.info("Reversing original inventory entries")
            
            # ✅ SAFE: Find ONLY inventory entries for THIS SPECIFIC INVOICE
            stock_entries = session.query(Inventory).filter_by(
                hospital_id=hospital_id,
                distributor_invoice_no=invoice.supplier_invoice_number,
                stock_type='Purchase'
            ).all()
            
            logger.info(f"Found {len(stock_entries)} stock entries for invoice {invoice.supplier_invoice_number}")
            
            if stock_entries:
                reversal_timestamp = datetime.now(timezone.utc)
                
                # Log what we're reversing for audit trail
                for i, stock_entry in enumerate(stock_entries, 1):
                    logger.info(f"Cancelling stock entry {i}: Medicine={stock_entry.medicine_name}, "
                            f"Batch={stock_entry.batch}, Qty={stock_entry.units}")
                
                for stock_entry in stock_entries:
                    # Create reversal stock entry with negative quantities
                    reversal_stock_entry = Inventory(
                        hospital_id=hospital_id,
                        branch_id=getattr(stock_entry, 'branch_id', None) or getattr(invoice, 'branch_id', None),
                        medicine_id=stock_entry.medicine_id,
                        medicine_name=getattr(stock_entry, 'medicine_name', 'Unknown Medicine'),
                        medicine_category=getattr(stock_entry, 'medicine_category', 'Unknown'),
                        batch=stock_entry.batch,
                        expiry=stock_entry.expiry,
                        pack_purchase_price=getattr(stock_entry, 'pack_purchase_price', 0),
                        pack_mrp=getattr(stock_entry, 'pack_mrp', 0),
                        units_per_pack=getattr(stock_entry, 'units_per_pack', 1),
                        unit_price=getattr(stock_entry, 'unit_price', 0),
                        sale_price=getattr(stock_entry, 'sale_price', 0),
                        units=-int(stock_entry.units),  # Negative quantity for reversal
                        percent_discount=getattr(stock_entry, 'percent_discount', 0),
                        cgst=getattr(stock_entry, 'cgst', 0),
                        sgst=getattr(stock_entry, 'sgst', 0),
                        igst=getattr(stock_entry, 'igst', 0),
                        total_gst=getattr(stock_entry, 'total_gst', 0),
                        stock_type='Purchase_Reversal',  # Mark as reversal
                        distributor_invoice_no=f"REV-{stock_entry.distributor_invoice_no or invoice.supplier_invoice_number}",
                        transaction_date=reversal_timestamp,
                        reason=f"Reversal of stock from cancelled invoice {invoice.supplier_invoice_number}",
                        current_stock=stock_entry.current_stock - int(stock_entry.units),
                        created_by=current_user_id
                    )
                    session.add(reversal_stock_entry)
                
                logger.info(f"✅ Created {len(stock_entries)} stock reversal entries")
            else:
                logger.warning(f"No stock entries found for invoice {invoice.supplier_invoice_number}")
                # ✅ NO DANGEROUS FALLBACKS - If no entries found, that's it!

        # ===================================================================
        # STEP 5: CHECK AND PREVENT PAYMENT REVERSALS (SAFETY CHECK)
        # ===================================================================
        
        # Additional check for any payments and their documents
        all_payments = session.query(SupplierPayment).filter_by(invoice_id=invoice_id).all()
        
        if all_payments:
            logger.warning(f"Found {len(all_payments)} payment records associated with invoice")
            
            # Check if any are approved
            approved_payments = [p for p in all_payments if p.workflow_status == 'approved']
            if approved_payments:
                raise ValueError(
                    f"Cannot cancel invoice. Found {len(approved_payments)} approved payments. "
                    f"Please reverse payments first before cancelling invoice."
                )
            
            # For draft/pending payments, we could optionally cancel them
            draft_payments = [p for p in all_payments if p.workflow_status in ['draft', 'pending_approval']]
            if draft_payments:
                logger.info(f"Found {len(draft_payments)} draft/pending payments that will be cancelled")
                for payment in draft_payments:
                    payment.workflow_status = 'cancelled'
                    payment.rejection_reason = f"Invoice {invoice.supplier_invoice_number} was cancelled"
                    payment.rejected_at = reversal_timestamp
                    payment.rejected_by = current_user_id

        # ===================================================================
        # STEP 6: UPDATE INVOICE WITH PROPER REVERSAL FLAGS
        # ===================================================================
        
        # Update invoice with proper reversal tracking
        invoice.is_reversed = True  # Use proper boolean flag
        invoice.payment_status = 'cancelled'
        invoice.reversal_reason = cancellation_reason[:200]  # Truncate to field limit
        invoice.reversed_at = reversal_timestamp
        invoice.reversed_by = current_user_id
        invoice.reversed_by_invoice_id = invoice_id  # Self-reference for cancellation
        
        # Update audit fields
        invoice.updated_at = reversal_timestamp
        invoice.updated_by = current_user_id
        
        # Reset posting flags to indicate reversal
        invoice.gl_posted = False
        invoice.inventory_posted = False
        
        # Update posting reference with proper truncation for 50-character limit
        original_ref = invoice.posting_reference or 'N/A'
        reversal_ref = f"REVERSED-{original_ref}"
        
        if len(reversal_ref) > 50:
            # Keep 'REVERSED-' prefix and truncate the original reference
            available_chars = 50 - len("REVERSED-")  # 41 chars available
            truncated_original = original_ref[:available_chars]
            invoice.posting_reference = f"REVERSED-{truncated_original}"
            logger.warning(f"Posting reference truncated to fit 50 char limit: {invoice.posting_reference}")
        else:
            invoice.posting_reference = reversal_ref

        # ===================================================================
        # STEP 7: HANDLE PAYMENT DOCUMENTS (MARK AS CANCELLED)
        # ===================================================================
        
        # Find and handle any payment documents
        payment_documents = session.query(PaymentDocument).join(SupplierPayment).filter(
            SupplierPayment.invoice_id == invoice_id
        ).all()
        
        if payment_documents:
            logger.info(f"Found {len(payment_documents)} payment documents to mark as cancelled")
            for doc in payment_documents:
                # Don't delete documents, just mark them as relating to cancelled invoice
                doc.notes = f"Related invoice {invoice.supplier_invoice_number} was cancelled on {reversal_timestamp.date()}"

        # ===================================================================
        # STEP 8: COMMIT AND RETURN RESULT
        # ===================================================================
        
        session.flush()  # Ensure all changes are saved
        
        logger.info(f"✅ Successfully cancelled invoice {invoice.supplier_invoice_number} with proper accounting reversal")
        logger.info(f"Reversal reference: {reversal_reference}")
        
        # Prepare result with comprehensive reversal information
        result = get_entity_dict(invoice)
        result.update({
            'reversal_reference': reversal_reference,
            'reversal_completed': True,
            'reversal_timestamp': reversal_timestamp.isoformat(),
            'accounting_properly_reversed': True,
            'gl_transactions_reversed': 1 if original_gl_transaction else 0,
            'gl_entries_reversed': gl_reversal_count if 'gl_reversal_count' in locals() else 0,
            'ap_entries_reversed': 1 if original_ap_entry else 0,
            'stock_entries_reversed': len(stock_entries) if 'stock_entries' in locals() else 0,
            'payments_cancelled': len(draft_payments) if 'draft_payments' in locals() else 0
        })
        
        return result
        
    except Exception as e:
        logger.error(f"Error cancelling invoice with proper reversal: {str(e)}")
        session.rollback()
        raise

# ===================================================================
# Supplier Payment Management - COMPLETE (already have branch support)
# ===================================================================

def record_supplier_payment(
    hospital_id: uuid.UUID,
    payment_data: Dict,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    SIMPLIFIED: Record supplier payment with mandatory accounting integration
    Removed create_gl_entries parameter - accounting is always created
    """
    logger.info(f"Recording supplier payment for hospital {hospital_id}")
    
    if session is not None:
        return _record_supplier_payment(session, hospital_id, payment_data, current_user_id)
    
    with get_db_session() as new_session:
        result = _record_supplier_payment(new_session, hospital_id, payment_data, current_user_id)
        new_session.commit()
        logger.info(f"Successfully recorded payment: {result.get('payment_id')}")
        return result



# ===================================================================
# REDESIGNED: _record_supplier_payment METHOD
# ===================================================================

def _record_supplier_payment(
    session: Session,
    hospital_id: uuid.UUID,
    payment_data: Dict,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    REDESIGNED: Record supplier payment with integrated accounting entries
    All operations in single transaction - payment + GL entries + AP subledger
    """
    try:
        logger.info(f"Recording supplier payment with integrated accounting for hospital {hospital_id}")
        
        # ===================================================================
        # STEP 1: VALIDATE AND CREATE PAYMENT (EXISTING LOGIC)
        # ===================================================================
        
        # Validate payment data
        validation_result = validate_payment_data(
            payment_data, hospital_id, current_user_id, session=session
        )
        if not validation_result['is_valid']:
            raise ValueError(f"Payment validation failed: {'; '.join(validation_result['errors'])}")
        
        # Validate supplier
        supplier_id = payment_data.get('supplier_id')
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            raise ValueError(f"Supplier with ID {supplier_id} not found")
        
        if supplier.black_listed:
            raise ValueError(f"Supplier '{supplier.supplier_name}' is blacklisted")
        
        # Validate invoice if provided
        invoice = None
        if payment_data.get('invoice_id'):
            invoice = session.query(SupplierInvoice).filter_by(
                invoice_id=uuid.UUID(payment_data['invoice_id']),
                hospital_id=hospital_id
            ).first()
            
            if not invoice:
                raise ValueError("Invoice not found")
            
            if invoice.payment_status == 'cancelled':
                raise ValueError("Cannot make payment to cancelled invoice")
        
        # Create payment record
        payment = SupplierPayment(
            hospital_id=hospital_id,
            branch_id=uuid.UUID(payment_data['branch_id']),
            supplier_id=supplier_id,
            invoice_id=uuid.UUID(payment_data['invoice_id']) if payment_data.get('invoice_id') else None,
            payment_date=payment_data.get('payment_date', datetime.now(timezone.utc)),
            payment_method=payment_data.get('payment_method', 'cash'),
            amount=Decimal(str(payment_data.get('amount', 0))),
            reference_no=payment_data.get('reference_no'),
            notes=payment_data.get('notes'),
            cash_amount=Decimal(str(payment_data.get('cash_amount', 0))),
            cheque_amount=Decimal(str(payment_data.get('cheque_amount', 0))),
            bank_transfer_amount=Decimal(str(payment_data.get('bank_transfer_amount', 0))),
            upi_amount=Decimal(str(payment_data.get('upi_amount', 0))),
            cheque_number=payment_data.get('cheque_number'),
            cheque_date=payment_data.get('cheque_date'),
            cheque_bank=payment_data.get('cheque_bank'),
            bank_account_name=payment_data.get('bank_account_name'),
            bank_reference_number=payment_data.get('bank_reference_number'),
            upi_transaction_id=payment_data.get('upi_transaction_id'),
            upi_app_name=payment_data.get('upi_app_name'),
            workflow_status='pending',
            created_by=current_user_id,
            created_at=datetime.now(timezone.utc)
        )
        
        # Auto-approve if amount is below threshold
        from app.config import PAYMENT_CONFIG
        auto_approve_limit = PAYMENT_CONFIG.get('AUTO_APPROVE_LIMIT', 5000)
        if float(payment.amount) <= auto_approve_limit:
            payment.workflow_status = 'approved'
            payment.approved_by = current_user_id
            payment.approved_at = datetime.now(timezone.utc)
        
        session.add(payment)
        session.flush()  # Get payment_id
        
        # Only create GL entries for approved payments
        if payment.workflow_status == 'approved':
            
            # ===================================================================
            # STEP 2: CREATE GL TRANSACTION AND ENTRIES (INTEGRATED)
            # ===================================================================
            
            logger.info(f"Creating GL entries for payment {payment.reference_no}")
            
            # Generate posting reference
            posting_reference = f"PAY-{payment.reference_no or str(payment.payment_id)[:8]}-{datetime.now().strftime('%Y%m%d%H%M%S')}"
            if len(posting_reference) > 50:
                posting_reference = posting_reference[:50]
            
            # Create GL Transaction
            gl_transaction = GLTransaction(
                hospital_id=hospital_id,
                transaction_date=payment.payment_date,
                transaction_type='SUPPLIER_PAYMENT',
                reference_id=str(payment.payment_id),
                description=f"Supplier Payment - {payment.reference_no}",
                source_document_type='SUPPLIER_PAYMENT',
                source_document_id=payment.payment_id,
                total_debit=payment.amount,
                total_credit=payment.amount,
                created_by=current_user_id
            )
            
            session.add(gl_transaction)
            session.flush()  # Get transaction_id
            
            # Get account mappings
            from app.services.posting_config_service import get_posting_config
            config = get_posting_config(str(hospital_id))
            
            ap_account_no = config.get('DEFAULT_AP_ACCOUNT', '2100')
            
            # Determine bank account based on payment method
            if payment.cash_amount > 0:
                bank_account_no = config.get('DEFAULT_CASH_ACCOUNT', '1101')
            else:
                bank_account_no = config.get('DEFAULT_BANK_ACCOUNT', '1200')
            
            # Get account records
            ap_account = session.query(ChartOfAccounts).filter_by(
                hospital_id=hospital_id,
                gl_account_no=ap_account_no,
                is_active=True
            ).first()
            
            bank_account = session.query(ChartOfAccounts).filter_by(
                hospital_id=hospital_id,
                gl_account_no=bank_account_no,
                is_active=True
            ).first()
            
            if not ap_account:
                raise ValueError(f"AP Account {ap_account_no} not found in chart of accounts")
            if not bank_account:
                raise ValueError(f"Bank Account {bank_account_no} not found in chart of accounts")
            
            gl_entries = []
            
            # 1. AP Debit (reduce liability)
            ap_entry = GLEntry(
                hospital_id=hospital_id,
                transaction_id=gl_transaction.transaction_id,
                account_id=ap_account.account_id,
                entry_date=payment.payment_date,
                description=f"AP Payment - {payment.reference_no}",
                debit_amount=payment.amount,
                credit_amount=Decimal('0'),
                source_document_type='SUPPLIER_PAYMENT',
                source_document_id=payment.payment_id,
                posting_reference=posting_reference,
                created_by=current_user_id
            )
            session.add(ap_entry)
            gl_entries.append(ap_entry)
            
            # 2. Bank Credit (reduce asset)
            bank_entry = GLEntry(
                hospital_id=hospital_id,
                transaction_id=gl_transaction.transaction_id,
                account_id=bank_account.account_id,
                entry_date=payment.payment_date,
                description=f"Payment - {payment.reference_no}",
                debit_amount=Decimal('0'),
                credit_amount=payment.amount,
                source_document_type='SUPPLIER_PAYMENT',
                source_document_id=payment.payment_id,
                posting_reference=posting_reference,
                created_by=current_user_id
            )
            session.add(bank_entry)
            gl_entries.append(bank_entry)
            
            # ===================================================================
            # STEP 3: CREATE AP SUBLEDGER ENTRY (INTEGRATED)
            # ===================================================================
            
            logger.info(f"Creating AP subledger entry for payment {payment.reference_no}")
            
            # Calculate current AP balance for this supplier
            current_balance = session.query(func.coalesce(func.sum(
                APSubledger.credit_amount - APSubledger.debit_amount
            ), 0)).filter_by(
                hospital_id=hospital_id,
                supplier_id=supplier_id
            ).scalar() or Decimal('0')
            
            new_balance = current_balance - payment.amount
            
            ap_subledger_entry = APSubledger(
                hospital_id=hospital_id,
                branch_id=payment.branch_id,
                transaction_date=payment.payment_date,
                entry_type='payment',
                reference_id=payment.payment_id,
                reference_type='payment',
                reference_number=payment.reference_no or str(payment.payment_id)[:8],
                supplier_id=supplier_id,
                debit_amount=payment.amount,
                credit_amount=Decimal('0'),
                current_balance=new_balance,
                gl_transaction_id=gl_transaction.transaction_id,
                created_by=current_user_id
            )
            
            session.add(ap_subledger_entry)
            
            # ===================================================================
            # STEP 4: MARK PAYMENT AS POSTED
            # ===================================================================
            
            payment.gl_posted = True
            payment.posting_date = datetime.now(timezone.utc)
            payment.posting_reference = posting_reference
            
            gl_transaction_id = str(gl_transaction.transaction_id)
            gl_entries_created = len(gl_entries)
            ap_subledger_created = True
            
        else:
            # Payment pending approval - no GL entries yet
            gl_transaction_id = None
            gl_entries_created = 0
            ap_subledger_created = False
        
        # ===================================================================
        # STEP 5: UPDATE INVOICE STATUS (IF APPLICABLE)
        # ===================================================================
        
        if invoice and payment.workflow_status == 'approved':
            # Calculate total approved payments for this invoice
            total_payments = session.query(func.sum(SupplierPayment.amount)).filter_by(
                invoice_id=invoice.invoice_id,
                workflow_status='approved'
            ).scalar() or Decimal('0')
            
            if total_payments >= invoice.total_amount:
                invoice.payment_status = 'paid'
            elif total_payments > 0:
                invoice.payment_status = 'partial'
            else:
                invoice.payment_status = 'unpaid'
            
            invoice.updated_at = datetime.now(timezone.utc)
            invoice.updated_by = current_user_id
        
        session.flush()  # Ensure everything is saved
        
        # ===================================================================
        # NEW: ADD DOCUMENT PROCESSING (MISSING FROM REDESIGN)
        # ===================================================================
        
        # Handle document uploads if provided
        if payment_data.get('documents'):
            try:
                logger.info(f"Processing {len(payment_data['documents'])} documents for payment {payment.reference_no}")
                _process_payment_documents(
                    payment_id=payment.payment_id,
                    documents=payment_data['documents'],
                    current_user_id=current_user_id,
                    session=session
                )
                logger.info(f"Document processing completed for payment {payment.reference_no}")
            except Exception as e:
                logger.warning(f"Document processing failed for payment {payment.reference_no}: {str(e)}")
                # Don't fail the entire payment for document errors


        # ===================================================================
        # STEP 6: RETURN RESULT
        # ===================================================================
        
        result = {
            'payment_id': str(payment.payment_id),
            'amount': float(payment.amount),
            'status': payment.workflow_status,
            'supplier_name': supplier.supplier_name,
            'invoice_number': invoice.supplier_invoice_number if invoice else None,
            'reference_no': payment.reference_no,
            'payment_method': payment.payment_method,
            'requires_approval': payment.workflow_status == 'pending',
            'gl_posted': payment.gl_posted,
            'gl_transaction_id': gl_transaction_id,
            'posting_reference': payment.posting_reference,
            'gl_entries_created': gl_entries_created,
            'ap_subledger_created': ap_subledger_created
        }
        
        logger.info(f"Successfully recorded payment {payment.reference_no} with complete accounting integration")
        
        return result
        
    except Exception as e:
        logger.error(f"Error recording supplier payment with accounting: {str(e)}")
        session.rollback()
        raise




def get_supplier_payment_history(
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    branch_id: Optional[uuid.UUID] = None,  # NEW: Optional branch filter
    current_user_id: Optional[str] = None,  # NEW: For auto-branch detection
    session: Optional[Session] = None
) -> List[Dict]:
    """
    Get payment history for a supplier - UPDATED to use branch_service
    """
    logger.info(f"Getting payment history for supplier {supplier_id}, branch {branch_id}")
    logger.debug(f"Date range: {start_date} to {end_date}")
    
    # NEW: Use branch service for branch determination if not specified
    if not branch_id and current_user_id:
        accessible_branches = get_user_accessible_branches(current_user_id, hospital_id)
        if len(accessible_branches) == 1:
            branch_id = uuid.UUID(accessible_branches[0]['branch_id'])
    
    if session is not None:
        return _get_supplier_payment_history(session, supplier_id, hospital_id, start_date, end_date, branch_id)
    
    with get_db_session() as new_session:
        return _get_supplier_payment_history(new_session, supplier_id, hospital_id, start_date, end_date, branch_id)

def _get_supplier_payment_history(
    session: Session,
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    branch_id: Optional[uuid.UUID] = None  # NEW: Added parameter
) -> List[Dict]:
    """
    Internal function to get payment history for a supplier within a session - UPDATED with branch filtering
    """
    try:
        # Validate supplier
        logger.debug(f"Validating supplier {supplier_id}")
        
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug("Building payment history query with branch support")
            
        # Query payments
        query = session.query(SupplierPayment).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        )
        
        # NEW: Apply branch filter if specified
        if branch_id:
            logger.debug(f"Filtering by branch: {branch_id}")
            query = query.filter(SupplierPayment.branch_id == branch_id)
        
        # Apply date filters
        if start_date:
            logger.debug(f"Filtering by start date: {start_date}")
            query = query.filter(SupplierPayment.payment_date >= start_date)
            
        if end_date:
            logger.debug(f"Filtering by end date: {end_date}")
            query = query.filter(SupplierPayment.payment_date <= end_date)
            
        # Order by date
        query = query.order_by(desc(SupplierPayment.payment_date))
        
        # Execute query
        payments = query.all()
        logger.debug(f"Found {len(payments)} payments")
        
        # Convert to dictionaries with invoice details
        results = []
        
        for payment in payments:
            payment_dict = get_entity_dict(payment)
            
            # Add invoice details if available
            if payment.invoice_id:
                logger.debug(f"Fetching invoice details for payment {payment.payment_id}")
                
                invoice = session.query(SupplierInvoice).filter_by(invoice_id=payment.invoice_id).first()
                if invoice:
                    payment_dict['invoice_number'] = invoice.supplier_invoice_number
                    payment_dict['invoice_date'] = invoice.invoice_date
                    payment_dict['invoice_amount'] = float(invoice.total_amount)
                    
            results.append(payment_dict)
            
        logger.info(f"Retrieved {len(results)} payment records")
            
        return results
        
    except Exception as e:
        logger.error(f"Error getting supplier payment history: {str(e)}", exc_info=True)
        raise

def get_pending_supplier_invoices(
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    branch_id: Optional[uuid.UUID] = None,  # NEW: Optional branch filter
    current_user_id: Optional[str] = None,  # NEW: For auto-branch detection
    session: Optional[Session] = None
) -> List[Dict]:
    """
    Get unpaid or partially paid supplier invoices - UPDATED to use branch_service
    """
    logger.info(f"Getting pending invoices for hospital {hospital_id}, branch {branch_id}")
    if supplier_id:
        logger.debug(f"Filtering by supplier: {supplier_id}")
    
    # NEW: Use branch service for branch determination if not specified
    if not branch_id and current_user_id:
        accessible_branches = get_user_accessible_branches(current_user_id, hospital_id)
        if len(accessible_branches) == 1:
            branch_id = uuid.UUID(accessible_branches[0]['branch_id'])
    
    if session is not None:
        return _get_pending_supplier_invoices(session, hospital_id, supplier_id, branch_id)
    
    with get_db_session() as new_session:
        return _get_pending_supplier_invoices(new_session, hospital_id, supplier_id, branch_id)

def _get_pending_supplier_invoices(
    session: Session,
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    branch_id: Optional[uuid.UUID] = None  # NEW: Added parameter
) -> List[Dict]:
    """
    Internal function to get pending supplier invoices within a session - UPDATED with branch filtering
    """
    try:
        logger.debug("Building pending invoices query with branch support")
        
        # Base query for unpaid or partially paid invoices
        query = session.query(SupplierInvoice).filter(
            SupplierInvoice.hospital_id == hospital_id,
            SupplierInvoice.payment_status.in_(['unpaid', 'partial'])
        )
        
        # NEW: Apply branch filter if specified
        if branch_id:
            logger.debug(f"Filtering by branch: {branch_id}")
            query = query.filter(SupplierInvoice.branch_id == branch_id)
        
        # Apply supplier filter if provided
        if supplier_id:
            logger.debug(f"Filtering by supplier: {supplier_id}")
            query = query.filter(SupplierInvoice.supplier_id == supplier_id)
            
        # Order by date (oldest first)
        query = query.order_by(SupplierInvoice.invoice_date)
        
        # Execute query
        invoices = query.all()
        logger.debug(f"Found {len(invoices)} pending invoices")
        
        # Convert to dictionaries with payment details
        results = []
        
        for invoice in invoices:
            invoice_dict = get_entity_dict(invoice)
            
            # Get supplier information
            logger.debug(f"Getting supplier info for invoice {invoice.invoice_id}")
            
            supplier = session.query(Supplier).filter_by(supplier_id=invoice.supplier_id).first()
            if supplier:
                invoice_dict['supplier_name'] = supplier.supplier_name
                
            # Get payment information
            logger.debug("Calculating payment totals")
            
            payment_total = session.query(func.sum(SupplierPayment.amount)).filter_by(
                invoice_id=invoice.invoice_id
            ).scalar() or 0
            
            payment_total = Decimal(payment_total)
            balance_due = invoice.total_amount - payment_total
            
            invoice_dict['payment_total'] = float(payment_total)
            invoice_dict['balance_due'] = float(balance_due)
            
            # Get payment history
            logger.debug("Getting payment history")
            
            payments = session.query(SupplierPayment).filter_by(
                invoice_id=invoice.invoice_id
            ).order_by(SupplierPayment.payment_date).all()
            
            invoice_dict['payments'] = [get_entity_dict(payment) for payment in payments]
            
            results.append(invoice_dict)
            
        logger.info(f"Retrieved {len(results)} pending invoices")
            
        return results
        
    except Exception as e:
        logger.error(f"Error getting pending supplier invoices: {str(e)}", exc_info=True)
        raise

# ===================================================================
# Utility Functions - COMPLETE WITH BRANCH SUPPORT
# ===================================================================

def get_supplier_statistics(
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    branch_id: Optional[uuid.UUID] = None,  # NEW: Optional branch filter
    current_user_id: Optional[str] = None,  # NEW: For auto-branch detection
    session: Optional[Session] = None
) -> Dict:
    """
    Get statistics for a supplier - UPDATED to use branch_service
    """
    logger.info(f"Getting statistics for supplier {supplier_id}, branch {branch_id}")
    
    # NEW: Use branch service for branch determination if not specified
    if not branch_id and current_user_id:
        accessible_branches = get_user_accessible_branches(current_user_id, hospital_id)
        if len(accessible_branches) == 1:
            branch_id = uuid.UUID(accessible_branches[0]['branch_id'])
    
    if session is not None:
        return _get_supplier_statistics(session, supplier_id, hospital_id, branch_id)
    
    with get_db_session() as new_session:
        return _get_supplier_statistics(new_session, supplier_id, hospital_id, branch_id)

def _get_supplier_statistics(
    session: Session,
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    branch_id: Optional[uuid.UUID] = None  # NEW: Added parameter
) -> Dict:
    """
    Internal function to get supplier statistics within a session - UPDATED with branch filtering
    """
    try:
        # Validate supplier
        logger.debug(f"Validating supplier {supplier_id}")
        
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug("Calculating supplier statistics with branch filtering")
        
        # Calculate statistics with optional branch filtering
        stats = {}
        
        # Base queries for branch filtering
        po_query = session.query(PurchaseOrderHeader).filter_by(supplier_id=supplier_id)
        invoice_query = session.query(SupplierInvoice).filter_by(supplier_id=supplier_id)
        payment_query = session.query(SupplierPayment).filter_by(supplier_id=supplier_id)
        
        # NEW: Apply branch filter if specified
        if branch_id:
            logger.debug(f"Applying branch filter: {branch_id}")
            po_query = po_query.filter_by(branch_id=branch_id)
            invoice_query = invoice_query.filter_by(branch_id=branch_id)
            payment_query = payment_query.filter_by(branch_id=branch_id)
        
        # Total purchase orders
        stats['total_purchase_orders'] = po_query.count()
        
        # Total invoices
        stats['total_invoices'] = invoice_query.count()
        
        # Total payments
        stats['total_payments'] = payment_query.count()
        
        # Total business volume
        total_business = invoice_query.with_entities(func.sum(SupplierInvoice.total_amount)).scalar() or 0
        stats['total_business_volume'] = float(total_business)
        
        # Outstanding balance
        total_invoiced = invoice_query.with_entities(func.sum(SupplierInvoice.total_amount)).scalar() or 0
        total_paid = payment_query.with_entities(func.sum(SupplierPayment.amount)).scalar() or 0
        
        stats['outstanding_balance'] = float(Decimal(total_invoiced) - Decimal(total_paid))
        
        # Average payment time (branch-filtered)
        paid_invoices = invoice_query.filter(SupplierInvoice.payment_status == 'paid').all()
        
        if paid_invoices:
            payment_times = []
            for invoice in paid_invoices:
                last_payment = payment_query.filter_by(
                    invoice_id=invoice.invoice_id
                ).order_by(desc(SupplierPayment.payment_date)).first()
                
                if last_payment:
                    payment_time = (last_payment.payment_date - invoice.invoice_date).days
                    payment_times.append(payment_time)
                    
            stats['average_payment_days'] = sum(payment_times) / len(payment_times) if payment_times else 0
        else:
            stats['average_payment_days'] = 0
            
        # Performance metrics
        stats['supplier_name'] = supplier.supplier_name
        stats['performance_rating'] = supplier.performance_rating
        stats['is_blacklisted'] = supplier.black_listed
        stats['status'] = supplier.status
        stats['branch_filtered'] = branch_id is not None  # NEW: Indicate if branch-filtered
        
        # Recent activity (branch-filtered)
        stats['last_po_date'] = None
        last_po = po_query.order_by(desc(PurchaseOrderHeader.po_date)).first()
        if last_po:
            stats['last_po_date'] = last_po.po_date
            
        stats['last_invoice_date'] = None
        last_invoice = invoice_query.order_by(desc(SupplierInvoice.invoice_date)).first()
        if last_invoice:
            stats['last_invoice_date'] = last_invoice.invoice_date
            
        stats['last_payment_date'] = None
        last_payment = payment_query.order_by(desc(SupplierPayment.payment_date)).first()
        if last_payment:
            stats['last_payment_date'] = last_payment.payment_date
            
        logger.info(f"Retrieved statistics for supplier {supplier.supplier_name}")
        
        return stats
        
    except Exception as e:
        logger.error(f"Error getting supplier statistics: {str(e)}", exc_info=True)
        raise

def export_supplier_list(
    hospital_id: uuid.UUID,
    filters: Optional[Dict] = None,
    format: str = 'csv',
    branch_id: Optional[uuid.UUID] = None,  # NEW: Optional branch filter
    current_user_id: Optional[str] = None,  # NEW: For auto-branch detection
    session: Optional[Session] = None
) -> Union[str, bytes]:
    """
    Export supplier list to file - UPDATED to use branch_service
    """
    logger.info(f"Exporting supplier list for hospital {hospital_id}, branch {branch_id} in {format} format")
    
    # NEW: Use branch service for branch determination if not specified
    if not branch_id and current_user_id:
        accessible_branches = get_user_accessible_branches(current_user_id, hospital_id)
        if len(accessible_branches) == 1:
            branch_id = uuid.UUID(accessible_branches[0]['branch_id'])
    
    if session is not None:
        return _export_supplier_list(session, hospital_id, filters, format, branch_id)
    
    with get_db_session() as new_session:
        return _export_supplier_list(new_session, hospital_id, filters, format, branch_id)

def _export_supplier_list(
    session: Session,
    hospital_id: uuid.UUID,
    filters: Optional[Dict] = None,
    format: str = 'csv',
    branch_id: Optional[uuid.UUID] = None  # NEW: Added parameter
) -> Union[str, bytes]:
    """
    Internal function to export supplier list within a session - UPDATED with branch filtering
    """
    try:
        logger.debug("Building supplier query for export with branch support")
        
        # Base query
        query = session.query(Supplier).filter_by(hospital_id=hospital_id)
        
        # NEW: Add branch filter if specified
        if branch_id:
            logger.debug(f"Filtering export by branch: {branch_id}")
            query = query.filter(Supplier.branch_id == branch_id)
        
        # Apply existing filters if provided
        if filters:
            if filters.get('name'):
                query = query.filter(Supplier.supplier_name.ilike(f"%{filters['name']}%"))
                
            if filters.get('category'):
                query = query.filter(Supplier.supplier_category == filters['category'])
                
            if filters.get('status'):
                query = query.filter(Supplier.status == filters['status'])
                
            if filters.get('blacklisted') is not None:
                query = query.filter(Supplier.black_listed == filters['blacklisted'])
                
        # Order by name
        query = query.order_by(Supplier.supplier_name)
        
        # Get all suppliers
        suppliers = query.all()
        logger.debug(f"Exporting {len(suppliers)} suppliers")
        
        # Prepare data for export with branch information
        data = []
        headers = [
            'Supplier ID', 'Supplier Name', 'Category', 'Status',
            'Branch Name',  # NEW: Added branch column
            'Contact Person', 'Phone', 'Email', 'GSTIN',
            'State Code', 'Payment Terms', 'Performance Rating',
            'Black Listed', 'Created Date'
        ]
        
        # Get branch names for lookup
        branch_names = {}
        if suppliers:
            branch_ids = [s.branch_id for s in suppliers if s.branch_id]
            if branch_ids:
                branches = session.query(Branch).filter(Branch.branch_id.in_(branch_ids)).all()
                branch_names = {str(b.branch_id): b.name for b in branches}
        
        for supplier in suppliers:
            row = [
                str(supplier.supplier_id),
                supplier.supplier_name,
                supplier.supplier_category or '',
                supplier.status,
                branch_names.get(str(supplier.branch_id), 'Unknown Branch'),  # NEW: Branch name
                supplier.contact_person_name or '',
                supplier.contact_info.get('phone', '') if supplier.contact_info else '',
                supplier.email or '',
                supplier.gst_registration_number or '',
                supplier.state_code or '',
                supplier.payment_terms or '',
                str(supplier.performance_rating) if supplier.performance_rating else '',
                'Yes' if supplier.black_listed else 'No',
                supplier.created_at.strftime('%Y-%m-%d %H:%M:%S') if supplier.created_at else ''
            ]
            data.append(row)
            
        # Export based on format (existing logic unchanged)
        if format == 'csv':
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(data)
            
            logger.info("CSV export completed")
            return output.getvalue()
            
        elif format == 'excel':
            import io
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Suppliers"
            
            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info("Excel export completed")
            return output.getvalue()
            
        else:
            error_msg = f"Unsupported export format: {format}"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
    except Exception as e:
        logger.error(f"Error exporting supplier list: {str(e)}", exc_info=True)
        raise

def delete_supplier(
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Delete a supplier (soft delete by marking as inactive) - UPDATED to use branch_service
    """
    logger.info(f"Deleting supplier {supplier_id} for hospital {hospital_id}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _delete_supplier(session, supplier_id, hospital_id, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _delete_supplier(new_session, supplier_id, hospital_id, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing supplier deletion: {supplier_id}")
        new_session.commit()
        
        logger.info(f"Successfully deleted supplier")
        return result

def _delete_supplier(
    session: Session,
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to delete a supplier within a session - UPDATED with branch validation
    """
    try:
        logger.debug(f"Fetching supplier {supplier_id}")
        
        # Get the supplier
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # NEW: Validate branch access using branch_service
        if current_user_id:
            if not validate_branch_access(current_user_id, hospital_id, supplier.branch_id):
                error_msg = f"Access denied - supplier belongs to different branch"
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        logger.debug("Checking for active transactions")
        
        # Check if supplier has active purchase orders
        active_pos = session.query(PurchaseOrderHeader).filter_by(
            supplier_id=supplier_id,
            status='approved'
        ).count()
        
        if active_pos > 0:
            error_msg = f"Cannot delete supplier with {active_pos} active purchase orders"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Check if supplier has unpaid invoices
        unpaid_invoices = session.query(SupplierInvoice).filter(
            SupplierInvoice.supplier_id == supplier_id,
            SupplierInvoice.payment_status.in_(['unpaid', 'partial'])
        ).count()
        
        if unpaid_invoices > 0:
            error_msg = f"Cannot delete supplier with {unpaid_invoices} unpaid invoices"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.debug(f"Marking supplier {supplier_id} as inactive")
        
        # Soft delete by marking as inactive
        supplier.status = 'inactive'
        supplier.deleted_flag = True
        
        if current_user_id:
            supplier.updated_by = current_user_id
            supplier.deleted_by = current_user_id
            
        supplier.deleted_at = datetime.now(timezone.utc)
        
        session.flush()
        
        logger.info(f"Supplier {supplier.supplier_name} deleted successfully")
        
        return {
            'supplier_id': str(supplier_id),
            'supplier_name': supplier.supplier_name,
            'branch_id': str(supplier.branch_id),
            'status': 'deleted',
            'deleted_at': supplier.deleted_at
        }
        
    except Exception as e:
        logger.error(f"Error deleting supplier: {str(e)}", exc_info=True)
        session.rollback()
        raise

# ===================================================================
# GST Calculation Functions - COMPLETE (unchanged)
# ===================================================================

def calculate_gst_values(
    quantity: float,
    unit_rate: float,
    gst_rate: float,
    discount_percent: float = 0,
    is_free_item: bool = False,
    is_interstate: bool = False,
    conversion_factor: float = 1.0
) -> dict:
    """
    Single source of truth for all GST calculations in the system - EXISTING FUNCTION (unchanged)
    """
    try:
        # Convert to Decimal for precision in financial calculations
        from decimal import Decimal, ROUND_HALF_UP
        
        quantity_dec = Decimal(str(quantity))
        unit_rate_dec = Decimal(str(unit_rate))
        gst_rate_dec = Decimal(str(gst_rate))
        discount_percent_dec = Decimal(str(discount_percent))
        conversion_factor_dec = Decimal(str(conversion_factor))
        
        # Initialize all values
        base_amount = Decimal('0')
        discount_amount = Decimal('0')
        discounted_rate = Decimal('0')
        taxable_amount = Decimal('0')
        cgst_rate = Decimal('0')
        sgst_rate = Decimal('0')
        igst_rate = Decimal('0')
        cgst_amount = Decimal('0')
        sgst_amount = Decimal('0')
        igst_amount = Decimal('0')
        total_gst_amount = Decimal('0')
        line_total = Decimal('0')
        sub_unit_price = Decimal('0')
        
        # CRITICAL FIX: Handle free items correctly - preserve quantity and conversion factor
        if is_free_item:
            # For free items: financial values are zero, but quantities are preserved
            sub_unit_price = Decimal('0')  # Free items have zero cost per sub-unit
            
            return {
                # Input values (PRESERVED for free items)
                'quantity': float(quantity_dec),                    # KEEP actual quantity
                'unit_rate': 0.0,                                   # Rate is zero for free items
                'discount_percent': 0.0,                           # No discount on free items
                'conversion_factor': float(conversion_factor_dec),  # KEEP conversion factor
                'is_free_item': is_free_item,
                'is_interstate': is_interstate,
                
                # Calculated amounts (all zero for free items)
                'base_amount': 0.0,
                'discount_amount': 0.0,
                'discounted_rate': 0.0,
                'taxable_amount': 0.0,
                
                # GST rates and amounts (all zero for free items)
                'gst_rate': float(gst_rate_dec),
                'cgst_rate': 0.0,
                'sgst_rate': 0.0,
                'igst_rate': 0.0,
                'cgst_amount': 0.0,
                'sgst_amount': 0.0,
                'igst_amount': 0.0,
                'total_gst_amount': 0.0,
                
                # Final amounts (zero cost but preserve inventory data)
                'line_total': 0.0,
                'sub_unit_price': 0.0,
                'total_sub_units': float(quantity_dec * conversion_factor_dec),  # PRESERVE for inventory
                
                # Database field mappings for free items
                'db_mappings': {
                    'units': float(quantity_dec),              # CRITICAL: Preserve actual quantity
                    'pack_purchase_price': 0.0,                # Zero rate for free items
                    'discount_percent': 0.0,
                    'discount_amount': 0.0,
                    'taxable_amount': 0.0,
                    'units_per_pack': float(conversion_factor_dec),  # PRESERVE conversion factor
                    'unit_price': 0.0,
                    'gst_rate': float(gst_rate_dec),
                    'cgst_rate': 0.0,
                    'sgst_rate': 0.0,
                    'igst_rate': 0.0,
                    'cgst': 0.0,
                    'sgst': 0.0,
                    'igst': 0.0,
                    'total_gst': 0.0,
                    'line_total': 0.0,
                    'is_free_item': True
                }
            }
        
        # Step 1: Calculate base amount (quantity × unit rate) - for non-free items
        base_amount = quantity_dec * unit_rate_dec
        
        # Step 2: Apply discount to get discounted rate
        if discount_percent_dec > 0:
            discount_amount = (base_amount * discount_percent_dec / 100).quantize(Decimal('0.01'), ROUND_HALF_UP)
            discounted_rate = (unit_rate_dec * (100 - discount_percent_dec) / 100).quantize(Decimal('0.01'), ROUND_HALF_UP)
        else:
            discount_amount = Decimal('0')
            discounted_rate = unit_rate_dec
        
        # Step 3: Calculate taxable amount (quantity × discounted rate)
        taxable_amount = (quantity_dec * discounted_rate).quantize(Decimal('0.01'), ROUND_HALF_UP)
        
        # Step 4: Determine GST structure based on interstate/intrastate
        if is_interstate:
            # Interstate transaction: Use IGST
            igst_rate = gst_rate_dec
            cgst_rate = Decimal('0')
            sgst_rate = Decimal('0')
        else:
            # Intrastate transaction: Split into CGST + SGST
            cgst_rate = gst_rate_dec / 2
            sgst_rate = gst_rate_dec / 2
            igst_rate = Decimal('0')
        
        # Step 5: Calculate GST amounts on taxable amount
        cgst_amount = (taxable_amount * cgst_rate / 100).quantize(Decimal('0.01'), ROUND_HALF_UP)
        sgst_amount = (taxable_amount * sgst_rate / 100).quantize(Decimal('0.01'), ROUND_HALF_UP)
        igst_amount = (taxable_amount * igst_rate / 100).quantize(Decimal('0.01'), ROUND_HALF_UP)
        total_gst_amount = cgst_amount + sgst_amount + igst_amount
        
        # Step 6: Calculate final line total
        line_total = taxable_amount + total_gst_amount
        
        # Step 7: Calculate sub-unit price for inventory
        if conversion_factor_dec > 0:
            sub_unit_price = discounted_rate / conversion_factor_dec
        else:
            sub_unit_price = discounted_rate
        
        # Return comprehensive calculation results for non-free items
        return {
            # Input values (for reference and validation)
            'quantity': float(quantity_dec),
            'unit_rate': float(unit_rate_dec),
            'discount_percent': float(discount_percent_dec),
            'conversion_factor': float(conversion_factor_dec),
            'is_free_item': is_free_item,
            'is_interstate': is_interstate,
            
            # Base calculations
            'base_amount': float(base_amount),
            'discount_amount': float(discount_amount),
            'discounted_rate': float(discounted_rate),
            'taxable_amount': float(taxable_amount),
            
            # GST rates
            'gst_rate': float(gst_rate_dec),
            'cgst_rate': float(cgst_rate),
            'sgst_rate': float(sgst_rate),
            'igst_rate': float(igst_rate),
            
            # GST amounts
            'cgst_amount': float(cgst_amount),
            'sgst_amount': float(sgst_amount),
            'igst_amount': float(igst_amount),
            'total_gst_amount': float(total_gst_amount),
            
            # Final amounts
            'line_total': float(line_total),
            'sub_unit_price': float(sub_unit_price),
            'total_sub_units': float(quantity_dec * conversion_factor_dec),
            
            # Database field mappings
            'db_mappings': {
                'units': float(quantity_dec),
                'pack_purchase_price': float(unit_rate_dec),
                'discount_percent': float(discount_percent_dec),
                'discount_amount': float(discount_amount),
                'taxable_amount': float(taxable_amount),
                'units_per_pack': float(conversion_factor_dec),
                'unit_price': float(sub_unit_price),
                'gst_rate': float(gst_rate_dec),
                'cgst_rate': float(cgst_rate),
                'sgst_rate': float(sgst_rate),
                'igst_rate': float(igst_rate),
                'cgst': float(cgst_amount),
                'sgst': float(sgst_amount),
                'igst': float(igst_amount),
                'total_gst': float(total_gst_amount),
                'line_total': float(line_total),
                'is_free_item': is_free_item
            }
        }
        
    except (ValueError, TypeError, Exception) as e:
        logger.error(f"Error in calculate_gst_values: {str(e)}")
        # Return safe defaults with preserved quantity for free items
        return {
            'quantity': float(quantity) if quantity else 0.0,
            'unit_rate': 0.0 if is_free_item else float(unit_rate) if unit_rate else 0.0,
            'discount_percent': 0.0,
            'conversion_factor': float(conversion_factor) if conversion_factor else 1.0,
            'is_free_item': is_free_item,
            'is_interstate': is_interstate,
            'base_amount': 0.0,
            'discount_amount': 0.0,
            'discounted_rate': 0.0,
            'taxable_amount': 0.0,
            'gst_rate': 0.0,
            'cgst_rate': 0.0,
            'sgst_rate': 0.0,
            'igst_rate': 0.0,
            'cgst_amount': 0.0,
            'sgst_amount': 0.0,
            'igst_amount': 0.0,
            'total_gst_amount': 0.0,
            'line_total': 0.0,
            'sub_unit_price': 0.0,
            'total_sub_units': float(quantity * conversion_factor) if quantity and conversion_factor else 0.0,
            'db_mappings': {
                'units': float(quantity) if quantity else 0.0,
                'pack_purchase_price': 0.0 if is_free_item else 0.0,
                'discount_percent': 0.0,
                'discount_amount': 0.0,
                'taxable_amount': 0.0,
                'units_per_pack': float(conversion_factor) if conversion_factor else 1.0,
                'unit_price': 0.0,
                'gst_rate': 0.0,
                'cgst_rate': 0.0,
                'sgst_rate': 0.0,
                'igst_rate': 0.0,
                'cgst': 0.0,
                'sgst': 0.0,
                'igst': 0.0,
                'total_gst': 0.0,
                'line_total': 0.0,
                'is_free_item': is_free_item
            }
        }

def apply_gst_calculations_to_line(line_obj, calculations):
    """
    Helper function to apply GST calculations with detailed logging - EXISTING FUNCTION (unchanged)
    """
    try:
        db_mappings = calculations.get('db_mappings', {})
        
        # Log what we're applying
        medicine_name = getattr(line_obj, 'medicine_name', 'Unknown')
        logger.debug(f"Applying calculations to line: {medicine_name}")
        
        # Apply calculations using original field names
        line_obj.units = db_mappings.get('units', 0)
        line_obj.pack_purchase_price = db_mappings.get('pack_purchase_price', 0)
        line_obj.discount_percent = db_mappings.get('discount_percent', 0)
        line_obj.discount_amount = db_mappings.get('discount_amount', 0)
        line_obj.taxable_amount = db_mappings.get('taxable_amount', 0)
        line_obj.units_per_pack = db_mappings.get('units_per_pack', 1)
        line_obj.unit_price = db_mappings.get('unit_price', 0)
        line_obj.gst_rate = db_mappings.get('gst_rate', 0)
        line_obj.cgst_rate = db_mappings.get('cgst_rate', 0)
        line_obj.sgst_rate = db_mappings.get('sgst_rate', 0)
        line_obj.igst_rate = db_mappings.get('igst_rate', 0)
        line_obj.cgst = db_mappings.get('cgst', 0)
        line_obj.sgst = db_mappings.get('sgst', 0)
        line_obj.igst = db_mappings.get('igst', 0)
        line_obj.total_gst = db_mappings.get('total_gst', 0)
        line_obj.line_total = db_mappings.get('line_total', 0)
        
        # Set free item flag if the model supports it
        if hasattr(line_obj, 'is_free_item'):
            line_obj.is_free_item = db_mappings.get('is_free_item', False)
        
        # Additional fields for some models
        # if hasattr(line_obj, 'unit_rate'):
        #     line_obj.unit_rate = db_mappings.get('pack_purchase_price', 0)
        # if hasattr(line_obj, 'discounted_rate'):
        #     line_obj.discounted_rate = db_mappings.get('pack_purchase_price', 0) * (1 - db_mappings.get('discount_percent', 0) / 100)
        
        logger.debug(f"Applied: Units={line_obj.units}, Rate={line_obj.pack_purchase_price}, "
                    f"Taxable={line_obj.taxable_amount}, GST={line_obj.total_gst}, "
                    f"Total={line_obj.line_total}, Free={getattr(line_obj, 'is_free_item', 'N/A')}")
        
    except Exception as e:
        logger.error(f"Error applying GST calculations to line: {str(e)}")
        raise

def apply_gst_calculations_to_invoice_line(line_obj, calculations):
    """
    Helper function to apply GST calculations to invoice lines - EXISTING FUNCTION (unchanged)
    """
    try:
        db_mappings = calculations.get('db_mappings', {})
        
        # Log what we're applying for invoice lines
        medicine_name = getattr(line_obj, 'medicine_name', 'Unknown')
        logger.debug(f"Applying invoice calculations to line: {medicine_name}")
        
        # Apply calculations using invoice line field names
        line_obj.units = db_mappings.get('units', 0)
        line_obj.pack_purchase_price = db_mappings.get('pack_purchase_price', 0)
        line_obj.discount_percent = db_mappings.get('discount_percent', 0)
        line_obj.discount_amount = db_mappings.get('discount_amount', 0)
        line_obj.taxable_amount = db_mappings.get('taxable_amount', 0)
        line_obj.units_per_pack = db_mappings.get('units_per_pack', 1)
        line_obj.unit_price = db_mappings.get('unit_price', 0)
        line_obj.gst_rate = db_mappings.get('gst_rate', 0)
        line_obj.cgst_rate = db_mappings.get('cgst_rate', 0)
        line_obj.sgst_rate = db_mappings.get('sgst_rate', 0)
        line_obj.igst_rate = db_mappings.get('igst_rate', 0)
        line_obj.cgst = db_mappings.get('cgst', 0)
        line_obj.sgst = db_mappings.get('sgst', 0)
        line_obj.igst = db_mappings.get('igst', 0)
        line_obj.total_gst = db_mappings.get('total_gst', 0)
        line_obj.line_total = db_mappings.get('line_total', 0)
        
        # Set free item flag if the model supports it
        if hasattr(line_obj, 'is_free_item'):
            line_obj.is_free_item = db_mappings.get('is_free_item', False)
        
        logger.debug(f"Applied invoice: Units={line_obj.units}, Rate={line_obj.pack_purchase_price}, "
                    f"Taxable={line_obj.taxable_amount}, GST={line_obj.total_gst}, "
                    f"Total={line_obj.line_total}, Free={getattr(line_obj, 'is_free_item', 'N/A')}")
        
    except Exception as e:
        logger.error(f"Error applying GST calculations to invoice line: {str(e)}")
        raise

def validate_purchase_order_data(po_data: Dict) -> List[str]:
    """
    Comprehensive validation for purchase order data - EXISTING FUNCTION (unchanged)
    """
    errors = []
    
    # Basic header validation
    if not po_data.get('supplier_id'):
        errors.append("Supplier ID is required")
    
    if not po_data.get('po_date'):
        errors.append("Purchase order date is required")
    
    line_items = po_data.get('line_items', [])
    if not line_items:
        errors.append("At least one line item is required")
    
    # Validate line items
    for idx, item_data in enumerate(line_items):
        line_errors = validate_po_line_item(item_data)
        for error in line_errors:
            errors.append(f"Line {idx + 1}: {error}")
    
    return errors

# ============================================================================
# GST CALCULATION HELPER METHODS
# ============================================================================

def get_transaction_state_codes(hospital_id, supplier_id=None):
    """
    Get hospital and supplier state codes for GST calculation
    
    Args:
        hospital_id: UUID of the hospital
        supplier_id: UUID of the supplier (optional)
        
    Returns:
        dict: {
            'hospital_state_code': str,
            'supplier_state_code': str,
            'is_interstate': bool
        }
    """
    from app.services.database_service import get_db_session
    from app.models.master import Hospital, Supplier
    import uuid
    
    result = {
        'hospital_state_code': None,
        'supplier_state_code': None,
        'is_interstate': False
    }
    
    try:
        with get_db_session(read_only=True) as session:
            # Get hospital state code
            hospital = session.query(Hospital).filter_by(
                hospital_id=hospital_id
            ).first()
            
            if hospital:
                result['hospital_state_code'] = hospital.state_code
                logger.info(f"Hospital state code: {result['hospital_state_code']}")
            
            # Get supplier state code if supplier_id provided
            if supplier_id:
                supplier_uuid = uuid.UUID(supplier_id) if isinstance(supplier_id, str) else supplier_id
                supplier = session.query(Supplier).filter_by(
                    supplier_id=supplier_uuid,
                    hospital_id=hospital_id,
                    is_deleted=False
                ).first()
                
                if supplier:
                    result['supplier_state_code'] = supplier.state_code
                    logger.info(f"Supplier state code: {result['supplier_state_code']}")
            
            # Determine if interstate
            if result['hospital_state_code'] and result['supplier_state_code']:
                result['is_interstate'] = (result['hospital_state_code'] != result['supplier_state_code'])
                logger.info(f"Interstate transaction: {result['is_interstate']}")
        
        return result
        
    except Exception as e:
        logger.error(f"Error getting state codes: {str(e)}")
        return result


def resolve_medicine_identifier(medicine_identifier, hospital_id):
    """
    Resolve medicine identifier to UUID (handles both UUID string and medicine name)
    
    Args:
        medicine_identifier: Medicine UUID string or medicine name
        hospital_id: Hospital UUID for filtering
        
    Returns:
        UUID: Resolved medicine UUID
        
    Raises:
        ValueError: If medicine not found
    """
    from app.models.master import Medicine
    from app.services.database_service import get_db_session
    import uuid
    
    if not medicine_identifier or medicine_identifier in ['undefined', 'null', '']:
        raise ValueError("Invalid medicine identifier")
    
    try:
        # First try to parse as UUID
        try:
            return uuid.UUID(medicine_identifier)
        except ValueError:
            # Not a UUID - search by name
            logger.info(f"Medicine identifier is not UUID, searching by name: {medicine_identifier}")
            
            with get_db_session(read_only=True) as session:
                medicine = session.query(Medicine).filter_by(
                    medicine_name=medicine_identifier,
                    hospital_id=hospital_id,
                    is_deleted=False
                ).first()
                
                if medicine:
                    logger.info(f"Found medicine by name '{medicine_identifier}', UUID: {medicine.medicine_id}")
                    return medicine.medicine_id
                else:
                    raise ValueError(f"Medicine not found: {medicine_identifier}")
    
    except Exception as e:
        logger.error(f"Error resolving medicine identifier: {str(e)}")
        raise ValueError(f"Cannot resolve medicine: {medicine_identifier}")


def calculate_line_item_with_gst(line_item_data, is_interstate=False):
    """
    Calculate GST for a single line item using centralized calculation
    
    Args:
        line_item_data: Dictionary with line item details (units, pack_purchase_price, etc.)
        is_interstate: Boolean indicating interstate transaction
        
    Returns:
        dict: Line item with calculated GST values
    """
    try:
        # Call existing centralized GST calculation
        gst_calculations = calculate_gst_values(
            quantity=line_item_data.get('units', 0),
            unit_rate=line_item_data.get('pack_purchase_price', 0),
            gst_rate=line_item_data.get('gst_rate', 0),
            discount_percent=line_item_data.get('discount_percent', 0),
            is_free_item=line_item_data.get('is_free_item', False),
            is_interstate=is_interstate,
            conversion_factor=line_item_data.get('units_per_pack', 1)
        )
        
        # Create a new dict with all original data plus GST calculations
        result = line_item_data.copy()
        
        # Update with calculated values
        result.update({
            'base_amount': gst_calculations['base_amount'],
            'discount_amount': gst_calculations['discount_amount'],
            'taxable_amount': gst_calculations['taxable_amount'],
            'cgst_rate': 0 if is_interstate else line_item_data.get('gst_rate', 0) / 2,
            'sgst_rate': 0 if is_interstate else line_item_data.get('gst_rate', 0) / 2,
            'igst_rate': line_item_data.get('gst_rate', 0) if is_interstate else 0,
            'cgst': gst_calculations['cgst_amount'],
            'sgst': gst_calculations['sgst_amount'],
            'igst': gst_calculations['igst_amount'],
            'total_gst': gst_calculations['total_gst_amount'],
            'line_total': gst_calculations['line_total'],
            'unit_price': gst_calculations.get('sub_unit_price', 0)
        })
        
        return result
        
    except Exception as e:
        logger.error(f"GST calculation error: {str(e)}")
        # Fallback calculation
        return calculate_line_item_with_gst_fallback(line_item_data, is_interstate)


def calculate_line_item_with_gst_fallback(line_item_data, is_interstate=False):
    """
    Fallback GST calculation if main calculation fails
    
    Args:
        line_item_data: Dictionary with line item details
        is_interstate: Boolean indicating interstate transaction
        
    Returns:
        dict: Line item with calculated GST values
    """
    result = line_item_data.copy()
    
    # Basic calculations
    units = float(line_item_data.get('units', 0))
    pack_price = float(line_item_data.get('pack_purchase_price', 0))
    discount_percent = float(line_item_data.get('discount_percent', 0))
    gst_rate = float(line_item_data.get('gst_rate', 0))
    units_per_pack = float(line_item_data.get('units_per_pack', 1)) or 1
    
    base_amount = units * pack_price
    discount_amount = base_amount * (discount_percent / 100)
    taxable_amount = base_amount - discount_amount
    total_gst = taxable_amount * (gst_rate / 100)
    
    result.update({
        'base_amount': base_amount,
        'discount_amount': discount_amount,
        'taxable_amount': taxable_amount,
        'cgst_rate': 0 if is_interstate else gst_rate / 2,
        'sgst_rate': 0 if is_interstate else gst_rate / 2,
        'igst_rate': gst_rate if is_interstate else 0,
        'cgst': 0 if is_interstate else total_gst / 2,
        'sgst': 0 if is_interstate else total_gst / 2,
        'igst': total_gst if is_interstate else 0,
        'total_gst': total_gst,
        'line_total': taxable_amount + total_gst,
        'unit_price': pack_price / units_per_pack
    })
    
    return result


def process_po_line_items_with_gst(line_items_raw, hospital_id, supplier_id=None):
    """
    Process multiple line items with GST calculations
    
    Args:
        line_items_raw: List of line item dictionaries
        hospital_id: Hospital UUID
        supplier_id: Supplier UUID (optional)
        
    Returns:
        list: Processed line items with GST calculations
    """
    if not line_items_raw:
        return []
    
    # Get state codes once for all line items
    state_info = get_transaction_state_codes(hospital_id, supplier_id)
    is_interstate = state_info['is_interstate']
    
    processed_items = []
    
    for idx, item in enumerate(line_items_raw, 1):
        try:
            # Calculate GST for this line item
            processed_item = calculate_line_item_with_gst(item, is_interstate)
            processed_items.append(processed_item)
            
            logger.info(f"Processed line {idx}: Total={processed_item['line_total']}, GST={processed_item['total_gst']}")
            
        except Exception as e:
            logger.error(f"Error processing line item {idx}: {str(e)}")
            raise ValueError(f"Error in line item {idx}: {str(e)}")
    
    return processed_items


def calculate_po_totals(line_items):
    """
    Calculate total amounts for a purchase order from line items
    
    Args:
        line_items: List of line items with GST calculations
        
    Returns:
        dict: {
            'subtotal': float,          # Total before discount and tax
            'total_discount': float,    # Total discount amount
            'total_taxable': float,     # Total taxable amount
            'total_cgst': float,        # Total CGST
            'total_sgst': float,        # Total SGST
            'total_igst': float,        # Total IGST
            'total_gst': float,         # Total GST amount
            'grand_total': float        # Final total including GST
        }
    """
    totals = {
        'subtotal': 0,
        'total_discount': 0,
        'total_taxable': 0,
        'total_cgst': 0,
        'total_sgst': 0,
        'total_igst': 0,
        'total_gst': 0,
        'grand_total': 0
    }
    
    for item in line_items:
        totals['subtotal'] += float(item.get('base_amount', 0))
        totals['total_discount'] += float(item.get('discount_amount', 0))
        totals['total_taxable'] += float(item.get('taxable_amount', 0))
        totals['total_cgst'] += float(item.get('cgst', 0))
        totals['total_sgst'] += float(item.get('sgst', 0))
        totals['total_igst'] += float(item.get('igst', 0))
        totals['total_gst'] += float(item.get('total_gst', 0))
        totals['grand_total'] += float(item.get('line_total', 0))
    
    return totals


# ===================================================================
# Payment Validation Functions - COMPLETE (unchanged - already have branch support)
# ===================================================================

# ENHANCED VALIDATION FUNCTION - with mixed payment method support
# CRITICAL FIX: Replace the validate_payment_data function in supplier_service.py
# The issue is that this function creates its own session inside validation
# Find this function and REPLACE it completely:

def validate_payment_data(
    payment_data: Dict, 
    hospital_id: uuid.UUID, 
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None  # NEW: Add optional session parameter
) -> Dict:
    """
    Validate payment data with MIXED PAYMENT METHOD support and partial payment business rules
    FIXED: Accept optional session to avoid creating conflicting sessions
    """
    errors = []
    warnings = []
    
    try:
        # Basic validations
        if not payment_data.get('amount'):
            errors.append("Payment amount is required")
            return {'is_valid': False, 'errors': errors, 'warnings': warnings}
        
        try:
            total_amount = Decimal(str(payment_data['amount']))
            if total_amount <= 0:
                errors.append("Payment amount must be greater than zero")
        except (ValueError, TypeError):
            errors.append("Invalid payment amount format")
            return {'is_valid': False, 'errors': errors, 'warnings': warnings}
        
        if not payment_data.get('supplier_id'):
            errors.append("Supplier selection is required")
        
        if not payment_data.get('branch_id'):
            errors.append("Branch selection is required")
        
        if not payment_data.get('payment_date'):
            errors.append("Payment date is required")
        
        # Payment method validation
        payment_method = payment_data.get('payment_method', 'cash')
        
        # ENHANCED: Extract payment method amounts
        cash_amount = Decimal(str(payment_data.get('cash_amount', 0)))
        cheque_amount = Decimal(str(payment_data.get('cheque_amount', 0)))
        bank_amount = Decimal(str(payment_data.get('bank_transfer_amount', 0)))
        upi_amount = Decimal(str(payment_data.get('upi_amount', 0)))
        
        # Calculate total of all payment methods
        method_total = cash_amount + cheque_amount + bank_amount + upi_amount
        
        logger.info(f"PAYMENT VALIDATION - Method amounts: Cash={cash_amount}, Cheque={cheque_amount}, Bank={bank_amount}, UPI={upi_amount}, Total={method_total}")
        
        # =====================================================================
        # CRITICAL FIX: Auto-correct payment method amounts for zero distribution
        # =====================================================================
        if method_total == 0 and total_amount > 0:
            # Auto-populate cash amount to match total amount
            payment_data['cash_amount'] = float(total_amount)
            payment_data['payment_method'] = 'cash'
            cash_amount = total_amount
            method_total = total_amount
            logger.info(f"AUTO-CORRECTED: Set cash_amount to {total_amount} for single-method payment")
        
        # BUSINESS RULE: Total of payment methods must equal declared payment amount
        if method_total > 0:
            if abs(method_total - total_amount) > Decimal('0.01'):
                errors.append(f"Sum of payment method amounts ( Rs.{method_total}) must equal total payment amount ( Rs.{total_amount})")
        
        # BUSINESS RULE: Method-specific validations - only for active methods
        
        # Cheque validation - only if cheque amount > 0
        if cheque_amount > 0:
            if not payment_data.get('cheque_number'):
                errors.append("Cheque number is required when cheque amount is specified")
            if not payment_data.get('cheque_bank'):
                errors.append("Bank name is required when cheque amount is specified")
        
        # Bank transfer validation - only if bank amount > 0
        if bank_amount > 0:
            if not payment_data.get('bank_reference_number'):
                errors.append("Bank reference number is required when bank transfer amount is specified")
            if not payment_data.get('bank_account_name'):
                errors.append("Bank account name is required when bank transfer amount is specified")
        
        # UPI validation - only if UPI amount > 0
        if upi_amount > 0:
            if not payment_data.get('upi_transaction_id'):
                errors.append("UPI transaction ID is required when UPI amount is specified")
        
        # BUSINESS RULE: If multiple payment methods used, set payment_method to 'mixed'
        active_methods = sum([1 for amt in [cash_amount, cheque_amount, bank_amount, upi_amount] if amt > 0])
        if active_methods > 1:
            payment_data['payment_method'] = 'mixed'  # Auto-correct payment method
            logger.info(f"BUSINESS RULE: Multiple payment methods detected, set payment_method to 'mixed'")
        elif active_methods == 1:
            # Single method - ensure payment_method matches the active method
            if cash_amount > 0 and payment_method not in ['cash', 'mixed']:
                payment_data['payment_method'] = 'cash'
            elif cheque_amount > 0 and payment_method not in ['cheque', 'mixed']:
                payment_data['payment_method'] = 'cheque'
            elif bank_amount > 0 and payment_method not in ['bank_transfer', 'mixed']:
                payment_data['payment_method'] = 'bank_transfer'
            elif upi_amount > 0 and payment_method not in ['upi', 'mixed']:
                payment_data['payment_method'] = 'upi'
        
        #   IMPROVED: Invoice validation - only if session is provided
        if payment_data.get('invoice_id') and session:
            try:
                current_balance = get_current_invoice_balance(
                    uuid.UUID(payment_data['invoice_id']), 
                    session
                )
                
                if float(total_amount) > current_balance:
                    overpayment = float(total_amount) - current_balance
                    errors.append(f"Payment amount exceeds remaining balance by  Rs.{overpayment:.2f}")
                elif float(total_amount) < current_balance:
                    remaining = current_balance - float(total_amount)
                    warnings.append(f"Partial payment:  Rs.{remaining:.2f} will remain unpaid")
                    
            except Exception as e:
                logger.error(f"Error validating against invoice: {str(e)}")
                warnings.append("Could not validate against invoice balance")
        elif payment_data.get('invoice_id'):
            # If invoice_id provided but no session, just add warning
            warnings.append("Invoice validation will be performed during payment recording")
        
        # Approval threshold check
        if float(total_amount) >= PAYMENT_CONFIG.get('APPROVAL_THRESHOLD_L1', 50000):
            warnings.append('This payment will require approval')
        
        # Auto-approve threshold check
        if float(total_amount) <= PAYMENT_CONFIG.get('AUTO_APPROVE_LIMIT', 5000):
            warnings.append('This payment will be auto-approved')
        
    except Exception as validation_error:
        logger.error(f"Error during payment validation: {str(validation_error)}")
        errors.append(f"Validation error: {str(validation_error)}")
    
    #   ENHANCED: Invoice validation with better error messages
    if payment_data.get('invoice_id') and session:
        try:
            current_balance = get_current_invoice_balance(
                uuid.UUID(payment_data['invoice_id']), 
                session
            )
            
            if float(total_amount) > current_balance:
                overpayment = float(total_amount) - current_balance
                #   ENHANCED: More descriptive error message
                errors.append(f"Payment amount exceeds remaining balance by  Rs.{overpayment:.2f}")
            elif float(total_amount) < current_balance:
                remaining = current_balance - float(total_amount)
                warnings.append(f"Partial payment:  Rs.{remaining:.2f} will remain unpaid")
                
        except Exception as e:
            logger.error(f"Error validating against invoice: {str(e)}")
            warnings.append("Could not validate against invoice balance")
    elif payment_data.get('invoice_id'):
        warnings.append("Invoice validation will be performed during payment recording")

    return {
        'is_valid': len(errors) == 0,
        'errors': errors,
        'warnings': warnings
    }

# ADDITIONAL INTERNAL VALIDATION FUNCTION - FIXED for payment_total issue
def _validate_payment_data_internal(
    session: Session,
    payment_data: Dict,
    hospital_id: uuid.UUID
) -> List[str]:
    """
    Internal payment data validation with database checks - FIXED to allow partial payments
    """
    errors = []
    
    try:
        # Validate supplier exists
        supplier_id = payment_data.get('supplier_id')
        if supplier_id:
            supplier = session.query(Supplier).filter_by(
                supplier_id=supplier_id,
                hospital_id=hospital_id
            ).first()
            
            if not supplier:
                errors.append(f"Supplier with ID {supplier_id} not found")
            elif supplier.black_listed:
                errors.append(f"Cannot make payment to blacklisted supplier: {supplier.supplier_name}")
        
        # Validate invoice if provided - FIXED to allow partial payments
        invoice_id = payment_data.get('invoice_id')
        if invoice_id:
            invoice = session.query(SupplierInvoice).filter_by(
                invoice_id=invoice_id,
                hospital_id=hospital_id
            ).first()
            
            if not invoice:
                errors.append(f"Invoice with ID {invoice_id} not found")
            elif invoice.payment_status == 'paid':
                errors.append("Invoice is already fully paid")
            elif str(invoice.supplier_id) != str(supplier_id):
                errors.append("Invoice supplier does not match payment supplier")
            else:
                # Calculate existing APPROVED payments only
                existing_payment_total = session.query(func.sum(SupplierPayment.amount)).filter_by(
                    invoice_id=invoice_id,
                    workflow_status='approved'
                ).scalar() or 0
                
                payment_amount = float(payment_data.get('amount', 0))
                total_with_new_payment = float(existing_payment_total) + payment_amount
                invoice_total = float(invoice.total_amount)
                
                #   CRITICAL FIX: Only error if payment would EXCEED invoice total
                if total_with_new_payment > invoice_total:
                    overpayment = total_with_new_payment - invoice_total
                    errors.append(f"Payment would exceed invoice balance by  Rs.{overpayment:.2f}")
                
                #   NEW: Log partial payment info (don't add to errors)
                remaining_balance = invoice_total - float(existing_payment_total)
                if payment_amount < remaining_balance:
                    logger.info(f"PARTIAL PAYMENT: Paying  Rs.{payment_amount:.2f} of  Rs.{remaining_balance:.2f} remaining")
    
    except Exception as e:
        errors.append(f"Error in internal validation: {str(e)}")
        logger.error(f"Internal validation error: {str(e)}")
    
    return errors


# BUSINESS RULES VALIDATION FUNCTION - FIXED for payment_total issue
def validate_payment_business_rules(
    session: Session,
    payment_data: Dict,
    hospital_id: uuid.UUID
) -> List[str]:
    """
    Validate business rules for supplier payments - FIXED to allow partial payments
    """
    errors = []
    
    try:
        invoice_id = payment_data.get('invoice_id')
        amount = float(payment_data.get('amount', 0))
        
        if invoice_id and amount > 0:
            # Get invoice details
            invoice = session.query(SupplierInvoice).filter_by(
                invoice_id=invoice_id,
                hospital_id=hospital_id
            ).first()
            
            if invoice:
                # Calculate existing APPROVED payments only
                existing_payments = session.query(func.sum(SupplierPayment.amount)).filter_by(
                    invoice_id=invoice_id,
                    workflow_status='approved'  # Only count approved payments
                ).scalar() or 0
                
                total_with_new_payment = float(existing_payments) + amount
                invoice_total = float(invoice.total_amount)
                
                #   CRITICAL FIX: Only error if payment EXCEEDS invoice total (allow partial payments)
                if total_with_new_payment > invoice_total:
                    overpayment = total_with_new_payment - invoice_total
                    errors.append(
                        f"Payment amount ( Rs.{amount:.2f}) would exceed invoice balance. "
                        f"Invoice total:  Rs.{invoice_total:.2f}, "
                        f"Already paid:  Rs.{float(existing_payments):.2f}, "
                        f"Overpayment:  Rs.{overpayment:.2f}"
                    )
                
                #   CRITICAL FIX: Only error if invoice is already FULLY paid
                if float(existing_payments) >= invoice_total:
                    errors.append("Invoice is already fully paid")
                
                #   NEW: Add informational logging for partial payments (don't add to errors)
                remaining_after_payment = invoice_total - total_with_new_payment
                if remaining_after_payment > 0.01:  # Small tolerance for decimal precision
                    logger.info(f"PARTIAL PAYMENT: Amount= Rs.{amount:.2f}, Remaining= Rs.{remaining_after_payment:.2f}")
                    
    except Exception as e:
        errors.append(f"Error in business rules validation: {str(e)}")
        logger.error(f"Business rules validation error: {str(e)}")
    
    return errors

    """
    Helper function to refresh entity in current session
    Prevents detached entity issues per developer guidelines
    """
    try:
        return session.query(entity_class).filter_by(
            **{f"{entity_class.__tablename__.rstrip('s')}_id": entity_id}
        ).first()
    except Exception as e:
        logger.error(f"Error refreshing entity in session: {str(e)}")
        return None

def enhanced_record_supplier_payment(
    hospital_id: uuid.UUID,
    payment_data: Dict,
    create_gl_entries: bool = True,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Enhanced payment recording with comprehensive validation - EXISTING FUNCTION (unchanged)
    """
    logger.info(f"Recording enhanced supplier payment for hospital {hospital_id}")
    
    # Perform basic validation first
    validation_errors = validate_payment_data(payment_data)
    if validation_errors:
        error_msg = "Payment validation failed:\n" + "\n".join(validation_errors)
        logger.error(error_msg)
        raise ValueError(error_msg)
    
    # Use the standard payment recording process
    return record_supplier_payment(
        hospital_id=hospital_id,
        payment_data=payment_data,
        create_gl_entries=create_gl_entries,
        current_user_id=current_user_id,
        session=session
    )

def get_supplier_branch_context(current_user_id, hospital_id):
    """
    Convenience function for views to get supplier-specific branch context
    """
    return get_user_branch_context(current_user_id, hospital_id, 'supplier')

def _determine_approval_requirement(payment_data: Dict, current_user_id: Optional[str] = None) -> bool:
    """Determine if payment requires approval"""
    try:
        amount = float(payment_data.get('amount', 0))
        
        if amount >= 50000.00:
            return True
        
        if current_user_id:
            try:
                from app.services.permission_service import has_permission
                user_obj = {'user_id': current_user_id}
                if has_permission(user_obj, 'payment', 'approve'):
                    return amount >= 200000.00
            except Exception as e:
                logger.warning(f"Role check failed: {str(e)}")
        
        return False
        
    except Exception as e:
        logger.error(f"Error determining approval requirement: {str(e)}")
        return True

def _get_approval_level(amount: float) -> str:
    """Get approval level based on amount"""
    if amount <= 5000:
        return 'auto_approved'
    elif amount <= 50000:
        return 'level_1'
    else:
        return 'level_2'

def _process_payment_documents(
    payment_id: uuid.UUID,
    documents: List[Dict],
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> None:
    """Process document uploads for payment - COMPLETE FIX"""
    try:
        import os
        from werkzeug.utils import secure_filename
        from flask import current_app
        
        # FIX: Get payment record FIRST to access hospital_id and branch_id
        payment = session.query(SupplierPayment).filter_by(payment_id=payment_id).first()
        if not payment:
            raise ValueError(f"Payment {payment_id} not found for document processing")
        
        logger.info(f"Processing {len(documents)} documents for payment {payment_id}")
        
        # Get base path for document storage
        base_path = current_app.config.get('PAYMENT_DOCUMENT_PATH', '/tmp/payment_documents')
        
        for i, doc_data in enumerate(documents):
            if 'file' not in doc_data:
                logger.warning(f"Document {i+1} has no file data - skipping")
                continue
                
            file = doc_data['file']
            document_type = doc_data.get('document_type', 'receipt')
            
            # Validate file
            if not file or not file.filename:
                logger.warning(f"Document {i+1} has no filename - skipping")
                continue
                
            # Validate and save file
            filename = secure_filename(file.filename)
            file_extension = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
            
            if file_extension not in {'pdf', 'jpg', 'jpeg', 'png'}:
                logger.error(f"Invalid file type .{file_extension} for document {i+1}")
                raise ValueError(f"File type .{file_extension} not allowed")
            
            # Create storage structure
            payment_folder = os.path.join(base_path, str(payment_id))
            os.makedirs(payment_folder, exist_ok=True)
            
            unique_filename = f"{document_type}_{datetime.now().strftime('%Y%m%d_%H%M%S')}.{file_extension}"
            file_path = os.path.join(payment_folder, unique_filename)
            
            # Save file to disk
            file.save(file_path)
            logger.info(f"File saved to: {file_path}")
            
            # Create document record with ALL required fields
            from app.models.transaction import PaymentDocument
            document = PaymentDocument(
                payment_id=payment_id,
                hospital_id=payment.hospital_id,  # Now payment is defined
                branch_id=payment.branch_id,      # Now payment is defined
                document_type=document_type,
                original_filename=filename,
                stored_filename=unique_filename,
                file_path=file_path,
                file_size=os.path.getsize(file_path),
                mime_type=file.mimetype if hasattr(file, 'mimetype') else None,
                file_extension=file_extension,
                description=doc_data.get('description', ''),
                required_for_approval=doc_data.get('required_for_approval', False),
                created_by=current_user_id
            )
            
            session.add(document)
            session.flush()  # Ensure document is saved immediately
            
            logger.info(f"Document saved to database: {document.document_id} ({document_type})")
        
        # Update payment document count
        if documents:
            session.query(SupplierPayment).filter_by(payment_id=payment_id).update({
                'total_documents_count': SupplierPayment.total_documents_count + len([d for d in documents if 'file' in d])
            })
        
        logger.info(f"Completed processing documents for payment {payment_id}")
        
    except Exception as e:
        logger.error(f"Error processing payment documents: {str(e)}")
        raise


# ADD new function for payment approval

def approve_supplier_payment(
    payment_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: str,
    approval_notes: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """Approve pending payment"""
    logger.info(f"Approving payment {payment_id}")
    
    if session is not None:
        return _approve_supplier_payment(session, payment_id, hospital_id, current_user_id, approval_notes)
    
    with get_db_session() as new_session:
        result = _approve_supplier_payment(new_session, payment_id, hospital_id, current_user_id, approval_notes)
        new_session.commit()
        return result

def _approve_supplier_payment(
    session: Session,
    payment_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: str,
    approval_notes: Optional[str] = None
) -> Dict:
    """Internal function for payment approval"""
    try:
        payment = session.query(SupplierPayment).filter_by(
            payment_id=payment_id,
            hospital_id=hospital_id
        ).first()
        
        if not payment:
            raise ValueError("Payment not found")
        
        if payment.workflow_status == 'approved':
            raise ValueError("Payment is already approved")
        
        # Validate approval permission
        if not _can_approve_payment(current_user_id, float(payment.amount)):
            raise ValueError("User does not have permission to approve this payment")
        
        # Update payment
        payment.workflow_status = 'approved'
        payment.approved_by = current_user_id
        payment.approved_at = datetime.now(timezone.utc)
        
        if approval_notes:
            payment.approval_notes = approval_notes
        
        session.flush()
        
        # Create GL entries
        try:
            gl_result = create_supplier_payment_gl_entries(payment_id, current_user_id, session)
            logger.info("GL entries created after approval")
        except Exception as e:
            logger.error(f"Error creating GL entries after approval: {str(e)}")
        
        return get_entity_dict(payment)
        
    except Exception as e:
        logger.error(f"Error approving payment: {str(e)}")
        session.rollback()
        raise

def _can_approve_payment(user_id: str, amount: float) -> bool:
    """Check approval permissions"""
    try:
        from app.services.permission_service import has_permission
        
        user_obj = {'user_id': user_id}
        
        if not has_permission(user_obj, 'payment', 'approve'):
            return False
        
        if amount >= 200000.00:
            return has_permission(user_obj, 'payment', 'approve_high_value')
        
        return True
        
    except Exception as e:
        logger.error(f"Error checking approval permission: {str(e)}")
        return False

# =====================================================================
# ADDITIONAL PAYMENT FUNCTIONS TO ADD TO supplier_service.py
# Add these functions to your existing supplier_service.py file
# =====================================================================

def search_supplier_payments(
    hospital_id: uuid.UUID,
    filters: Dict = None,
    branch_id: Optional[uuid.UUID] = None,
    current_user_id: Optional[str] = None,
    page: int = 1,
    per_page: int = 20,
    session: Optional[Session] = None
) -> Dict:
    """Search supplier payments with filtering support"""
    logger.info(f"Searching supplier payments for hospital {hospital_id}")
    
    if session is not None:
        return _search_supplier_payments(session, hospital_id, filters, branch_id, current_user_id, page, per_page)
    
    with get_db_session(read_only=True) as new_session:
        return _search_supplier_payments(new_session, hospital_id, filters, branch_id, current_user_id, page, per_page)


def _search_supplier_payments(
    session: Session,
    hospital_id: uuid.UUID,
    filters: Dict = None,
    branch_id: Optional[uuid.UUID] = None,
    current_user_id: Optional[str] = None,
    page: int = 1,
    per_page: int = 20
) -> Dict:
    """DEBUG VERSION: Find exact hanging point"""
    import time
    start_time = time.time()
    
    # CRITICAL DEBUG: Log everything
    logger.info(f"🔍 PAYMENT SEARCH START - Time: {time.strftime('%H:%M:%S')}")
    logger.info(f"🔍 Filters received: {filters}")
    logger.info(f"🔍 Hospital ID: {hospital_id}")
    logger.info(f"🔍 Branch ID: {branch_id}")
    logger.info(f"🔍 User ID: {current_user_id}")
    
    try:
        filters = filters or {}
        
        # STEP 1: Base query
        step_start = time.time()
        query = session.query(SupplierPayment).filter_by(hospital_id=hospital_id)
        logger.info(f"✅ STEP 1 - Base query: {time.time() - step_start:.3f}s")
        
        # STEP 2: Branch filtering
        step_start = time.time()
        if branch_id:
            query = query.filter(SupplierPayment.branch_id == branch_id)
            logger.info(f"Applied direct branch filter: {branch_id}")
        elif current_user_id:
            try:
                accessible_branches = get_user_accessible_branches(current_user_id, hospital_id)
                if accessible_branches:
                    branch_ids = [uuid.UUID(b['branch_id']) for b in accessible_branches]
                    query = query.filter(SupplierPayment.branch_id.in_(branch_ids))
                    logger.info(f"Applied user branch filter: {len(branch_ids)} branches")
            except Exception as e:
                logger.warning(f"Branch filtering failed: {str(e)}")
        logger.info(f"✅ STEP 2 - Branch filter: {time.time() - step_start:.3f}s")
        
        # STEP 3: Individual filters - TEST EACH ONE
        
        # Supplier ID
        if filters.get('supplier_id'):
            step_start = time.time()
            try:
                supplier_uuid = uuid.UUID(filters['supplier_id'])
                query = query.filter(SupplierPayment.supplier_id == supplier_uuid)
                logger.info(f"✅ STEP 3a - Supplier ID filter: {time.time() - step_start:.3f}s")
            except Exception as e:
                logger.error(f"❌ STEP 3a - Supplier ID filter FAILED: {e}")
        
        # Payment method (single)
        if filters.get('payment_method'):
            step_start = time.time()
            method = filters['payment_method']
            logger.info(f"🔍 Applying single payment method: {method}")
            
            if method == 'cash':
                query = query.filter(SupplierPayment.cash_amount > 0)
            elif method == 'cheque':
                query = query.filter(SupplierPayment.cheque_amount > 0)
            elif method == 'bank_transfer':
                query = query.filter(SupplierPayment.bank_transfer_amount > 0)
            elif method == 'upi':
                query = query.filter(SupplierPayment.upi_amount > 0)
            
            logger.info(f"✅ STEP 3b - Single payment method: {time.time() - step_start:.3f}s")
        
        # Payment methods (multiple)
        elif filters.get('payment_methods'):
            step_start = time.time()
            payment_methods = filters['payment_methods']
            logger.info(f"🔍 Applying multiple payment methods: {payment_methods}")
            
            method_conditions = []
            for method in payment_methods:
                if method == 'cash':
                    method_conditions.append(SupplierPayment.cash_amount > 0)
                elif method == 'cheque':
                    method_conditions.append(SupplierPayment.cheque_amount > 0)
                elif method == 'bank_transfer':
                    method_conditions.append(SupplierPayment.bank_transfer_amount > 0)
                elif method == 'upi':
                    method_conditions.append(SupplierPayment.upi_amount > 0)
            
            if method_conditions:
                from sqlalchemy import or_
                query = query.filter(or_(*method_conditions))
            
            logger.info(f"✅ STEP 3c - Multiple payment methods: {time.time() - step_start:.3f}s")
        
        # Status filters
        if filters.get('workflow_status'):
            step_start = time.time()
            query = query.filter(SupplierPayment.workflow_status == filters['workflow_status'])
            logger.info(f"✅ STEP 3d - Single status: {time.time() - step_start:.3f}s")
        elif filters.get('statuses'):
            step_start = time.time()
            statuses = filters['statuses']
            query = query.filter(SupplierPayment.workflow_status.in_(statuses))
            logger.info(f"✅ STEP 3e - Multiple statuses: {time.time() - step_start:.3f}s")
        
        # Date filters
        if filters.get('start_date'):
            step_start = time.time()
            try:
                start_date = datetime.strptime(filters['start_date'], '%Y-%m-%d').date()
                query = query.filter(SupplierPayment.payment_date >= start_date)
                logger.info(f"✅ STEP 3f - Start date: {time.time() - step_start:.3f}s")
            except Exception as e:
                logger.error(f"❌ STEP 3f - Start date FAILED: {e}")
        
        if filters.get('end_date'):
            step_start = time.time()
            try:
                end_date = datetime.strptime(filters['end_date'], '%Y-%m-%d').date()
                query = query.filter(SupplierPayment.payment_date <= end_date)
                logger.info(f"✅ STEP 3g - End date: {time.time() - step_start:.3f}s")
            except Exception as e:
                logger.error(f"❌ STEP 3g - End date FAILED: {e}")
        
        # CRITICAL: Amount filters - this often causes hanging
        if filters.get('min_amount'):
            step_start = time.time()
            logger.info(f"🔍 CRITICAL: Applying min_amount filter: {filters['min_amount']}")
            
            try:
                min_amount = float(filters['min_amount'])
                # CRITICAL: Use direct amount field, NOT calculated
                query = query.filter(SupplierPayment.amount >= min_amount)
                logger.info(f"✅ STEP 3h - Min amount (DIRECT): {time.time() - step_start:.3f}s")
            except Exception as e:
                logger.error(f"❌ STEP 3h - Min amount FAILED: {e}")
        
        # Invoice search
        if filters.get('invoice_id'):
            step_start = time.time()
            logger.info(f"🔍 CRITICAL: Applying invoice search: {filters['invoice_id']}")
            
            try:
                invoice_search = str(filters['invoice_id']).strip()
                invoice_subquery = session.query(SupplierInvoice.invoice_id).filter(
                    SupplierInvoice.hospital_id == hospital_id,
                    SupplierInvoice.supplier_invoice_number.ilike(f'%{invoice_search}%')
                ).subquery()
                query = query.filter(SupplierPayment.invoice_id.in_(invoice_subquery))
                logger.info(f"✅ STEP 3i - Invoice search (SUBQUERY): {time.time() - step_start:.3f}s")
            except Exception as e:
                logger.error(f"❌ STEP 3i - Invoice search FAILED: {e}")
        
        # Supplier name search
        if filters.get('supplier_name_search') and not filters.get('supplier_id'):
            step_start = time.time()
            logger.info(f"🔍 CRITICAL: Applying supplier name search: {filters['supplier_name_search']}")
            
            try:
                supplier_search = str(filters['supplier_name_search']).strip()
                supplier_subquery = session.query(Supplier.supplier_id).filter(
                    Supplier.hospital_id == hospital_id,
                    Supplier.supplier_name.ilike(f'%{supplier_search}%')
                ).subquery()
                query = query.filter(SupplierPayment.supplier_id.in_(supplier_subquery))
                logger.info(f"✅ STEP 3j - Supplier name search (SUBQUERY): {time.time() - step_start:.3f}s")
            except Exception as e:
                logger.error(f"❌ STEP 3j - Supplier name search FAILED: {e}")
        
        # STEP 4: Query execution - this is where most hangs occur
        logger.info(f"🔍 CRITICAL: About to execute query with {len(filters)} filters")
        
        # Ordering
        step_start = time.time()
        query = query.order_by(SupplierPayment.payment_date.desc())
        logger.info(f"✅ STEP 4a - Ordering: {time.time() - step_start:.3f}s")
        
        # Count query (often hangs here)
        step_start = time.time()
        logger.info(f"🔍 CRITICAL: Executing COUNT query...")
        total_count = query.count()
        count_time = time.time() - step_start
        logger.info(f"✅ STEP 4b - Count query: {count_time:.3f}s - Found {total_count} records")
        
        if count_time > 5.0:
            logger.error(f"❌ COUNT QUERY TOO SLOW: {count_time:.3f}s - This is the problem!")
        
        # Main query execution (often hangs here)
        step_start = time.time()
        logger.info(f"🔍 CRITICAL: Executing MAIN query...")
        offset = (page - 1) * per_page
        payments = query.offset(offset).limit(per_page).all()
        main_time = time.time() - step_start
        logger.info(f"✅ STEP 4c - Main query: {main_time:.3f}s - Retrieved {len(payments)} records")
        
        if main_time > 5.0:
            logger.error(f"❌ MAIN QUERY TOO SLOW: {main_time:.3f}s - This is the problem!")
        
        # Result processing
        step_start = time.time()
        payment_dicts = []
        
        for payment in payments:
            payment_dict = get_entity_dict(payment)
            
            # CRITICAL FIX: Add supplier information
            supplier = session.query(Supplier).filter_by(supplier_id=payment.supplier_id).first()
            if supplier:
                payment_dict['supplier_name'] = supplier.supplier_name
                payment_dict['supplier_code'] = getattr(supplier, supplier.supplier_id)
            else:
                payment_dict['supplier_name'] = 'N/A'
                payment_dict['supplier_code'] = 'N/A'
            
            # Add branch information if available
            if payment.branch_id:
                branch = session.query(Branch).filter_by(branch_id=payment.branch_id).first()
                if branch:
                    payment_dict['branch_name'] = branch.name
            
            # Add invoice information if linked
            if payment.invoice_id:
                invoice = session.query(SupplierInvoice).filter_by(invoice_id=payment.invoice_id).first()
                if invoice:
                    payment_dict['invoice_number'] = invoice.supplier_invoice_number
                    payment_dict['invoice_amount'] = float(invoice.total_amount)
            
            payment_dicts.append(payment_dict)
        
        # Summary calculation
        step_start = time.time()
        try:
            # Use the existing _calculate_payment_summary function
            summary = _calculate_payment_summary(
                session=session,
                hospital_id=hospital_id,
                filters=filters,
                branch_id=branch_id,
                current_user_id=current_user_id
            )
            logger.info(f"✅ STEP 6 - Summary calculation: {time.time() - step_start:.3f}s")
            logger.info(f"Summary calculated: total_amount={summary.get('total_amount', 0)}, total_count={summary.get('total_count', 0)}")
        except Exception as e:
            logger.error(f"Error calculating payment summary: {str(e)}")
            # Fallback to safe defaults if calculation fails
            summary = {
                'total_count': total_count,
                'total_amount': 0.0,
                'pending_count': 0,
                'this_month_count': 0
            }
        logger.info(f"✅ STEP 6 - Summary: {time.time() - step_start:.3f}s")
        
        # Final result
        total_time = time.time() - start_time
        logger.info(f"🎯 TOTAL SEARCH TIME: {total_time:.3f}s")
        
        if total_time > 10.0:
            logger.error(f"❌ SEARCH TOO SLOW: {total_time:.3f}s - PERFORMANCE ISSUE!")
        else:
            logger.info(f"✅ SEARCH COMPLETED SUCCESSFULLY in {total_time:.3f}s")
        
        return {
            'payments': payment_dicts,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total_count': total_count,
                'total_pages': (total_count + per_page - 1) // per_page
            },
            'summary': summary
        }
        
    except Exception as e:
        error_time = time.time() - start_time
        logger.error(f"💥 ERROR after {error_time:.3f}s: {str(e)}", exc_info=True)
        raise

def _calculate_payment_summary(
    session: Session, 
    hospital_id: uuid.UUID, 
    filters: Dict, 
    branch_id: Optional[uuid.UUID],
    current_user_id: Optional[str]
) -> Dict:
    """Calculate payment summary efficiently without complex joins"""
    try:
        # Build base query for summary (same filters as main query)
        summary_query = session.query(SupplierPayment).filter_by(hospital_id=hospital_id)
        
        # Apply same filters as main query
        if branch_id:
            summary_query = summary_query.filter(SupplierPayment.branch_id == branch_id)
        
        # ❌ PROBLEM: This comment says "Apply other filters" but then doesn't!
        # Apply other filters similarly to main query
        # (simplified for performance)
        
        # ✅ SOLUTION: Actually apply the filters like the main query does
        
        # Date filters
        if filters.get('start_date'):
            try:
                start_date = datetime.strptime(filters['start_date'], '%Y-%m-%d').date()
                summary_query = summary_query.filter(SupplierPayment.payment_date >= start_date)
            except Exception:
                pass
        
        if filters.get('end_date'):
            try:
                end_date = datetime.strptime(filters['end_date'], '%Y-%m-%d').date()
                summary_query = summary_query.filter(SupplierPayment.payment_date <= end_date)
            except Exception:
                pass
        
        # Supplier filters
        if filters.get('supplier_id'):
            try:
                supplier_uuid = uuid.UUID(filters['supplier_id'])
                summary_query = summary_query.filter(SupplierPayment.supplier_id == supplier_uuid)
            except Exception:
                pass
        
        # Supplier name search (for text input fallback)
        if filters.get('supplier_name_search') and not filters.get('supplier_id'):
            try:
                supplier_search = str(filters['supplier_name_search']).strip()
                supplier_subquery = session.query(Supplier.supplier_id).filter(
                    Supplier.hospital_id == hospital_id,
                    Supplier.supplier_name.ilike(f'%{supplier_search}%')
                ).subquery()
                summary_query = summary_query.filter(SupplierPayment.supplier_id.in_(supplier_subquery))
            except Exception:
                pass
        
        # Reference number search
        if filters.get('reference_no_search'):
            try:
                reference_search = str(filters['reference_no_search']).strip()
                summary_query = summary_query.filter(SupplierPayment.reference_no.ilike(f'%{reference_search}%'))
            except Exception:
                pass

        # Payment method filters
        if filters.get('payment_method'):
            method = filters['payment_method']
            if method == 'cash':
                summary_query = summary_query.filter(SupplierPayment.cash_amount > 0)
            elif method == 'cheque':
                summary_query = summary_query.filter(SupplierPayment.cheque_amount > 0)
            elif method == 'bank_transfer':
                summary_query = summary_query.filter(SupplierPayment.bank_transfer_amount > 0)
            elif method == 'upi':
                summary_query = summary_query.filter(SupplierPayment.upi_amount > 0)
        
        # Multiple payment methods
        elif filters.get('payment_methods'):
            payment_methods = filters['payment_methods']
            method_conditions = []
            for method in payment_methods:
                if method == 'cash':
                    method_conditions.append(SupplierPayment.cash_amount > 0)
                elif method == 'cheque':
                    method_conditions.append(SupplierPayment.cheque_amount > 0)
                elif method == 'bank_transfer':
                    method_conditions.append(SupplierPayment.bank_transfer_amount > 0)
                elif method == 'upi':
                    method_conditions.append(SupplierPayment.upi_amount > 0)
            
            if method_conditions:
                from sqlalchemy import or_
                summary_query = summary_query.filter(or_(*method_conditions))
        
        # Status filters
        if filters.get('workflow_status'):
            summary_query = summary_query.filter(SupplierPayment.workflow_status == filters['workflow_status'])
        elif filters.get('statuses'):
            summary_query = summary_query.filter(SupplierPayment.workflow_status.in_(filters['statuses']))
        
        # Amount filters
        if filters.get('min_amount'):
            try:
                min_amount = float(filters['min_amount'])
                summary_query = summary_query.filter(SupplierPayment.amount >= min_amount)
            except Exception:
                pass
        
        if filters.get('max_amount'):
            try:
                max_amount = float(filters['max_amount'])
                summary_query = summary_query.filter(SupplierPayment.amount <= max_amount)
            except Exception:
                pass
        
        # Invoice search
        if filters.get('invoice_id'):
            try:
                invoice_search = str(filters['invoice_id']).strip()
                invoice_subquery = session.query(SupplierInvoice.invoice_id).filter(
                    SupplierInvoice.hospital_id == hospital_id,
                    SupplierInvoice.supplier_invoice_number.ilike(f'%{invoice_search}%')
                ).subquery()
                summary_query = summary_query.filter(SupplierPayment.invoice_id.in_(invoice_subquery))
            except Exception:
                pass
        
        # Calculate summary efficiently with all filters applied
        total_count = summary_query.count()
        total_amount = summary_query.with_entities(
            func.sum(SupplierPayment.amount)
        ).scalar() or 0
        
        pending_count = summary_query.filter(
            SupplierPayment.workflow_status == 'pending'
        ).count()
        
        # This month count
        from datetime import datetime, date
        this_month_start = date.today().replace(day=1)
        this_month_count = summary_query.filter(
            SupplierPayment.payment_date >= this_month_start
        ).count()
        
        return {
            'total_count': total_count,
            'total_amount': float(total_amount),
            'pending_count': pending_count,
            'this_month_count': this_month_count
        }
        
    except Exception as e:
        logger.warning(f"Error calculating summary: {str(e)}")
        return {
            'total_count': 0,
            'total_amount': 0.0,
            'pending_count': 0,
            'this_month_count': 0
        }


def _determine_payment_method_display(payment):
    """
    Helper function to determine the correct payment method display label
    """
    cash_amount = float(payment.cash_amount or 0)
    cheque_amount = float(payment.cheque_amount or 0)
    bank_amount = float(payment.bank_transfer_amount or 0)
    upi_amount = float(payment.upi_amount or 0)
    
    # Count active payment methods
    active_methods = []
    if cash_amount > 0:
        active_methods.append('cash')
    if cheque_amount > 0:
        active_methods.append('cheque')
    if bank_amount > 0:
        active_methods.append('bank_transfer')
    if upi_amount > 0:
        active_methods.append('upi')
    
    # Determine display method
    if len(active_methods) == 1:
        return active_methods[0]
    elif len(active_methods) > 1:
        return 'mixed'
    else:
        return 'cash'  # Default fallback

def get_supplier_payment_by_id(
    payment_id: uuid.UUID,
    hospital_id: uuid.UUID,
    include_documents: bool = False,
    include_approvals: bool = False,
    session: Optional[Session] = None
) -> Optional[Dict]:
    """Get supplier payment by ID with optional related data"""
    logger.info(f"Getting payment {payment_id} for hospital {hospital_id}")
    
    if session is not None:
        return _get_supplier_payment_by_id(session, payment_id, hospital_id, include_documents, include_approvals)
    
    with get_db_session(read_only=True) as new_session:
        return _get_supplier_payment_by_id(new_session, payment_id, hospital_id, include_documents, include_approvals)


def _get_supplier_payment_by_id(
    session: Session,
    payment_id: uuid.UUID,
    hospital_id: uuid.UUID,
    include_documents: bool = False,
    include_approvals: bool = False
) -> Optional[Dict]:
    """Internal function to get payment by ID"""
    try:
        payment = session.query(SupplierPayment).filter_by(
            payment_id=payment_id,
            hospital_id=hospital_id
        ).first()
        
        if not payment:
            return None
        
        payment_dict = get_entity_dict(payment)
        
        # Add supplier information
        supplier = session.query(Supplier).filter_by(supplier_id=payment.supplier_id).first()
        if supplier:
            payment_dict['supplier_name'] = supplier.supplier_name
            payment_dict['supplier_code'] = getattr(supplier, supplier.supplier_id)
        
        # Add branch information
        if payment.branch_id:
            branch = session.query(Branch).filter_by(branch_id=payment.branch_id).first()
            if branch:
                payment_dict['branch_name'] = branch.name
        
        # Add invoice information if linked
        if payment.invoice_id:
            invoice = session.query(SupplierInvoice).filter_by(invoice_id=payment.invoice_id).first()
            if invoice:
                payment_dict['invoice_number'] = invoice.supplier_invoice_number
                payment_dict['invoice_amount'] = float(invoice.total_amount)
        
        # Add documents if requested
        if include_documents:
            documents = session.query(PaymentDocument).filter_by(
                payment_id=payment_id
            ).order_by(PaymentDocument.updated_at.desc()).all()
            
            payment_dict['documents'] = [get_entity_dict(doc) for doc in documents]
            payment_dict['document_count'] = len(documents)
        
        # Include approval history if requested
        if include_approvals:
            # This could be enhanced with a proper approval history table
            payment_dict['approval_history'] = []
            if payment.approved_by and payment.approved_at:
                payment_dict['approval_history'].append({
                    'action': 'approved',
                    'user': payment.approved_by,
                    'timestamp': payment.approved_at,
                    'notes': getattr(payment, 'approval_notes', None)
                })
        
        return payment_dict
        
    except Exception as e:
        logger.error(f"Error getting payment by ID: {str(e)}")
        raise


def add_payment_document(
    payment_id: uuid.UUID,
    document_data: Dict,
    current_user_id: str,
    session: Optional[Session] = None
) -> Dict:
    """Add document to payment"""
    logger.info(f"Adding document to payment {payment_id}")
    
    if session is not None:
        return _add_payment_document(session, payment_id, document_data, current_user_id)
    
    with get_db_session() as new_session:
        result = _add_payment_document(new_session, payment_id, document_data, current_user_id)
        new_session.commit()
        return result


def _add_payment_document(
    session: Session,
    payment_id: uuid.UUID,
    document_data: Dict,
    current_user_id: str
) -> Dict:
    """Internal function to add payment document - COMPLETE FIX"""
    try:
        import os
        from werkzeug.utils import secure_filename
        from flask import current_app
        
        # FIX: Get payment to verify it exists and get required fields
        payment = session.query(SupplierPayment).filter_by(payment_id=payment_id).first()
        if not payment:
            raise ValueError("Payment not found")
        
        file = document_data['file']
        document_type = document_data.get('document_type', 'other')
        
        # Validate file
        if not file or not file.filename:
            raise ValueError("No file provided")
        
        filename = secure_filename(file.filename)
        file_extension = filename.rsplit('.', 1)[1].lower() if '.' in filename else ''
        
        # Validate file type
        allowed_extensions = current_app.config.get('PAYMENT_CONFIG', {}).get('ALLOWED_FILE_TYPES', ['pdf', 'jpg', 'jpeg', 'png'])
        if file_extension not in allowed_extensions:
            raise ValueError(f"File type .{file_extension} not allowed")
        
        # Create storage path
        base_path = current_app.config.get('PAYMENT_CONFIG', {}).get('DOCUMENT_STORAGE_PATH', '/tmp/payment_documents')
        payment_folder = os.path.join(base_path, str(payment_id))
        os.makedirs(payment_folder, exist_ok=True)
        
        # Generate unique filename
        import time
        unique_filename = f"{document_type}_{int(time.time())}.{file_extension}"
        file_path = os.path.join(payment_folder, unique_filename)
        
        # Save file
        file.save(file_path)
        
        # Create document record with ALL required fields
        document = PaymentDocument(
            payment_id=payment_id,
            hospital_id=payment.hospital_id,  # Now payment is defined
            branch_id=payment.branch_id,      # Now payment is defined
            document_type=document_type,
            original_filename=filename,
            stored_filename=unique_filename,
            file_path=file_path,
            file_size=os.path.getsize(file_path),
            mime_type=file.mimetype if hasattr(file, 'mimetype') else None,
            file_extension=file_extension,
            description=document_data.get('description', ''),
            required_for_approval=document_data.get('required_for_approval', False),
            created_by=current_user_id
        )
        
        session.add(document)
        session.flush()
        
        # Update payment document count
        payment.total_documents_count = (payment.total_documents_count or 0) + 1
        
        logger.info(f"Standalone document added: {document.document_id} for payment {payment_id}")
        
        return {
            'document_id': document.document_id,
            'filename': unique_filename,
            'file_path': file_path,
            'document_type': document_type
        }
        
    except Exception as e:
        logger.error(f"Error adding payment document: {str(e)}")
        session.rollback()
        raise


def export_supplier_payments(
    hospital_id: uuid.UUID,
    filters: Optional[Dict] = None,
    format: str = 'csv',
    branch_id: Optional[uuid.UUID] = None,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> str:
    """
    Export supplier payments to CSV/Excel format
    """
    logger.info(f"Exporting supplier payments for hospital {hospital_id} in {format} format")
    
    if session is not None:
        return _export_supplier_payments(session, hospital_id, filters, format, branch_id)
    
    with get_db_session() as new_session:
        return _export_supplier_payments(new_session, hospital_id, filters, format, branch_id)

def _export_supplier_payments(
    session: Session,
    hospital_id: uuid.UUID,
    filters: Optional[Dict] = None,
    format: str = 'csv',
    branch_id: Optional[uuid.UUID] = None
) -> str:
    """
    Internal function to export supplier payments within a session
    """
    try:
        import csv
        import io
        from datetime import datetime
        
        logger.debug("Building payment export query")
        
        # Base query with joins
        query = session.query(
            SupplierPayment,
            Supplier.supplier_name,
            Branch.name.label('branch_name')
        ).join(
            Supplier, SupplierPayment.supplier_id == Supplier.supplier_id
        ).outerjoin(
            Branch, SupplierPayment.branch_id == Branch.branch_id
        ).filter(
            SupplierPayment.hospital_id == hospital_id
        )
        
        # Apply branch filter if specified
        if branch_id:
            query = query.filter(SupplierPayment.branch_id == branch_id)
        
        # Apply filters if provided
        if filters:
            if filters.get('supplier_id'):
                query = query.filter(SupplierPayment.supplier_id == uuid.UUID(filters['supplier_id']))
            
            if filters.get('workflow_status'):
                query = query.filter(SupplierPayment.workflow_status == filters['workflow_status'])
            
            if filters.get('payment_method'):
                method = filters['payment_method']
                if method == 'cash':
                    query = query.filter(SupplierPayment.cash_amount > 0)
                elif method == 'cheque':
                    query = query.filter(SupplierPayment.cheque_amount > 0)
                elif method == 'bank_transfer':
                    query = query.filter(SupplierPayment.bank_transfer_amount > 0)
                elif method == 'upi':
                    query = query.filter(SupplierPayment.upi_amount > 0)
            
            if filters.get('start_date'):
                try:
                    start_date = datetime.strptime(filters['start_date'], '%Y-%m-%d').date()
                    query = query.filter(SupplierPayment.payment_date >= start_date)
                except ValueError:
                    logger.warning(f"Invalid start_date format: {filters['start_date']}")
            
            if filters.get('end_date'):
                try:
                    end_date = datetime.strptime(filters['end_date'], '%Y-%m-%d').date()
                    query = query.filter(SupplierPayment.payment_date <= end_date)
                except ValueError:
                    logger.warning(f"Invalid end_date format: {filters['end_date']}")
            
            if filters.get('min_amount'):
                try:
                    query = query.filter(SupplierPayment.amount >= float(filters['min_amount']))
                except (ValueError, TypeError):
                    logger.warning(f"Invalid min_amount: {filters['min_amount']}")
            
            if filters.get('max_amount'):
                try:
                    query = query.filter(SupplierPayment.amount <= float(filters['max_amount']))
                except (ValueError, TypeError):
                    logger.warning(f"Invalid max_amount: {filters['max_amount']}")
        
        # Order by payment date (newest first)
        query = query.order_by(SupplierPayment.payment_date.desc())
        
        # Execute query
        results = query.all()
        
        # Create CSV
        output = io.StringIO()
        writer = csv.writer(output)
        
        # Write header
        headers = [
            'Payment Date',
            'Supplier Name',
            'Branch',
            'Reference Number',
            'Total Amount',
            'Cash Amount',
            'Cheque Amount',
            'Bank Transfer Amount',
            'UPI Amount',
            'Payment Method',
            'Status',
            'Notes',
            'Created Date',
            'Created By'
        ]
        writer.writerow(headers)
        
        # Write data rows
        for payment, supplier_name, branch_name in results:
            # FIXED: Determine primary payment method - improved logic
            payment_method = 'cash'  # Default
            active_methods = []

            if payment.cash_amount and payment.cash_amount > 0:
                active_methods.append('cash')
            if payment.cheque_amount and payment.cheque_amount > 0:
                active_methods.append('cheque')
            if payment.bank_transfer_amount and payment.bank_transfer_amount > 0:
                active_methods.append('bank_transfer')
            if payment.upi_amount and payment.upi_amount > 0:
                active_methods.append('upi')

            # FIXED: Only show 'mixed' if truly multiple methods
            if len(active_methods) == 1:
                payment_method = active_methods[0]
            elif len(active_methods) > 1:
                payment_method = 'mixed'
            # If no active methods (shouldn't happen), defaults to 'cash'
            
            row = [
                payment.payment_date.strftime('%Y-%m-%d') if payment.payment_date else '',
                supplier_name or '',
                branch_name or '',
                payment.reference_no or '',
                f"{float(payment.amount):.2f}" if payment.amount else '0.00',
                f"{float(payment.cash_amount):.2f}" if payment.cash_amount else '0.00',
                f"{float(payment.cheque_amount):.2f}" if payment.cheque_amount else '0.00',
                f"{float(payment.bank_transfer_amount):.2f}" if payment.bank_transfer_amount else '0.00',
                f"{float(payment.upi_amount):.2f}" if payment.upi_amount else '0.00',
                payment_method,
                payment.workflow_status or 'unknown',
                payment.notes or '',
                payment.created_at.strftime('%Y-%m-%d %H:%M:%S') if payment.created_at else '',
                payment.created_by or ''
            ]
            writer.writerow(row)
        
        csv_content = output.getvalue()
        output.close()
        
        logger.info(f"Export completed: {len(results)} payments exported")
        return csv_content
        
    except Exception as e:
        logger.error(f"Error exporting payments: {str(e)}", exc_info=True)
        raise

def log_document_access(
    document_id: uuid.UUID,
    user_id: str,
    access_type: str,
    hospital_id: uuid.UUID,
    session: Optional[Session] = None
) -> bool:
    """Log document access for audit purposes"""
    logger.info(f"Logging document access: {document_id} by {user_id}")
    
    if session is not None:
        return _log_document_access(session, document_id, user_id, access_type, hospital_id)
    
    with get_db_session() as new_session:
        return _log_document_access(new_session, document_id, user_id, access_type, hospital_id)

def _log_document_access(
    session: Session,
    document_id: uuid.UUID,
    user_id: str,
    access_type: str,
    hospital_id: uuid.UUID
) -> bool:
    """Internal function to log document access"""
    try:
        # Create access log entry
        access_log = PaymentDocumentAccessLog(
            document_id=document_id,
            user_id=user_id,
            access_type=access_type,
            hospital_id=hospital_id,
            accessed_at=datetime.now(timezone.utc),
            ip_address=None,  # Could be enhanced to capture IP
            user_agent=None   # Could be enhanced to capture user agent
        )
        
        session.add(access_log)
        session.flush()
        
        logger.info(f"Document access logged: {access_type} of {document_id} by {user_id}")
        return True
        
    except Exception as e:
        logger.error(f"Error logging document access: {str(e)}")
        return False


def approve_supplier_payment(
    payment_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: str,
    approval_notes: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """Approve supplier payment"""
    logger.info(f"Approving payment {payment_id}")
    
    if session is not None:
        return _approve_supplier_payment(session, payment_id, hospital_id, current_user_id, approval_notes)
    
    with get_db_session() as new_session:
        result = _approve_supplier_payment(new_session, payment_id, hospital_id, current_user_id, approval_notes)
        new_session.commit()
        return result


def _approve_supplier_payment(
    session: Session,
    payment_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: str,
    approval_notes: Optional[str] = None
) -> Dict:
    """Internal function to approve payment"""
    try:
        payment = session.query(SupplierPayment).filter_by(
            payment_id=payment_id,
            hospital_id=hospital_id
        ).first()
        
        if not payment:
            raise ValueError("Payment not found")
        
        if payment.workflow_status == 'approved':
            raise ValueError("Payment is already approved")
        
        # Update payment status
        payment.workflow_status = 'approved'
        payment.approved_by = current_user_id
        payment.approved_at = datetime.now(timezone.utc)
        
        if approval_notes:
            payment.approval_notes = approval_notes
        
        session.flush()
        
        # Create GL entries if not already created
        try:
            gl_result = create_supplier_payment_gl_entries(payment_id, current_user_id, session)
            logger.info("GL entries created after approval")
        except Exception as e:
            logger.warning(f"Error creating GL entries after approval: {str(e)}")
        
        return get_entity_dict(payment)
        
    except Exception as e:
        logger.error(f"Error approving payment: {str(e)}")
        session.rollback()
        raise


def reject_supplier_payment(
    payment_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: str,
    rejection_reason: str,
    session: Optional[Session] = None
) -> Dict:
    """Reject supplier payment"""
    logger.info(f"Rejecting payment {payment_id}")
    
    if session is not None:
        return _reject_supplier_payment(session, payment_id, hospital_id, current_user_id, rejection_reason)
    
    with get_db_session() as new_session:
        result = _reject_supplier_payment(new_session, payment_id, hospital_id, current_user_id, rejection_reason)
        new_session.commit()
        return result


def _reject_supplier_payment(
    session: Session,
    payment_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: str,
    rejection_reason: str
) -> Dict:
    """Internal function to reject payment"""
    try:
        payment = session.query(SupplierPayment).filter_by(
            payment_id=payment_id,
            hospital_id=hospital_id
        ).first()
        
        if not payment:
            raise ValueError("Payment not found")
        
        if payment.workflow_status == 'approved':
            raise ValueError("Cannot reject approved payment")
        
        # Update payment status
        payment.workflow_status = 'rejected'
        payment.rejected_by = current_user_id
        payment.rejected_at = datetime.now(timezone.utc)
        payment.rejection_reason = rejection_reason
        
        session.flush()
        
        return get_entity_dict(payment)
        
    except Exception as e:
        logger.error(f"Error rejecting payment: {str(e)}")
        session.rollback()
        raise


def update_supplier_payment(
    payment_id: uuid.UUID,
    payment_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: str,
    session: Optional[Session] = None
) -> Dict:
    """Update existing supplier payment"""
    logger.info(f"Updating payment {payment_id}")
    
    if session is not None:
        return _update_supplier_payment(session, payment_id, payment_data, hospital_id, current_user_id)
    
    with get_db_session() as new_session:
        result = _update_supplier_payment(new_session, payment_id, payment_data, hospital_id, current_user_id)
        new_session.commit()
        return result


def _update_supplier_payment(
    session: Session,
    payment_id: uuid.UUID,
    payment_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: str
) -> Dict:
    """Internal function to update payment"""
    try:
        payment = session.query(SupplierPayment).filter_by(
            payment_id=payment_id,
            hospital_id=hospital_id
        ).first()
        
        if not payment:
            raise ValueError("Payment not found")
        
        if payment.workflow_status == 'approved':
            raise ValueError("Cannot edit approved payment")
        
        # Validate payment data
        validation_result = validate_payment_data(payment_data, hospital_id, current_user_id)
        if not validation_result['is_valid']:
            raise ValueError(f"Payment validation failed: {'; '.join(validation_result['errors'])}")
        
        # Update payment fields
        payment.amount = Decimal(str(payment_data.get('amount', 0)))
        payment.payment_date = payment_data.get('payment_date')
        payment.cash_amount = Decimal(str(payment_data.get('cash_amount', 0)))
        payment.cheque_amount = Decimal(str(payment_data.get('cheque_amount', 0)))
        payment.bank_transfer_amount = Decimal(str(payment_data.get('bank_transfer_amount', 0)))
        payment.upi_amount = Decimal(str(payment_data.get('upi_amount', 0)))
        
        # Update method-specific details
        payment.cheque_number = payment_data.get('cheque_number')
        payment.cheque_date = payment_data.get('cheque_date')
        payment.cheque_bank = payment_data.get('cheque_bank')
        payment.bank_account_name = payment_data.get('bank_account_name')
        payment.bank_reference_number = payment_data.get('bank_reference_number')
        payment.upi_transaction_id = payment_data.get('upi_transaction_id')
        payment.upi_app_name = payment_data.get('upi_app_name')
        
        # Update other fields
        payment.reference_no = payment_data.get('reference_no')
        payment.notes = payment_data.get('notes')
        payment.updated_by = current_user_id
        payment.updated_at = datetime.now(timezone.utc)
        
        session.flush()
        
        # Handle new document uploads
        if payment_data.get('documents'):
            try:
                _process_payment_documents(payment.payment_id, payment_data['documents'], current_user_id, session)
            except Exception as e:
                logger.warning(f"Document processing failed: {str(e)}")
        
        return get_entity_dict(payment)
        
    except Exception as e:
        logger.error(f"Error updating payment: {str(e)}")
        session.rollback()
        raise

def refresh_entity_in_session(session: Session, entity_class, entity_id: uuid.UUID) -> Optional[object]:
    """
    Helper function to refresh entity in current session
    Prevents detached entity issues per developer guidelines
    """
    try:
        return session.query(entity_class).filter_by(
            **{f"{entity_class.__tablename__.rstrip('s')}_id": entity_id}
        ).first()
    except Exception as e:
        logger.error(f"Error refreshing entity in session: {str(e)}")
        return None
    
def get_current_invoice_balance(invoice_id: uuid.UUID, session: Session) -> float:
    """
    Calculate current balance due for an invoice based on approved payments
    """
    try:
        # Get invoice
        invoice = session.query(SupplierInvoice).filter_by(invoice_id=invoice_id).first()
        if not invoice:
            return 0.0
        
        # Calculate approved payments
        approved_payments = session.query(func.sum(SupplierPayment.amount)).filter_by(
            invoice_id=invoice_id,
            workflow_status='approved'
        ).scalar() or 0
        
        # Calculate remaining balance
        invoice_total = float(invoice.total_amount or 0)
        paid_amount = float(approved_payments)
        balance_due = invoice_total - paid_amount
        
        # Ensure balance is not negative
        return max(0.0, balance_due)
        
    except Exception as e:
        logger.error(f"Error calculating invoice balance: {str(e)}")
        return 0.0
    
def get_payment_documents(
    payment_id: uuid.UUID,
    hospital_id: uuid.UUID,
    session: Optional[Session] = None
) -> List[Dict]:
    """Get all documents for a payment"""
    logger.info(f"Getting documents for payment {payment_id}")
    
    if session is not None:
        return _get_payment_documents(session, payment_id, hospital_id)
    
    with get_db_session(read_only=True) as new_session:
        return _get_payment_documents(new_session, payment_id, hospital_id)

def _get_payment_documents(
    session: Session,
    payment_id: uuid.UUID,
    hospital_id: uuid.UUID
) -> List[Dict]:
    """Internal function to get payment documents"""
    try:
        # Verify payment exists and belongs to hospital
        payment = session.query(SupplierPayment).filter_by(
            payment_id=payment_id,
            hospital_id=hospital_id
        ).first()
        
        if not payment:
            logger.warning(f"Payment {payment_id} not found")
            return []
        
        # Get documents
        documents = session.query(PaymentDocument).filter_by(
            payment_id=payment_id
        ).order_by(PaymentDocument.updated_at.desc()).all()
        
        document_list = []
        for doc in documents:
            doc_dict = get_entity_dict(doc)
            document_list.append(doc_dict)
        
        logger.info(f"Found {len(document_list)} documents for payment {payment_id}")
        return document_list
        
    except Exception as e:
        logger.error(f"Error getting payment documents: {str(e)}")
        return []

def get_payment_document(
    payment_id: uuid.UUID,
    document_id: uuid.UUID,
    hospital_id: uuid.UUID,
    session: Optional[Session] = None
) -> Optional[Dict]:
    """Get specific payment document with content - FIXED SIGNATURE"""
    logger.info(f"Getting document {document_id} for payment {payment_id}")
    
    if session is not None:
        return _get_payment_document(session, payment_id, document_id, hospital_id)
    
    with get_db_session(read_only=True) as new_session:
        return _get_payment_document(new_session, payment_id, document_id, hospital_id)
    
def _get_payment_document(
    session: Session,
    payment_id: uuid.UUID,
    document_id: uuid.UUID,
    hospital_id: uuid.UUID
) -> Optional[Dict]:
    """Internal function to get payment document - FIXED VERSION"""
    try:
        # Verify payment exists and belongs to hospital
        payment = session.query(SupplierPayment).filter_by(
            payment_id=payment_id,
            hospital_id=hospital_id
        ).first()
        
        if not payment:
            logger.warning(f"Payment {payment_id} not found")
            return None
        
        # Get document
        document = session.query(PaymentDocument).filter_by(
            document_id=document_id,
            payment_id=payment_id
        ).first()
        
        if not document:
            logger.warning(f"Document {document_id} not found for payment {payment_id}")
            return None
        
        doc_dict = get_entity_dict(document)
        
        # FIX: Read file content from disk for download
        if hasattr(document, 'file_path') and document.file_path:
            try:
                import os
                if os.path.exists(document.file_path):
                    with open(document.file_path, 'rb') as f:
                        doc_dict['content'] = f.read()
                else:
                    logger.error(f"File not found: {document.file_path}")
                    doc_dict['content'] = None
            except Exception as e:
                logger.error(f"Error reading file {document.file_path}: {str(e)}")
                doc_dict['content'] = None
        
        # FIX: Use original_filename instead of filename
        logger.info(f"Retrieved document {document.original_filename} for payment {payment_id}")
        # Add file content reading for download
        if hasattr(document, 'file_path') and document.file_path:
            try:
                import os
                if os.path.exists(document.file_path):
                    with open(document.file_path, 'rb') as f:
                        doc_dict['content'] = f.read()
                else:
                    logger.error(f"File not found: {document.file_path}")
                    doc_dict['content'] = None
            except Exception as e:
                logger.error(f"Error reading file {document.file_path}: {str(e)}")
                doc_dict['content'] = None
        return doc_dict
        
    except Exception as e:
        logger.error(f"Error getting payment document: {str(e)}")
        return None
    
# ===================================================================
# AUTHORITATIVE PAYMENT CALCULATION - SINGLE SOURCE OF TRUTH
# Replace ALL existing calculation logic with this centralized approach
# ===================================================================

def _get_zero_payment_totals() -> Dict:
    """
    Return safe zero values if calculation fails
    Used as fallback when invoice not found or calculation errors
    """
    return {
        # Core Financial Data
        'invoice_total': 0.0,
        'gross_payments': 0.0,
        'credit_adjustments': 0.0,
        'net_payment': 0.0,
        'balance_due': 0.0,
        'payment_status': 'unpaid',
        
        # Backward Compatibility Fields
        'positive_payments_total': 0.0,
        'credit_adjustments_total': 0.0,
        'payment_total': 0.0,
        
        # Indicators
        'has_credit_notes': False,
        'is_fully_paid': False,
        'is_overpaid': False,
        
        # Audit Trail
        'calculation_method': 'error_fallback',
        'calculated_at': datetime.now(timezone.utc).isoformat(),
        'invoice_number': 'unknown'
    }

# ===================================================================
# COMPLETE CORRECTED calculate_invoice_payment_totals FUNCTION
# (with the missing helper function included)
# ===================================================================

def calculate_invoice_payment_totals(invoice_id: uuid.UUID, session: Session) -> Dict:
    """
    FIXED: Authoritative calculation with complete working credit note detection logic
    This is the single source of truth for payment calculations across the entire application
    """
    logger.debug(f"Calculating authoritative payment totals for invoice {invoice_id}")
    
    try:
        # Get invoice details
        invoice = session.query(SupplierInvoice).filter_by(invoice_id=invoice_id).first()
        if not invoice:
            logger.error(f"Invoice {invoice_id} not found")
            return _get_zero_payment_totals()
        
        # 1. POSITIVE PAYMENTS (Original payments made TO supplier)
        positive_payments = session.query(func.sum(SupplierPayment.amount)).filter_by(
            invoice_id=invoice_id,
            workflow_status='approved'
        ).filter(SupplierPayment.amount > 0).scalar() or 0
        
        logger.debug(f"Found positive payments: {positive_payments}")
        
        # 2. COMPLETE CREDIT ADJUSTMENTS DETECTION - COPIED FROM WORKING VERSION
        invoice_number = invoice.supplier_invoice_number
        hospital_id = invoice.hospital_id
        
        # Method A: Credit note invoices (from working logic)
        credit_note_invoices = session.query(SupplierInvoice).filter(
            SupplierInvoice.hospital_id == hospital_id,
            SupplierInvoice.is_credit_note == True,
            or_(
                SupplierInvoice.original_invoice_id == invoice_id,
                SupplierInvoice.notes.like(f'%{invoice_number}%'),
                SupplierInvoice.notes.like(f'%{invoice_id}%')
            )
        ).all()
        
        logger.debug(f"Found {len(credit_note_invoices)} credit note invoices")
        
        # Track unique credit payment IDs to avoid double-counting (working logic)
        found_credit_payment_ids = set()
        credit_payments_total = 0
        
        for credit_invoice in credit_note_invoices:
            logger.debug(f"Processing credit note: {credit_invoice.supplier_invoice_number}, Amount: {credit_invoice.total_amount}")
            
            # Get payments for this credit note
            credit_payments = session.query(SupplierPayment).filter_by(
                invoice_id=credit_invoice.invoice_id
            ).all()
            
            for cp in credit_payments:
                logger.debug(f"Credit payment: ID={cp.payment_id}, Amount={cp.amount}, Status={cp.workflow_status}, Ref={cp.reference_no}")
                
                # Only count each credit payment once (working logic)
                if cp.amount < 0 and str(cp.payment_id) not in found_credit_payment_ids:
                    credit_payments_total += abs(float(cp.amount))
                    found_credit_payment_ids.add(str(cp.payment_id))
        
        # Method B: Search by payment reference patterns (COMPLETE working logic)
        credit_payment_refs = session.query(SupplierPayment).filter(
            SupplierPayment.hospital_id == hospital_id,
            or_(
                # Original patterns
                SupplierPayment.reference_no.like(f'%CN-ADJ%{invoice_number}%'),
                SupplierPayment.reference_no.like(f'%CN-PAY%{invoice_number}%'),
                # CRITICAL: Add the missing pattern that your credit notes actually use
                SupplierPayment.reference_no.like(f'%CN-ADJ-PAY%{invoice_number}%'),
                # Search in notes field too
                SupplierPayment.notes.like(f'%{invoice_number}%'),
                SupplierPayment.notes.like(f'%{invoice_id}%')
            ),
            SupplierPayment.amount < 0
        ).all()
        
        logger.debug(f"Found {len(credit_payment_refs)} credit payments by reference pattern")
        
        for cp in credit_payment_refs:
            logger.debug(f"Ref credit payment: ID={cp.payment_id}, Amount={cp.amount}, Ref={cp.reference_no}")
            
            # Only count if not already found (working logic)
            if cp.amount < 0 and str(cp.payment_id) not in found_credit_payment_ids:
                credit_payments_total += abs(float(cp.amount))
                found_credit_payment_ids.add(str(cp.payment_id))
        
        # Method C: Direct credit payments (negative amounts on same invoice)
        direct_credit_payments = session.query(func.sum(SupplierPayment.amount)).filter_by(
            invoice_id=invoice_id,
            workflow_status='approved'
        ).filter(SupplierPayment.amount < 0).scalar() or 0
        direct_credit_total = abs(float(direct_credit_payments))
        
        logger.debug(f"Direct credit payments: {direct_credit_total}")
        
        # Take the maximum to avoid double-counting (but log all methods for debugging)
        total_credit_adjustments = max(credit_payments_total, direct_credit_total)
        
        logger.info(f"CREDIT DETECTION SUMMARY for {invoice.supplier_invoice_number}:")
        logger.info(f"  Method A+B (Working Logic): {credit_payments_total}")
        logger.info(f"  Method C (Direct): {direct_credit_total}")
        logger.info(f"  FINAL Credit Adjustments: {total_credit_adjustments}")
        
        # 3. CALCULATE FINAL AUTHORITATIVE TOTALS (working logic)
        positive_payments_total = float(positive_payments)
        credit_adjustments = total_credit_adjustments  # Use the working logic result
        effective_payment = positive_payments_total - credit_adjustments
        
        invoice_total = float(invoice.total_amount or 0)
        balance_due = max(0, invoice_total - effective_payment)
        
        # 4. DETERMINE AUTHORITATIVE PAYMENT STATUS (working logic)
        if invoice.payment_status == 'cancelled':
            authoritative_status = 'cancelled'
        elif effective_payment >= invoice_total:
            authoritative_status = 'paid'
        elif effective_payment > 0:
            authoritative_status = 'partial'
        else:
            authoritative_status = 'unpaid'
        
        # 5. RETURN AUTHORITATIVE CALCULATION RESULT
        result = {
            # Core Financial Data
            'invoice_total': invoice_total,
            'gross_payments': positive_payments_total,
            'credit_adjustments': credit_adjustments,
            'net_payment': effective_payment,
            'balance_due': balance_due,
            'payment_status': authoritative_status,
            
            # Backward Compatibility Fields (for existing code)
            'positive_payments_total': positive_payments_total,
            'credit_adjustments_total': credit_adjustments,
            'payment_total': effective_payment,
            
            # Indicators
            'has_credit_notes': credit_adjustments > 0,
            'is_fully_paid': effective_payment >= invoice_total,
            'is_overpaid': effective_payment > invoice_total,
            
            # Debug information for troubleshooting
            'credit_detection_details': {
                'working_logic_credits': credit_payments_total,
                'direct_credits': direct_credit_total,
                'credit_note_invoices_found': len(credit_note_invoices),
                'credit_payment_refs_found': len(credit_payment_refs)
            },
            
            # Audit Trail
            'calculation_method': 'authoritative_v3_with_working_logic',
            'calculated_at': datetime.now(timezone.utc).isoformat(),
            'invoice_number': invoice.supplier_invoice_number
        }
        
        logger.info(f"FIXED calculation for {invoice.supplier_invoice_number}: "
                   f"Gross={positive_payments_total}, Credits={credit_adjustments}, "
                   f"Net={effective_payment}, Balance={balance_due}, Status={authoritative_status}")
        
        return result
        
    except Exception as e:
        logger.error(f"Error in authoritative payment calculation: {str(e)}")
        return _get_zero_payment_totals()
    
def get_suppliers_for_choice(hospital_id: str) -> dict:
    """Get suppliers for dropdown choices"""
    try:
        # Use your existing supplier query logic here
        # This is just a placeholder - adapt to your actual code
        return {
            'success': True,
            'suppliers': []  # Replace with actual supplier query
        }
    except Exception as e:
        return {
            'success': False,
            'suppliers': [],
            'error': str(e)
        }
    
def search_supplier_payments_universal(
    hospital_id: uuid.UUID,
    filters: Dict = None,
    branch_id: Optional[uuid.UUID] = None,
    current_user_id: Optional[str] = None,
    page: int = 1,
    per_page: int = 20,
    session: Optional[Session] = None
) -> Dict:
    """
    Universal wrapper for supplier payments search
    Calls the enhanced universal implementation
    """
    try:
        from app.services.universal_supplier_service import EnhancedUniversalSupplierService
        
        service = EnhancedUniversalSupplierService()
        return service._search_supplier_payments_universal(
            hospital_id=hospital_id,
            filters=filters,
            branch_id=branch_id,
            current_user_id=current_user_id,
            page=page,
            per_page=per_page,
            session=session
        )
    except Exception as e:
        logger.error(f"Error in universal supplier payment search: {str(e)}")
        # Fallback to original method
        return search_supplier_payments(
            hospital_id=hospital_id,
            filters=filters,
            branch_id=branch_id,
            current_user_id=current_user_id,
            page=page,
            per_page=per_page,
            session=session
        )