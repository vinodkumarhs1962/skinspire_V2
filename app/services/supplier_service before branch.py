# app/services/supplier_service.py
# Part 1 of 4
from datetime import datetime, timezone, date
import uuid
from typing import Dict, List, Optional, Tuple, Union
from decimal import Decimal
import logging

from sqlalchemy import and_, or_, func, desc
from sqlalchemy.orm import Session
from sqlalchemy.exc import IntegrityError, SQLAlchemyError

from app.models.master import Supplier, Medicine, ChartOfAccounts, Hospital, Branch
from app.models.transaction import (
    PurchaseOrderHeader, PurchaseOrderLine,
    SupplierInvoice, SupplierInvoiceLine, SupplierPayment
)
from app.services.database_service import get_db_session, get_entity_dict
from app.services.gl_service import create_supplier_invoice_gl_entries, create_supplier_payment_gl_entries
from app.services.inventory_service import record_stock_from_supplier_invoice

logger = logging.getLogger(__name__)

# ===================================================================
# Supplier Management
# ===================================================================

def create_supplier(
    hospital_id: uuid.UUID,
    supplier_data: Dict,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Create a new supplier record
    
    Args:
        hospital_id: Hospital UUID
        supplier_data: Dictionary containing supplier information
        current_user_id: ID of the user creating the supplier
        session: Database session (optional)
        
    Returns:
        Dictionary containing created supplier details
    """
    logger.info(f"Creating supplier for hospital {hospital_id}")
    logger.debug(f"Supplier data: {supplier_data}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _create_supplier(session, hospital_id, supplier_data, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _create_supplier(new_session, hospital_id, supplier_data, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing supplier creation: {result.get('supplier_id')}")
        new_session.commit()
        
        logger.info(f"Successfully created supplier: {result.get('supplier_name')}")
        return result

def _create_supplier(
    session: Session,
    hospital_id: uuid.UUID,
    supplier_data: Dict,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to create a new supplier record within a session
    """
    try:
        logger.debug(f"Validating supplier name: {supplier_data.get('supplier_name')}")
        
        # Check for duplicate supplier name
        existing_supplier = session.query(Supplier).filter(
            Supplier.hospital_id == hospital_id,
            Supplier.supplier_name == supplier_data['supplier_name']
        ).first()
        
        if existing_supplier:
            error_msg = f"Supplier with name '{supplier_data['supplier_name']}' already exists"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Check for duplicate GST number if provided
        if supplier_data.get('gst_registration_number'):
            logger.debug(f"Validating GST number: {supplier_data.get('gst_registration_number')}")
            
            existing_gst = session.query(Supplier).filter(
                Supplier.hospital_id == hospital_id,
                Supplier.gst_registration_number == supplier_data['gst_registration_number']
            ).first()
            
            if existing_gst:
                error_msg = f"Supplier with GST number '{supplier_data['gst_registration_number']}' already exists"
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        logger.debug("Creating new supplier object")
        
        # Create new supplier
        supplier = Supplier(
            hospital_id=hospital_id,
            supplier_name=supplier_data['supplier_name'],
            supplier_category=supplier_data.get('supplier_category'),
            supplier_address=supplier_data.get('supplier_address', {}),
            contact_person_name=supplier_data.get('contact_person_name'),
            contact_info=supplier_data.get('contact_info', {}),
            manager_name=supplier_data.get('manager_name'),
            manager_contact_info=supplier_data.get('manager_contact_info', {}),
            email=supplier_data.get('email'),
            black_listed=supplier_data.get('black_listed', False),
            performance_rating=supplier_data.get('performance_rating'),
            payment_terms=supplier_data.get('payment_terms'),
            gst_registration_number=supplier_data.get('gst_registration_number'),
            pan_number=supplier_data.get('pan_number'),
            tax_type=supplier_data.get('tax_type'),
            state_code=supplier_data.get('state_code'),
            bank_details=supplier_data.get('bank_details', {}),
            remarks=supplier_data.get('remarks'),
            status=supplier_data.get('status', 'active')
        )
        
        if current_user_id:
            supplier.created_by = current_user_id
            logger.debug(f"Setting created_by to: {current_user_id}")
            
        session.add(supplier)
        session.flush()
        
        logger.info(f"Supplier created with ID: {supplier.supplier_id}")
        
        # Return the created supplier
        return get_entity_dict(supplier)
        
    except IntegrityError as e:
        logger.error(f"Database integrity error creating supplier: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error creating supplier: {str(e)}", exc_info=True)
        session.rollback()
        raise

def update_supplier(
    supplier_id: uuid.UUID,
    supplier_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Update an existing supplier record
    
    Args:
        supplier_id: Supplier UUID
        supplier_data: Dictionary containing supplier information updates
        hospital_id: Hospital UUID for security
        current_user_id: ID of the user updating the supplier
        session: Database session (optional)
        
    Returns:
        Dictionary containing updated supplier details
    """
    logger.info(f"Updating supplier {supplier_id} for hospital {hospital_id}")
    logger.debug(f"Update data: {supplier_data}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _update_supplier(session, supplier_id, supplier_data, hospital_id, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _update_supplier(new_session, supplier_id, supplier_data, hospital_id, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing supplier update: {supplier_id}")
        new_session.commit()
        
        logger.info(f"Successfully updated supplier: {result.get('supplier_name')}")
        return result

def _update_supplier(
    session: Session,
    supplier_id: uuid.UUID,
    supplier_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to update an existing supplier record within a session
    """
    try:
        logger.debug(f"Fetching supplier {supplier_id}")
        
        # Get the supplier
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Check for duplicate supplier name if being changed
        if 'supplier_name' in supplier_data and supplier_data['supplier_name'] != supplier.supplier_name:
            logger.debug(f"Checking for duplicate name: {supplier_data['supplier_name']}")
            
            existing_supplier = session.query(Supplier).filter(
                Supplier.hospital_id == hospital_id,
                Supplier.supplier_name == supplier_data['supplier_name'],
                Supplier.supplier_id != supplier_id
            ).first()
            
            if existing_supplier:
                error_msg = f"Supplier with name '{supplier_data['supplier_name']}' already exists"
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        # Check for duplicate GST number if being changed
        if 'gst_registration_number' in supplier_data and supplier_data['gst_registration_number'] != supplier.gst_registration_number:
            logger.debug(f"Checking for duplicate GST: {supplier_data['gst_registration_number']}")
            
            existing_gst = session.query(Supplier).filter(
                Supplier.hospital_id == hospital_id,
                Supplier.gst_registration_number == supplier_data['gst_registration_number'],
                Supplier.supplier_id != supplier_id
            ).first()
            
            if existing_gst:
                error_msg = f"Supplier with GST number '{supplier_data['gst_registration_number']}' already exists"
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        logger.debug("Updating supplier fields")
        
        # Update fields
        updatable_fields = [
            'supplier_name', 'supplier_category', 'supplier_address',
            'contact_person_name', 'contact_info', 'manager_name', 
            'manager_contact_info', 'email', 'black_listed', 
            'performance_rating', 'payment_terms', 'gst_registration_number',
            'pan_number', 'tax_type', 'state_code', 'bank_details',
            'remarks', 'status'
        ]
        
        for field in updatable_fields:
            if field in supplier_data:
                logger.debug(f"Updating field {field}: {supplier_data[field]}")
                setattr(supplier, field, supplier_data[field])
        
        if current_user_id:
            supplier.updated_by = current_user_id
            logger.debug(f"Setting updated_by to: {current_user_id}")
            
        # Update the timestamp
        supplier.updated_at = datetime.now(timezone.utc)
        
        session.flush()
        
        logger.info(f"Supplier {supplier_id} updated successfully")
        
        # Return the updated supplier
        return get_entity_dict(supplier)
        
    except IntegrityError as e:
        logger.error(f"Database integrity error updating supplier: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error updating supplier: {str(e)}", exc_info=True)
        session.rollback()
        raise

def get_supplier_by_id(
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    session: Optional[Session] = None
) -> Dict:
    """
    Get a supplier by ID
    
    Args:
        supplier_id: Supplier UUID
        hospital_id: Hospital UUID for security
        session: Database session (optional)
        
    Returns:
        Dictionary containing supplier details
    """
    logger.info(f"Fetching supplier {supplier_id} for hospital {hospital_id}")
    
    if session is not None:
        return _get_supplier_by_id(session, supplier_id, hospital_id)
    
    with get_db_session() as new_session:
        return _get_supplier_by_id(new_session, supplier_id, hospital_id)

def _get_supplier_by_id(
    session: Session,
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID
) -> Dict:
    """
    Internal function to get a supplier by ID within a session
    """
    try:
        logger.debug(f"Querying supplier {supplier_id}")
        
        # Get the supplier
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.info(f"Found supplier: {supplier.supplier_name}")
        
        # Return the supplier as a dictionary
        return get_entity_dict(supplier)
        
    except Exception as e:
        logger.error(f"Error getting supplier: {str(e)}", exc_info=True)
        raise

def search_suppliers(
    hospital_id: uuid.UUID,
    name: Optional[str] = None,
    category: Optional[str] = None,
    gst_number: Optional[str] = None,
    status: Optional[str] = None,
    blacklisted: Optional[bool] = None,
    page: int = 1,
    per_page: int = 20,
    session: Optional[Session] = None
) -> Dict:
    """
    Search suppliers with filtering and pagination
    
    Args:
        hospital_id: Hospital UUID
        name: Filter by supplier name
        category: Filter by supplier category
        gst_number: Filter by GST number
        status: Filter by status
        blacklisted: Filter by blacklisted status
        page: Page number for pagination
        per_page: Number of items per page
        session: Database session (optional)
        
    Returns:
        Dictionary containing supplier list and pagination info
    """
    logger.info(f"Searching suppliers for hospital {hospital_id}")
    logger.debug(f"Filters: name={name}, category={category}, gst={gst_number}, status={status}, blacklisted={blacklisted}")
    logger.debug(f"Pagination: page={page}, per_page={per_page}")
    
    if session is not None:
        return _search_suppliers(
            session, hospital_id, name, category, gst_number, 
            status, blacklisted, page, per_page
        )
    
    with get_db_session() as new_session:
        return _search_suppliers(
            new_session, hospital_id, name, category, gst_number, 
            status, blacklisted, page, per_page
        )

def _search_suppliers(
    session: Session,
    hospital_id: uuid.UUID,
    name: Optional[str] = None,
    category: Optional[str] = None,
    gst_number: Optional[str] = None,
    status: Optional[str] = None,
    blacklisted: Optional[bool] = None,
    page: int = 1,
    per_page: int = 20
) -> Dict:
    """
    Internal function to search suppliers within a session
    """
    try:
        logger.debug("Building supplier search query")
        
        # Base query
        query = session.query(Supplier).filter_by(hospital_id=hospital_id)
        
        # Apply filters
        if name:
            logger.debug(f"Filtering by name: {name}")
            query = query.filter(Supplier.supplier_name.ilike(f"%{name}%"))
            
        if category:
            logger.debug(f"Filtering by category: {category}")
            query = query.filter(Supplier.supplier_category == category)
            
        if gst_number:
            logger.debug(f"Filtering by GST: {gst_number}")
            query = query.filter(Supplier.gst_registration_number.ilike(f"%{gst_number}%"))
            
        if status:
            logger.debug(f"Filtering by status: {status}")
            query = query.filter(Supplier.status == status)
            
        if blacklisted is not None:
            logger.debug(f"Filtering by blacklisted: {blacklisted}")
            query = query.filter(Supplier.black_listed == blacklisted)
            
        # Count total for pagination
        total_count = query.count()
        logger.debug(f"Total suppliers found: {total_count}")
        
        # Apply pagination
        query = query.order_by(Supplier.supplier_name)
        query = query.offset((page - 1) * per_page).limit(per_page)
        
        # Execute query
        suppliers = query.all()
        logger.debug(f"Retrieved {len(suppliers)} suppliers for current page")
        
        # Convert to dictionaries
        supplier_list = [get_entity_dict(supplier) for supplier in suppliers]
        
        # Prepare pagination info
        pagination = {
            'page': page,
            'per_page': per_page,
            'total_count': total_count,
            'total_pages': (total_count + per_page - 1) // per_page
        }
        
        logger.info(f"Search completed: {len(supplier_list)} suppliers returned")
        
        return {
            'suppliers': supplier_list,
            'pagination': pagination
        }
        
    except Exception as e:
        logger.error(f"Error searching suppliers: {str(e)}", exc_info=True)
        raise

# app/services/supplier_service.py
# Part 2a of 2

# ===================================================================
# Purchase Order Management
# ===================================================================

def validate_po_line_item(line_data: Dict) -> List[str]:
    """
    Enhanced validation - focus on business rules for free and regular items
    
    Args:
        line_data: Dictionary containing line item data
        
    Returns:
        List of validation error messages
    """
    errors = []
    
    try:
        # Extract values safely with proper type conversion
        quantity = float(line_data.get('units', 0))
        unit_rate = float(line_data.get('pack_purchase_price', 0))
        unit_mrp = float(line_data.get('pack_mrp', 0))
        discount_percent = float(line_data.get('discount_percent', 0))
        is_free_item = line_data.get('is_free_item', False)
        
        # Convert string boolean to actual boolean if needed
        if isinstance(is_free_item, str):
            is_free_item = is_free_item.lower() in ('true', '1', 'yes')
        
        # CRITICAL: Quantity validation applies to ALL items (including free items)
        if quantity <= 0:
            errors.append("Quantity must be greater than 0")
            
        # Discount validation
        if discount_percent > 100 or discount_percent < 0:
            errors.append("Discount must be between 0 and 100%")
        
        # Free item specific validation
        if is_free_item:
            # Free items: zero rate, zero discount, but quantity > 0 is allowed
            if unit_rate > 0:
                errors.append("Free items should have zero rate")
            if discount_percent > 0:
                errors.append("Free items cannot have discount")
            # Note: MRP can be non-zero for free items (for reference purposes)
            
        else:
            # Non-free items: must have positive rate
            if unit_rate <= 0:
                errors.append("Rate must be greater than 0 for non-free items")
            
            # MRP validation for non-free items (optional but recommended)
            if unit_mrp > 0 and unit_rate > 0:
                # Check if rate + GST doesn't exceed MRP (basic business rule)
                gst_rate = float(line_data.get('gst_rate', 0))
                discounted_rate = unit_rate * (1 - discount_percent / 100)
                
                # Simple check: if discounted rate exceeds MRP, warn
                if discounted_rate > unit_mrp:
                    difference = discounted_rate - unit_mrp
                    errors.append(
                        f"Purchase rate ( Rs.{discounted_rate:.2f}) exceeds MRP ( Rs.{unit_mrp:.2f}) by  Rs.{difference:.2f}"
                    )
        
        # Medicine ID validation
        medicine_id = line_data.get('medicine_id')
        if not medicine_id:
            errors.append("Medicine selection is required")
        
        # Log validation for debugging free items
        if is_free_item:
            logger.debug(f"Validating free item: Qty={quantity}, Rate={unit_rate}, Errors={len(errors)}")
        
    except (ValueError, TypeError) as e:
        errors.append(f"Invalid data format: {str(e)}")
        logger.error(f"Validation error for line item: {str(e)}")
        
    return errors

# REPLACE your existing create_purchase_order function with this enhanced version:

def create_purchase_order(
    hospital_id: uuid.UUID,
    po_data: Dict,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None,
    skip_validation: bool = False  # NEW: Backward compatibility flag
) -> Dict:
    """
    Create a new purchase order with optional server-side validation
    
    Args:
        hospital_id: Hospital UUID
        po_data: Dictionary containing PO header and line items
        current_user_id: ID of the user creating the PO
        session: Database session (optional)
        skip_validation: Skip server-side validation (for backward compatibility)
        
    Returns:
        Dictionary containing created PO details
    """
    logger.info(f"Creating purchase order for hospital {hospital_id}")
    logger.debug(f"PO data: {po_data}")
    
    # Server-side validation (unless skipped for backward compatibility)
    if not skip_validation:
        logger.debug("Performing server-side validation")
        
        # Validate line items
        line_items = po_data.get('line_items', [])
        if not line_items:
            error_msg = "At least one line item is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Validate each line item
        validation_errors = []
        for idx, item_data in enumerate(line_items):
            line_errors = validate_po_line_item(item_data)
            if line_errors:
                for error in line_errors:
                    validation_errors.append(f"Line {idx + 1}: {error}")
        
        # If validation errors exist, raise exception
        if validation_errors:
            error_msg = "Purchase order validation failed:\n" + "\n".join(validation_errors)
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.debug("Server-side validation passed")
    else:
        logger.debug("Skipping server-side validation (backward compatibility mode)")
    
    # Proceed with PO creation using existing logic
    if session is not None:
        logger.debug("Using provided session")
        return _create_purchase_order(session, hospital_id, po_data, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _create_purchase_order(new_session, hospital_id, po_data, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing purchase order creation: {result.get('po_number')}")
        new_session.commit()
        
        logger.info(f"Successfully created purchase order: {result.get('po_number')}")
        return result


# OPTIONAL: Add a validation-enabled wrapper function for explicit validation
def create_purchase_order_with_validation(
    hospital_id: uuid.UUID,
    po_data: Dict,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Create a new purchase order with mandatory server-side validation
    
    This function explicitly enables validation and cannot be bypassed.
    Use this for new implementations where validation is required.
    """
    return create_purchase_order(
        hospital_id=hospital_id,
        po_data=po_data,
        current_user_id=current_user_id,
        session=session,
        skip_validation=False  # Force validation
    )


# OPTIONAL: Add a legacy wrapper function for old code
def create_purchase_order_legacy(
    hospital_id: uuid.UUID,
    po_data: Dict,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Create a new purchase order without validation (legacy mode)
    
    This function skips validation for backward compatibility.
    Use only for existing code that needs to be updated gradually.
    """
    return create_purchase_order(
        hospital_id=hospital_id,
        po_data=po_data,
        current_user_id=current_user_id,
        session=session,
        skip_validation=True  # Skip validation
    )

# def create_purchase_order(
#     hospital_id: uuid.UUID,
#     po_data: Dict,
#     current_user_id: Optional[str] = None,
#     session: Optional[Session] = None
# ) -> Dict:
#     """
#     Create a new purchase order
    
#     Args:
#         hospital_id: Hospital UUID
#         po_data: Dictionary containing PO header and line items
#         current_user_id: ID of the user creating the PO
#         session: Database session (optional)
        
#     Returns:
#         Dictionary containing created PO details
#     """
#     logger.info(f"Creating purchase order for hospital {hospital_id}")
#     logger.debug(f"PO data: {po_data}")
    
#     if session is not None:
#         logger.debug("Using provided session")
#         return _create_purchase_order(session, hospital_id, po_data, current_user_id)
    
#     logger.debug("Creating new session with explicit commit")
#     with get_db_session() as new_session:
#         result = _create_purchase_order(new_session, hospital_id, po_data, current_user_id)
        
#         # Add explicit commit for this critical operation
#         logger.info(f"Committing purchase order creation: {result.get('po_number')}")
#         new_session.commit()
        
#         logger.info(f"Successfully created purchase order: {result.get('po_number')}")
#         return result


def _create_purchase_order(
    session: Session,
    hospital_id: uuid.UUID,
    po_data: Dict,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to create a purchase order within a session - FIXED VERSION
    """
    try:
        # Validate supplier
        supplier_id = po_data.get('supplier_id')
        if not supplier_id:
            error_msg = "Supplier ID is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Validating supplier {supplier_id}")
        
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Check if supplier is blacklisted
        if supplier.black_listed:
            error_msg = f"Supplier '{supplier.supplier_name}' is blacklisted"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Validate line items
        line_items = po_data.get('line_items', [])
        if not line_items:
            error_msg = "At least one line item is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Processing {len(line_items)} line items")
            
        # Generate PO number
        po_number = _generate_po_number(session, hospital_id)
        logger.info(f"Generated PO number: {po_number}")
        
        # Get hospital state code for interstate calculation
        from app.models.master import Hospital
        hospital = session.query(Hospital).filter_by(hospital_id=hospital_id).first()
        hospital_state_code = hospital.state_code if hospital else None
        supplier_state_code = supplier.state_code
        is_interstate = (hospital_state_code != supplier_state_code) if (hospital_state_code and supplier_state_code) else False
        
        logger.debug(f"Interstate transaction: {is_interstate} (Hospital: {hospital_state_code}, Supplier: {supplier_state_code})")
        
        # Create PO header
        po_header = PurchaseOrderHeader(
            hospital_id=hospital_id,
            po_number=po_number,
            po_date=po_data.get('po_date', datetime.now(timezone.utc)),
            supplier_id=supplier_id,
            quotation_id=po_data.get('quotation_id'),
            quotation_date=po_data.get('quotation_date'),
            expected_delivery_date=po_data.get('expected_delivery_date'),
            currency_code=po_data.get('currency_code', 'INR'),
            exchange_rate=po_data.get('exchange_rate', 1.0),
            status=po_data.get('status', 'draft'),  # Allow status to be set from data
            deleted_flag=False
        )
        
        # Add additional fields from po_data
        if po_data.get('delivery_instructions'):
            po_header.delivery_instructions = po_data.get('delivery_instructions')
        if po_data.get('terms_conditions'):
            po_header.terms_conditions = po_data.get('terms_conditions')
        if po_data.get('notes'):
            po_header.notes = po_data.get('notes')
        
        if current_user_id:
            po_header.created_by = current_user_id
            logger.debug(f"Setting created_by to: {current_user_id}")
            
        session.add(po_header)
        session.flush()  # To get the po_id
        
        logger.debug(f"Created PO header with ID: {po_header.po_id}")
        
        # Create PO lines using standardized GST calculation
        po_lines = []
        total_amount = Decimal('0')
        
        for idx, item_data in enumerate(line_items):
            logger.debug(f"Processing line item {idx + 1}")
            
            # Validate medicine
            medicine_id = item_data.get('medicine_id')
            if not medicine_id:
                error_msg = f"Medicine ID is required for line item {idx + 1}"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            medicine = session.query(Medicine).filter_by(
                medicine_id=medicine_id,
                hospital_id=hospital_id
            ).first()
            
            if not medicine:
                error_msg = f"Medicine with ID {medicine_id} not found"
                logger.error(error_msg)
                raise ValueError(error_msg)
            
            # Extract line item data using business-friendly terms
            quantity = float(item_data.get('units', 0))  # Purchase quantity
            unit_rate = float(item_data.get('pack_purchase_price', 0))  # Vendor's rate per unit
            unit_mrp = float(item_data.get('pack_mrp', 0))  # MRP per unit
            conversion_factor = float(item_data.get('units_per_pack', 1))  # Sub-units per unit
            discount_percent = float(item_data.get('discount_percent', 0))
            is_free_item = item_data.get('is_free_item', False)
            
            # Get GST rate from medicine master or override from item data
            gst_rate = float(item_data.get('gst_rate', medicine.gst_rate or 0))
            
            logger.debug(f"Line {idx + 1}: Qty={quantity}, Rate={unit_rate}, GST={gst_rate}%, Discount={discount_percent}%, Free={is_free_item}")
            
            # Calculate all amounts using standardized function
            calculations = calculate_gst_values(
                quantity=quantity,
                unit_rate=unit_rate,
                gst_rate=gst_rate,
                discount_percent=discount_percent,
                is_free_item=is_free_item,
                is_interstate=is_interstate,
                conversion_factor=conversion_factor
            )
            
            logger.debug(f"Line {idx + 1} calculations completed")
            
            # Create PO line object with all required fields
            po_line = PurchaseOrderLine(
                hospital_id=hospital_id,
                po_id=po_header.po_id,
                medicine_id=medicine_id,
                medicine_name=medicine.medicine_name,
                pack_mrp=unit_mrp,
                hsn_code=medicine.hsn_code or item_data.get('hsn_code', ''),
                expected_delivery_date=item_data.get('expected_delivery_date', po_header.expected_delivery_date)
            )
            
            # Apply standardized GST calculations to the line object
            apply_gst_calculations_to_line(po_line, calculations)
            
            if current_user_id:
                po_line.created_by = current_user_id
                
            session.add(po_line)
            po_lines.append(po_line)
            
            # Add to total (using the calculated line total)
            total_amount += Decimal(str(po_line.line_total))
            
            logger.debug(f"Line {idx + 1} stored: Rate={po_line.unit_rate}, Discounted={po_line.discounted_rate}, GST={po_line.total_gst}, Total={po_line.line_total}")
            
        # Update header total amount
        po_header.total_amount = total_amount
        logger.info(f"PO total amount: {total_amount}")
        
        session.flush()
        
        # Return the created PO with lines
        result = get_entity_dict(po_header)
        result['line_items'] = [get_entity_dict(line) for line in po_lines]
        result['supplier_name'] = supplier.supplier_name
        
        logger.info(f"Successfully created PO {po_number} with {len(po_lines)} line items")
        
        return result
        
    except IntegrityError as e:
        logger.error(f"Database integrity error creating purchase order: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error creating purchase order: {str(e)}", exc_info=True)
        session.rollback()
        raise

def _generate_po_number(session: Session, hospital_id: uuid.UUID) -> str:
    """
    Generate a sequential purchase order number
    
    Args:
        session: Database session
        hospital_id: Hospital UUID
        
    Returns:
        Formatted PO number
    """
    try:
        # Get current financial year
        today = datetime.now()
        if today.month >= 4:  # Financial year starts in April
            fin_year_start = today.year
            fin_year_end = today.year + 1
        else:
            fin_year_start = today.year - 1
            fin_year_end = today.year
            
        fin_year = f"{fin_year_start}-{fin_year_end}"
        logger.debug(f"Generating PO number for financial year: {fin_year}")
        
        # Get the last PO number for this hospital and financial year
        last_po = session.query(PurchaseOrderHeader).filter(
            PurchaseOrderHeader.hospital_id == hospital_id,
            PurchaseOrderHeader.po_number.like(f"PO/{fin_year}/%")
        ).order_by(PurchaseOrderHeader.po_number.desc()).first()
        
        if last_po:
            # Extract the sequence number from the last PO number
            try:
                seq_num = int(last_po.po_number.split('/')[-1])
                new_seq_num = seq_num + 1
                logger.debug(f"Last PO sequence: {seq_num}, new sequence: {new_seq_num}")
            except:
                # If parsing fails, start from 1
                new_seq_num = 1
                logger.warning("Failed to parse last PO number, starting from 1")
        else:
            # No existing PO for this hospital and financial year
            new_seq_num = 1
            logger.debug("No existing PO for this year, starting from 1")
            
        # Format the new PO number
        po_number = f"PO/{fin_year}/{new_seq_num:05d}"
        logger.info(f"Generated PO number: {po_number}")
        
        return po_number
        
    except Exception as e:
        logger.error(f"Error generating PO number: {str(e)}", exc_info=True)
        raise

# app/services/supplier_service.py
# Part 2b of 2

def update_purchase_order_status(
    po_id: uuid.UUID,
    status: str,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Update the status of a purchase order
    
    Args:
        po_id: Purchase Order UUID
        status: New status (draft, approved, received, cancelled)
        hospital_id: Hospital UUID for security
        current_user_id: ID of the user updating the status
        session: Database session (optional)
        
    Returns:
        Dictionary containing updated PO details
    """
    logger.info(f"Updating PO {po_id} status to {status}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _update_purchase_order_status(session, po_id, status, hospital_id, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _update_purchase_order_status(new_session, po_id, status, hospital_id, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing PO status update: {po_id}")
        new_session.commit()
        
        logger.info(f"Successfully updated PO status to: {status}")
        return result

def _update_purchase_order_status(
    session: Session,
    po_id: uuid.UUID,
    status: str,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to update the status of a purchase order within a session
    """
    try:
        # Validate status
        valid_statuses = ['draft', 'approved', 'received', 'cancelled']
        if status not in valid_statuses:
            error_msg = f"Invalid status '{status}'. Must be one of: {', '.join(valid_statuses)}"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Fetching PO {po_id}")
            
        # Get the PO
        po = session.query(PurchaseOrderHeader).filter_by(
            po_id=po_id,
            hospital_id=hospital_id
        ).first()
        
        if not po:
            error_msg = f"Purchase order with ID {po_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Check if status transition is allowed
        current_status = po.status
        logger.debug(f"Current status: {current_status}, new status: {status}")
        
        # Rules for status transitions
        if current_status == 'cancelled':
            error_msg = "Cannot change status of a cancelled purchase order"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        if current_status == 'received' and status != 'cancelled':
            error_msg = "Received purchase orders can only be cancelled"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        if current_status == 'approved' and status == 'draft':
            error_msg = "Cannot change status from approved to draft"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Update the status
        po.status = status
        
        if status == 'approved':
            po.approved_by = current_user_id
            logger.debug(f"Setting approved_by to: {current_user_id}")
            
        if current_user_id:
            po.updated_by = current_user_id
            
        # Update the timestamp
        po.updated_at = datetime.now(timezone.utc)
        
        session.flush()
        
        logger.info(f"PO {po.po_number} status updated to {status}")
        
        # Return the updated PO with lines
        return get_po_with_lines(session, po)
        
    except Exception as e:
        logger.error(f"Error updating purchase order status: {str(e)}", exc_info=True)
        session.rollback()
        raise

def get_po_with_lines(session: Session, po: PurchaseOrderHeader) -> Dict:
    """
    Helper function to get a PO with its line items
    
    Args:
        session: Database session
        po: Purchase order header
        
    Returns:
        Dictionary containing PO details with line items
    """
    try:
        logger.debug(f"Getting PO {po.po_id} with line items")
        
        # Get the line items
        lines = session.query(PurchaseOrderLine).filter_by(po_id=po.po_id).all()
        
        logger.debug(f"Found {len(lines)} line items")
        
        # Get the supplier
        supplier = session.query(Supplier).filter_by(supplier_id=po.supplier_id).first()
        
        if not supplier:
            logger.warning(f"Supplier {po.supplier_id} not found for PO")
        
        # Convert to dictionary
        result = get_entity_dict(po)
        result['line_items'] = [get_entity_dict(line) for line in lines]
        
        if supplier:
            result['supplier_name'] = supplier.supplier_name
            
        return result
        
    except Exception as e:
        logger.error(f"Error getting PO with lines: {str(e)}", exc_info=True)
        raise

def get_purchase_order_by_id(
    po_id: uuid.UUID,
    hospital_id: uuid.UUID,
    session: Optional[Session] = None
) -> Dict:
    """
    Get a purchase order by ID with its line items
    
    Args:
        po_id: Purchase Order UUID
        hospital_id: Hospital UUID for security
        session: Database session (optional)
        
    Returns:
        Dictionary containing PO details with line items
    """
    logger.info(f"Fetching purchase order {po_id} for hospital {hospital_id}")
    
    if session is not None:
        return _get_purchase_order_by_id(session, po_id, hospital_id)
    
    with get_db_session() as new_session:
        return _get_purchase_order_by_id(new_session, po_id, hospital_id)

def _get_purchase_order_by_id(
    session: Session,
    po_id: uuid.UUID,
    hospital_id: uuid.UUID
) -> Dict:
    """
    Internal function to get a purchase order by ID within a session
    """
    try:
        logger.debug(f"Querying purchase order {po_id}")
        
        # Get the PO
        po = session.query(PurchaseOrderHeader).filter_by(
            po_id=po_id,
            hospital_id=hospital_id
        ).first()
        
        if not po:
            error_msg = f"Purchase order with ID {po_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.info(f"Found purchase order: {po.po_number}")
            
        # Return the PO with lines
        return get_po_with_lines(session, po)
        
    except Exception as e:
        logger.error(f"Error getting purchase order: {str(e)}", exc_info=True)
        raise

def search_purchase_orders(
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    po_number: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = 1,
    per_page: int = 20,
    session: Optional[Session] = None
) -> Dict:
    """
    Search purchase orders with filtering and pagination
    
    Args:
        hospital_id: Hospital UUID
        supplier_id: Filter by supplier ID
        po_number: Filter by PO number
        status: Filter by status
        start_date: Filter by start date
        end_date: Filter by end date
        page: Page number for pagination
        per_page: Number of items per page
        session: Database session (optional)
        
    Returns:
        Dictionary containing PO list and pagination info
    """
    logger.info(f"Searching purchase orders for hospital {hospital_id}")
    logger.debug(f"Filters: supplier={supplier_id}, po={po_number}, status={status}")
    logger.debug(f"Date range: {start_date} to {end_date}")
    logger.debug(f"Pagination: page={page}, per_page={per_page}")
    
    if session is not None:
        return _search_purchase_orders(
            session, hospital_id, supplier_id, po_number, 
            status, start_date, end_date, page, per_page
        )
    
    with get_db_session() as new_session:
        return _search_purchase_orders(
            new_session, hospital_id, supplier_id, po_number, 
            status, start_date, end_date, page, per_page
        )

def _search_purchase_orders(
    session: Session,
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    po_number: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = 1,
    per_page: int = 20
) -> Dict:
    """
    Internal function to search purchase orders within a session
    """
    try:
        logger.debug("Building purchase order search query")
        
        # Base query
        query = session.query(PurchaseOrderHeader).filter_by(hospital_id=hospital_id)
        
        # Apply filters
        if supplier_id:
            logger.debug(f"Filtering by supplier: {supplier_id}")
            query = query.filter(PurchaseOrderHeader.supplier_id == supplier_id)
            
        if po_number:
            logger.debug(f"Filtering by PO number: {po_number}")
            query = query.filter(PurchaseOrderHeader.po_number.ilike(f"%{po_number}%"))
            
        if status:
            logger.debug(f"Filtering by status: {status}")
            query = query.filter(PurchaseOrderHeader.status == status)
            
        if start_date:
            logger.debug(f"Filtering by start date: {start_date}")
            query = query.filter(PurchaseOrderHeader.po_date >= start_date)
            
        if end_date:
            logger.debug(f"Filtering by end date: {end_date}")
            query = query.filter(PurchaseOrderHeader.po_date <= end_date)
            
        # Count total for pagination
        total_count = query.count()
        logger.debug(f"Total purchase orders found: {total_count}")
        
        # Apply pagination
        query = query.order_by(desc(PurchaseOrderHeader.po_date))
        query = query.offset((page - 1) * per_page).limit(per_page)
        
        # Execute query
        pos = query.all()
        logger.debug(f"Retrieved {len(pos)} purchase orders for current page")
        
        # Convert to dictionaries
        po_list = []
        
        for po in pos:
            po_dict = get_entity_dict(po)
            
            # Get the supplier name
            supplier = session.query(Supplier).filter_by(supplier_id=po.supplier_id).first()
            if supplier:
                po_dict['supplier_name'] = supplier.supplier_name
                
            # Add line count
            line_count = session.query(PurchaseOrderLine).filter_by(po_id=po.po_id).count()
            po_dict['line_count'] = line_count
            
            # Get line items for GST calculation
            lines = session.query(PurchaseOrderLine).filter_by(po_id=po.po_id).all()
            line_items = [get_entity_dict(line) for line in lines]
            
            # Calculate GST totals using standardized function
            subtotal = 0
            total_gst = 0
            
            for line in line_items:
                # Calculate standardized GST values
                gst_calculations = calculate_gst_values(
                    price=line.get('pack_purchase_price', 0),
                    gst_rate=line.get('gst_rate', 0),
                    quantity=line.get('units', 1)
                )
                
                # Update running totals
                subtotal += gst_calculations['taxable_value']
                total_gst += gst_calculations['gst_amount']
            
            # Add calculated totals to PO dictionary
            po_dict['calculated_subtotal'] = subtotal
            po_dict['calculated_total_gst'] = total_gst
            po_dict['calculated_total_with_gst'] = subtotal + total_gst
            
            po_list.append(po_dict)
        
        # Prepare pagination info
        pagination = {
            'page': page,
            'per_page': per_page,
            'total_count': total_count,
            'total_pages': (total_count + per_page - 1) // per_page
        }
        
        logger.info(f"Search completed: {len(po_list)} purchase orders returned")
        
        return {
            'purchase_orders': po_list,
            'pagination': pagination
        }
        
    except Exception as e:
        logger.error(f"Error searching purchase orders: {str(e)}", exc_info=True)
        raise

def update_purchase_order(
    po_id: uuid.UUID,
    po_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Update an existing purchase order (draft status only)
    
    Args:
        po_id: Purchase Order UUID
        po_data: Dictionary containing PO updates (header + line items)
        hospital_id: Hospital UUID for security
        current_user_id: ID of the user updating the PO
        session: Database session (optional)
        
    Returns:
        Dictionary containing updated PO details with line items
    """
    logger.info(f"Updating purchase order {po_id}")
    
    if session is not None:
        return _update_purchase_order(session, po_id, po_data, hospital_id, current_user_id)
    
    with get_db_session() as new_session:
        result = _update_purchase_order(new_session, po_id, po_data, hospital_id, current_user_id)
        new_session.commit()
        logger.info(f"Successfully updated PO: {result.get('po_number')}")
        return result

def _update_purchase_order(
    session: Session,
    po_id: uuid.UUID,
    po_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    FIXED: Internal function to update purchase order within a session with safe field updates
    """
    try:
        # Get existing PO
        po = session.query(PurchaseOrderHeader).filter_by(
            po_id=po_id,
            hospital_id=hospital_id
        ).first()
        
        if not po:
            raise ValueError(f"Purchase order with ID {po_id} not found")
        
        # Only allow updating draft POs
        if po.status != 'draft':
            raise ValueError(f"Cannot update purchase order in '{po.status}' status. Only draft purchase orders can be edited.")
        
        logger.debug(f"Updating PO {po.po_number} from draft status")
        
        # FIXED: Safe field updates - only update fields that actually exist on the model
        updatable_fields = [
            'supplier_id', 'po_date', 'expected_delivery_date', 'quotation_id',
            'quotation_date', 'currency_code', 'exchange_rate', 'notes'
        ]
        
        # Add optional fields only if they exist on the model
        optional_fields = ['delivery_instructions', 'terms_conditions']
        for field in optional_fields:
            if hasattr(po, field):
                updatable_fields.append(field)
                logger.debug(f"Model has field: {field}")
            else:
                logger.warning(f"Model missing field: {field} - skipping")
        
        for field in updatable_fields:
            if field in po_data:
                old_value = getattr(po, field, None)
                new_value = po_data[field]
                if old_value != new_value:
                    setattr(po, field, new_value)
                    logger.debug(f"Updated {field}: {old_value} -> {new_value}")
        
        # Update metadata
        po.updated_by = current_user_id
        po.updated_at = datetime.now(timezone.utc)
        
        # CRITICAL: Delete existing line items with logging
        existing_lines = session.query(PurchaseOrderLine).filter_by(po_id=po_id).all()
        logger.debug(f"Found {len(existing_lines)} existing line items to delete")
        
        for line in existing_lines:
            logger.debug(f"Deleting line: {line.medicine_name}, Qty: {line.units}, Price: {line.pack_purchase_price}, Free: {getattr(line, 'is_free_item', 'N/A')}")
        
        deleted_count = session.query(PurchaseOrderLine).filter_by(po_id=po_id).delete()
        logger.debug(f"Deleted {deleted_count} existing line items")
        
        # Recreate line items with fresh GST calculations
        line_items = po_data.get('line_items', [])
        if not line_items:
            raise ValueError("At least one line item is required")
        
        logger.debug(f"Processing {len(line_items)} line items for update")
        
        # Get supplier and hospital for interstate calculation
        supplier = session.query(Supplier).filter_by(supplier_id=po.supplier_id).first()
        if not supplier:
            raise ValueError(f"Supplier with ID {po.supplier_id} not found")
            
        hospital = session.query(Hospital).filter_by(hospital_id=hospital_id).first()
        hospital_state_code = hospital.state_code if hospital else None
        supplier_state_code = supplier.state_code
        is_interstate = (hospital_state_code != supplier_state_code) if (
            hospital_state_code and supplier_state_code
        ) else False
        
        logger.debug(f"Interstate transaction: {is_interstate} (Hospital: {hospital_state_code}, Supplier: {supplier_state_code})")
        
        # Process line items with fresh calculations
        total_amount = Decimal('0')
        
        for idx, item_data in enumerate(line_items):
            logger.debug(f"Processing line item {idx + 1}")
            
            # Validate medicine
            medicine_id = item_data.get('medicine_id')
            if not medicine_id:
                raise ValueError(f"Medicine ID is required for line item {idx + 1}")
                
            medicine = session.query(Medicine).filter_by(
                medicine_id=medicine_id,
                hospital_id=hospital_id
            ).first()
            
            if not medicine:
                raise ValueError(f"Medicine with ID {medicine_id} not found")
            
            # Extract line item data with explicit logging
            quantity = float(item_data.get('units', 0))
            unit_rate = float(item_data.get('pack_purchase_price', 0))
            unit_mrp = float(item_data.get('pack_mrp', 0))
            conversion_factor = float(item_data.get('units_per_pack', 1))
            discount_percent = float(item_data.get('discount_percent', 0))
            is_free_item = item_data.get('is_free_item', False)
            gst_rate = float(item_data.get('gst_rate', medicine.gst_rate or 0))
            
            logger.debug(f"Line {idx + 1} INPUT: Medicine={medicine.medicine_name}, "
                        f"Qty={quantity}, Rate={unit_rate}, GST={gst_rate}%, "
                        f"Discount={discount_percent}%, Free={is_free_item}")
            
            # Calculate using standardized GST function
            calculations = calculate_gst_values(
                quantity=quantity,
                unit_rate=unit_rate,
                gst_rate=gst_rate,
                discount_percent=discount_percent,
                is_free_item=is_free_item,
                is_interstate=is_interstate,
                conversion_factor=conversion_factor
            )
            
            # Create new PO line object
            po_line = PurchaseOrderLine(
                hospital_id=hospital_id,
                po_id=po.po_id,
                medicine_id=medicine_id,
                medicine_name=medicine.medicine_name,
                pack_mrp=unit_mrp,
                hsn_code=medicine.hsn_code or item_data.get('hsn_code', ''),
                expected_delivery_date=item_data.get('expected_delivery_date', po.expected_delivery_date)
            )
            
            # Apply standardized GST calculations
            apply_gst_calculations_to_line(po_line, calculations)
            
            if current_user_id:
                po_line.created_by = current_user_id
                po_line.updated_by = current_user_id
                
            session.add(po_line)
            total_amount += Decimal(str(po_line.line_total))
            
            logger.debug(f"Line {idx + 1} STORED: Qty={po_line.units}, "
                        f"Rate={po_line.pack_purchase_price}, "
                        f"Taxable={po_line.taxable_amount}, "
                        f"GST={po_line.total_gst}, Total={po_line.line_total}, "
                        f"Free={getattr(po_line, 'is_free_item', 'N/A')}")
        
        # Update header total amount
        old_total = po.total_amount
        po.total_amount = total_amount
        logger.info(f"Updated PO total amount: {old_total} -> {total_amount}")
        
        session.flush()
        
        # Verify the update was successful by re-querying
        verification_lines = session.query(PurchaseOrderLine).filter_by(po_id=po_id).all()
        logger.debug(f"Verification: Created {len(verification_lines)} new line items")
        
        for i, line in enumerate(verification_lines):
            logger.debug(f"Verification Line {i+1}: {line.medicine_name}, "
                        f"Qty={line.units}, Rate={line.pack_purchase_price}, "
                        f"Total={line.line_total}")
        
        # Return updated PO with lines
        result = get_po_with_lines(session, po)
        
        logger.info(f"Successfully updated PO {po.po_number} with {len(line_items)} line items")
        
        return result
        
    except IntegrityError as e:
        logger.error(f"Database integrity error updating purchase order: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error updating purchase order: {str(e)}", exc_info=True)
        session.rollback()
        raise

# ===================================================================
# Supplier Invoice Management
# ===================================================================

def create_supplier_invoice(
    hospital_id: uuid.UUID,
    invoice_data: Dict,
    create_stock_entries: bool = True,
    create_gl_entries: bool = True,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Create a supplier invoice
    
    Args:
        hospital_id: Hospital UUID
        invoice_data: Dictionary containing invoice header and line items
        create_stock_entries: Whether to create inventory entries
        create_gl_entries: Whether to create GL entries
        current_user_id: ID of the user creating the invoice
        session: Database session (optional)
        
    Returns:
        Dictionary containing created invoice details
    """
    logger.info(f"Creating supplier invoice for hospital {hospital_id}")
    logger.debug(f"Invoice data: {invoice_data}")
    logger.debug(f"Stock entries: {create_stock_entries}, GL entries: {create_gl_entries}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _create_supplier_invoice(
            session, hospital_id, invoice_data, 
            create_stock_entries, create_gl_entries, current_user_id
        )
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _create_supplier_invoice(
            new_session, hospital_id, invoice_data, 
            create_stock_entries, create_gl_entries, current_user_id
        )
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing supplier invoice creation: {result.get('supplier_invoice_number')}")
        new_session.commit()
        
        logger.info(f"Successfully created supplier invoice: {result.get('supplier_invoice_number')}")
        return result

def _create_supplier_invoice(
    session: Session,
    hospital_id: uuid.UUID,
    invoice_data: Dict,
    create_stock_entries: bool = True,
    create_gl_entries: bool = True,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to create a supplier invoice within a session
    UPDATED: Uses centralized GST calculation service
    """
    try:
        # Validate supplier
        supplier_id = invoice_data.get('supplier_id')
        if not supplier_id:
            error_msg = "Supplier ID is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Validating supplier {supplier_id}")
        
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Check if supplier is blacklisted
        if supplier.black_listed:
            error_msg = f"Supplier '{supplier.supplier_name}' is blacklisted"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Validate line items
        line_items = invoice_data.get('line_items', [])
        if not line_items:
            error_msg = "At least one line item is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Processing {len(line_items)} line items")
            
        # Get PO if provided
        po = None
        po_id = invoice_data.get('po_id')
        if po_id:
            logger.debug(f"Validating purchase order {po_id}")
            
            po = session.query(PurchaseOrderHeader).filter_by(
                po_id=po_id,
                hospital_id=hospital_id
            ).first()
            
            if not po:
                error_msg = f"Purchase order with ID {po_id} not found"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            # Check if PO is cancelled
            if po.status == 'cancelled':
                error_msg = "Cannot create invoice for a cancelled purchase order"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            # Check if supplier matches - convert both to strings for comparison
            if str(po.supplier_id) != str(supplier_id):
                error_msg = "Invoice supplier does not match purchase order supplier"
                logger.error(f"Supplier mismatch: PO supplier={po.supplier_id}, Invoice supplier={supplier_id}")
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        # Create invoice header
        logger.debug("Creating invoice header")
        
        invoice = SupplierInvoice(
            hospital_id=hospital_id,
            po_id=po_id,
            supplier_id=supplier_id,
            supplier_invoice_number=invoice_data.get('supplier_invoice_number'),
            invoice_date=invoice_data.get('invoice_date', datetime.now(timezone.utc)),
            supplier_gstin=invoice_data.get('supplier_gstin', supplier.gst_registration_number),
            place_of_supply=invoice_data.get('place_of_supply', supplier.state_code),
            notes=invoice_data.get('notes'),
            reverse_charge=invoice_data.get('reverse_charge', False),
            currency_code=invoice_data.get('currency_code', 'INR'),
            exchange_rate=invoice_data.get('exchange_rate', 1.0),
            cgst_amount=Decimal('0'),
            sgst_amount=Decimal('0'),
            igst_amount=Decimal('0'),
            total_gst_amount=Decimal('0'),
            total_amount=Decimal('0'),
            itc_eligible=invoice_data.get('itc_eligible', True),
            payment_status='unpaid',
            due_date=invoice_data.get('due_date')
        )
        
        if current_user_id:
            invoice.created_by = current_user_id
            logger.debug(f"Setting created_by to: {current_user_id}")
            
        session.add(invoice)
        session.flush()  # To get the invoice_id
        
        logger.debug(f"Created invoice header with ID: {invoice.invoice_id}")
        
        # Get hospital state code for interstate calculation
        hospital = session.query(Hospital).filter_by(hospital_id=hospital_id).first()
        hospital_state_code = hospital.state_code if hospital else None
        supplier_state_code = supplier.state_code
        is_interstate = (hospital_state_code != supplier_state_code) if (hospital_state_code and supplier_state_code) else False
        
        logger.debug(f"Interstate transaction: {is_interstate} (Hospital: {hospital_state_code}, Supplier: {supplier_state_code})")
        
        # Create invoice lines using standardized GST calculation
        invoice_lines = []
        total_amount = Decimal('0')
        total_cgst = Decimal('0')
        total_sgst = Decimal('0')
        total_igst = Decimal('0')
        
        for idx, item_data in enumerate(line_items):
            logger.debug(f"Processing line item {idx + 1}")
            
            # Validate medicine
            medicine_id = item_data.get('medicine_id')
            if not medicine_id:
                error_msg = f"Medicine ID is required for line item {idx + 1}"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            medicine = session.query(Medicine).filter_by(
                medicine_id=medicine_id,
                hospital_id=hospital_id
            ).first()
            
            if not medicine:
                error_msg = f"Medicine with ID {medicine_id} not found"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            # Extract line item data with proper type conversion
            quantity = float(item_data.get('units', 0))
            unit_rate = float(item_data.get('pack_purchase_price', 0))
            unit_mrp = float(item_data.get('pack_mrp', 0))
            conversion_factor = float(item_data.get('units_per_pack', 1))
            discount_percent = float(item_data.get('discount_percent', 0))
            is_free_item = item_data.get('is_free_item', False)
            gst_rate = float(item_data.get('gst_rate', medicine.gst_rate or 0))
            
            logger.debug(f"Line {idx + 1}: Medicine={medicine.medicine_name}, "
                        f"Qty={quantity}, Rate={unit_rate}, GST={gst_rate}%, "
                        f"Discount={discount_percent}%, Free={is_free_item}")
            
            # Calculate using standardized GST function
            calculations = calculate_gst_values(
                quantity=quantity,
                unit_rate=unit_rate,
                gst_rate=gst_rate,
                discount_percent=discount_percent,
                is_free_item=is_free_item,
                is_interstate=is_interstate,
                conversion_factor=conversion_factor
            )
            
            logger.debug(f"Line {idx + 1} calculations completed")
            
            # Create invoice line object with all required fields
            invoice_line = SupplierInvoiceLine(
                hospital_id=hospital_id,
                invoice_id=invoice.invoice_id,
                medicine_id=medicine_id,
                medicine_name=medicine.medicine_name,
                pack_mrp=unit_mrp,
                hsn_code=medicine.hsn_code or item_data.get('hsn_code'),
                batch_number=item_data.get('batch_number'),
                manufacturing_date=item_data.get('manufacturing_date'),
                expiry_date=item_data.get('expiry_date'),
                itc_eligible=item_data.get('itc_eligible', invoice.itc_eligible)
            )
            
            # Apply standardized GST calculations to the line object
            apply_gst_calculations_to_invoice_line(invoice_line, calculations)
            
            if current_user_id:
                invoice_line.created_by = current_user_id
                
            session.add(invoice_line)
            invoice_lines.append(invoice_line)
            
            # Add to totals (using the calculated line total)
            total_amount += Decimal(str(invoice_line.line_total))
            total_cgst += Decimal(str(invoice_line.cgst))
            total_sgst += Decimal(str(invoice_line.sgst))
            total_igst += Decimal(str(invoice_line.igst))
            
            logger.debug(f"Line {idx + 1} stored: Rate={invoice_line.pack_purchase_price}, "
                        f"Taxable={invoice_line.taxable_amount}, GST={invoice_line.total_gst}, "
                        f"Total={invoice_line.line_total}")
            
        # Update invoice totals
        invoice.cgst_amount = total_cgst
        invoice.sgst_amount = total_sgst
        invoice.igst_amount = total_igst
        invoice.total_gst_amount = total_cgst + total_sgst + total_igst
        invoice.total_amount = total_amount
        
        logger.info(f"Invoice total amount: {total_amount}")
        
        session.flush()
        
        # Create stock entries if requested
        if create_stock_entries:
            logger.debug("Creating stock entries")
            try:
                stock_result = record_stock_from_supplier_invoice(
                    hospital_id=hospital_id,
                    invoice_id=invoice.invoice_id,
                    current_user_id=current_user_id,
                    session=session
                )
                logger.info("Stock entries created successfully")
            except Exception as e:
                logger.error(f"Error creating stock entries: {str(e)}")
                if not invoice_data.get('skip_stock_error', False):
                    raise
                logger.warning("Continuing despite stock entry error")
            
        # Create GL entries if requested
        gl_transaction = None
        if create_gl_entries:
            logger.debug("Creating GL entries")
            try:
                gl_result = create_supplier_invoice_gl_entries(
                    invoice.invoice_id,
                    current_user_id,
                    session
                )
                logger.info("GL entries created successfully")
            except Exception as e:
                logger.error(f"Error creating GL entries: {str(e)}")
                if not invoice_data.get('skip_gl_error', False):
                    raise
                logger.warning("Continuing despite GL entry error")
            
        # Return the created invoice with lines
        result = get_entity_dict(invoice)
        result['line_items'] = [get_entity_dict(line) for line in invoice_lines]
        result['supplier_name'] = supplier.supplier_name
        
        return result
        
    except IntegrityError as e:
        logger.error(f"Database integrity error creating supplier invoice: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error creating supplier invoice: {str(e)}", exc_info=True)
        session.rollback()
        raise


def update_supplier_invoice(
    invoice_id: uuid.UUID,
    invoice_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Update an existing supplier invoice (unpaid invoices only)
    
    Args:
        invoice_id: Supplier Invoice UUID
        invoice_data: Dictionary containing invoice updates (header + line items)
        hospital_id: Hospital UUID for security
        current_user_id: ID of the user updating the invoice
        session: Database session (optional)
        
    Returns:
        Dictionary containing updated invoice details with line items
    """
    logger.info(f"Updating supplier invoice {invoice_id}")
    
    if session is not None:
        return _update_supplier_invoice(session, invoice_id, invoice_data, hospital_id, current_user_id)
    
    with get_db_session() as new_session:
        result = _update_supplier_invoice(new_session, invoice_id, invoice_data, hospital_id, current_user_id)
        new_session.commit()
        logger.info(f"Successfully updated invoice: {result.get('supplier_invoice_number')}")
        return result

def _update_supplier_invoice(
    session: Session,
    invoice_id: uuid.UUID,
    invoice_data: Dict,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to update supplier invoice within a session with safe field updates
    """
    try:
        # Get existing invoice
        invoice = session.query(SupplierInvoice).filter_by(
            invoice_id=invoice_id,
            hospital_id=hospital_id
        ).first()
        
        if not invoice:
            raise ValueError(f"Supplier invoice with ID {invoice_id} not found")
        
        # Validate invoice can be edited
        if invoice.payment_status == 'cancelled':
            raise ValueError(f"Cannot update cancelled invoice {invoice.supplier_invoice_number}")
        
        if invoice.payment_status == 'paid':
            raise ValueError(f"Cannot update paid invoice {invoice.supplier_invoice_number}. Create a credit note instead.")
        
        # Check for existing payments
        existing_payments = session.query(SupplierPayment).filter_by(invoice_id=invoice_id).count()
        if existing_payments > 0:
            raise ValueError(f"Cannot update invoice {invoice.supplier_invoice_number} with recorded payments")
        
        logger.debug(f"Updating invoice {invoice.supplier_invoice_number}")
        
        # Update header fields safely
        updatable_fields = [
            'supplier_id', 'supplier_invoice_number', 'invoice_date', 'po_id',
            'supplier_gstin', 'place_of_supply', 'due_date', 'reverse_charge',
            'itc_eligible', 'currency_code', 'exchange_rate', 'notes'
        ]
        
        for field in updatable_fields:
            if field in invoice_data:
                old_value = getattr(invoice, field, None)
                new_value = invoice_data[field]
                if old_value != new_value:
                    setattr(invoice, field, new_value)
                    logger.debug(f"Updated {field}: {old_value} -> {new_value}")
        
        # Update metadata
        invoice.updated_by = current_user_id
        invoice.updated_at = datetime.now(timezone.utc)
        
        # Delete existing line items with logging
        existing_lines = session.query(SupplierInvoiceLine).filter_by(invoice_id=invoice_id).all()
        logger.debug(f"Found {len(existing_lines)} existing line items to delete")
        
        for line in existing_lines:
            logger.debug(f"Deleting line: {line.medicine_name}, Qty: {line.units}, Batch: {line.batch_number}")
        
        deleted_count = session.query(SupplierInvoiceLine).filter_by(invoice_id=invoice_id).delete()
        logger.debug(f"Deleted {deleted_count} existing line items")
        
        # Recreate line items with fresh GST calculations
        line_items = invoice_data.get('line_items', [])
        if not line_items:
            raise ValueError("At least one line item is required")
        
        logger.debug(f"Processing {len(line_items)} line items for update")
        
        # Get supplier and hospital for interstate calculation
        supplier = session.query(Supplier).filter_by(
            supplier_id=invoice.supplier_id,
            hospital_id=hospital_id
        ).first()
        if not supplier:
            raise ValueError(f"Supplier with ID {invoice.supplier_id} not found")
            
        hospital = session.query(Hospital).filter_by(hospital_id=hospital_id).first()
        hospital_state_code = hospital.state_code if hospital else None
        supplier_state_code = invoice_data.get('place_of_supply')  # Use form data for place of supply
        is_interstate = (hospital_state_code != supplier_state_code) if (
            hospital_state_code and supplier_state_code
        ) else False
        
        logger.debug(f"Interstate transaction: {is_interstate} (Hospital: {hospital_state_code}, Supplier: {supplier_state_code})")
        
        # Process line items with fresh calculations
        total_amount = Decimal('0')
        total_cgst = Decimal('0')
        total_sgst = Decimal('0')
        total_igst = Decimal('0')
        
        for idx, item_data in enumerate(line_items):
            logger.debug(f"Processing line item {idx + 1}")
            
            # Validate medicine
            medicine_id = item_data.get('medicine_id')
            if not medicine_id:
                raise ValueError(f"Medicine ID is required for line item {idx + 1}")
                
            medicine = session.query(Medicine).filter_by(
                medicine_id=medicine_id,
                hospital_id=hospital_id
            ).first()
            
            if not medicine:
                raise ValueError(f"Medicine with ID {medicine_id} not found")
            
            # Extract line item data with explicit logging
            quantity = float(item_data.get('units', 0))
            unit_rate = float(item_data.get('pack_purchase_price', 0))
            unit_mrp = float(item_data.get('pack_mrp', 0))
            conversion_factor = float(item_data.get('units_per_pack', 1))
            discount_percent = float(item_data.get('discount_percent', 0))
            is_free_item = item_data.get('is_free_item', False)
            gst_rate = float(item_data.get('gst_rate', medicine.gst_rate or 0))
            batch_number = item_data.get('batch_number', '')
            expiry_date = item_data.get('expiry_date')
            
            logger.debug(f"Line {idx + 1} INPUT: Medicine={medicine.medicine_name}, "
                        f"Qty={quantity}, Rate={unit_rate}, GST={gst_rate}%, "
                        f"Batch={batch_number}, Expiry={expiry_date}")
            
            # Calculate using standardized GST function
            calculations = calculate_gst_values(
                quantity=quantity,
                unit_rate=unit_rate,
                gst_rate=gst_rate,
                discount_percent=discount_percent,
                is_free_item=is_free_item,
                is_interstate=is_interstate,
                conversion_factor=conversion_factor
            )
            
            # Create new invoice line object
            invoice_line = SupplierInvoiceLine(
                hospital_id=hospital_id,
                invoice_id=invoice.invoice_id,
                medicine_id=medicine_id,
                medicine_name=medicine.medicine_name,
                pack_mrp=unit_mrp,
                hsn_code=medicine.hsn_code or item_data.get('hsn_code', ''),
                batch_number=batch_number,
                manufacturing_date=item_data.get('manufacturing_date'),
                expiry_date=expiry_date,
                itc_eligible=item_data.get('itc_eligible', invoice.itc_eligible)
            )
            
            # Apply standardized GST calculations
            apply_gst_calculations_to_invoice_line(invoice_line, calculations)
            
            if current_user_id:
                invoice_line.created_by = current_user_id
                invoice_line.updated_by = current_user_id
                
            session.add(invoice_line)
            
            # Add to totals
            total_amount += Decimal(str(invoice_line.line_total))
            total_cgst += Decimal(str(invoice_line.cgst))
            total_sgst += Decimal(str(invoice_line.sgst))
            total_igst += Decimal(str(invoice_line.igst))
            
            logger.debug(f"Line {idx + 1} STORED: Qty={invoice_line.units}, "
                        f"Rate={invoice_line.pack_purchase_price}, "
                        f"Batch={invoice_line.batch_number}, "
                        f"Total={invoice_line.line_total}")
        
        # Update invoice totals
        old_total = invoice.total_amount
        invoice.cgst_amount = total_cgst
        invoice.sgst_amount = total_sgst
        invoice.igst_amount = total_igst
        invoice.total_gst_amount = total_cgst + total_sgst + total_igst
        invoice.total_amount = total_amount
        
        logger.info(f"Updated invoice total amount: {old_total} -> {total_amount}")
        
        session.flush()
        
        # Verify the update was successful
        verification_lines = session.query(SupplierInvoiceLine).filter_by(invoice_id=invoice_id).all()
        logger.debug(f"Verification: Created {len(verification_lines)} new line items")
        
        # Return updated invoice with lines
        result = get_entity_dict(invoice)
        result['line_items'] = [get_entity_dict(line) for line in verification_lines]
        result['supplier_name'] = supplier.supplier_name
        
        logger.info(f"Successfully updated invoice {invoice.supplier_invoice_number} with {len(line_items)} line items")
        
        return result
        
    except IntegrityError as e:
        logger.error(f"Database integrity error updating supplier invoice: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error updating supplier invoice: {str(e)}", exc_info=True)
        session.rollback()
        raise


def get_supplier_invoice_by_id(
    invoice_id: uuid.UUID,
    hospital_id: uuid.UUID,
    include_payments: bool = True,
    session: Optional[Session] = None
) -> Dict:
    """
    Get a supplier invoice by ID with its line items
    
    Args:
        invoice_id: Supplier Invoice UUID
        hospital_id: Hospital UUID for security
        include_payments: Whether to include payment details
        session: Database session (optional)
        
    Returns:
        Dictionary containing invoice details with line items
    """
    logger.info(f"Fetching supplier invoice {invoice_id} for hospital {hospital_id}")
    logger.debug(f"Include payments: {include_payments}")
    
    if session is not None:
        return _get_supplier_invoice_by_id(session, invoice_id, hospital_id, include_payments)
    
    with get_db_session() as new_session:
        return _get_supplier_invoice_by_id(new_session, invoice_id, hospital_id, include_payments)

def _get_supplier_invoice_by_id(
    session: Session,
    invoice_id: uuid.UUID,
    hospital_id: uuid.UUID,
    include_payments: bool = True
) -> Dict:
    """
    Internal function to get a supplier invoice by ID within a session
    UPDATED: Uses stored GST values only, no recalculation
    """
    try:
        logger.debug(f"Querying supplier invoice {invoice_id}")
        
        # Get the invoice
        invoice = session.query(SupplierInvoice).filter_by(
            invoice_id=invoice_id,
            hospital_id=hospital_id
        ).first()
        
        if not invoice:
            error_msg = f"Supplier invoice with ID {invoice_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug("Fetching invoice line items")
            
        # Get the line items
        lines = session.query(SupplierInvoiceLine).filter_by(
            invoice_id=invoice_id
        ).all()
        
        logger.debug(f"Found {len(lines)} line items")
        
        # Get the supplier
        supplier = session.query(Supplier).filter_by(
            supplier_id=invoice.supplier_id
        ).first()
        
        if not supplier:
            logger.warning(f"Supplier {invoice.supplier_id} not found for invoice")
        
        # Convert to dictionary
        result = get_entity_dict(invoice)
        result['line_items'] = [get_entity_dict(line) for line in lines]
        
        if supplier:
            result['supplier_name'] = supplier.supplier_name
            result['supplier_address'] = supplier.supplier_address
            result['contact_info'] = supplier.contact_info
            
        # Use stored GST values only (no recalculation in view)
        subtotal = 0
        total_gst = 0
        
        for line in result['line_items']:
            # Use stored values directly from database
            line_subtotal = float(line.get('taxable_amount', 0))
            line_gst = float(line.get('total_gst', 0))
            line_total = float(line.get('line_total', 0))
            
            # Add display values using stored data
            line['subtotal'] = line_subtotal
            line['gst_amount'] = line_gst
            line['total_with_gst'] = line_total
            
            # Update running totals
            subtotal += line_subtotal
            total_gst += line_gst
        
        # Add totals to invoice dictionary using stored values
        result['calculated_subtotal'] = subtotal
        result['calculated_total_gst'] = total_gst
        result['calculated_total_with_gst'] = subtotal + total_gst
        
        # Log verification of stored vs calculated totals
        stored_total = float(result.get('total_amount', 0))
        calculated_total = subtotal + total_gst
        
        logger.debug(f"Invoice {invoice.supplier_invoice_number}: "
                    f"Stored total={stored_total}, Calculated total={calculated_total}")
        
        if abs(stored_total - calculated_total) > 0.01:
            logger.warning(f"Invoice total mismatch - using stored total as authoritative")
            
        # Include payments if requested
        if include_payments:
            logger.debug("Fetching payment details")
            
            payments = session.query(SupplierPayment).filter_by(
                invoice_id=invoice_id
            ).all()
            
            logger.debug(f"Found {len(payments)} payments")
            
            result['payments'] = [get_entity_dict(payment) for payment in payments]
            
        logger.info(f"Successfully retrieved invoice {invoice.supplier_invoice_number} using stored values")
            
        return result
        
    except Exception as e:
        logger.error(f"Error getting supplier invoice: {str(e)}", exc_info=True)
        raise
    
# Updated search_supplier_invoices function with po_id parameter

def search_supplier_invoices(
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    invoice_number: Optional[str] = None,
    po_id: Optional[uuid.UUID] = None,  # Add this parameter
    payment_status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = 1,
    per_page: int = 20,
    session: Optional[Session] = None
) -> Dict:
    """
    Search supplier invoices with filtering and pagination
    
    Args:
        hospital_id: Hospital UUID
        supplier_id: Filter by supplier ID
        invoice_number: Filter by invoice number
        po_id: Filter by purchase order ID  # Add this
        payment_status: Filter by payment status
        start_date: Filter by start date
        end_date: Filter by end date
        page: Page number for pagination
        per_page: Number of items per page
        session: Database session (optional)
        
    Returns:
        Dictionary containing invoice list and pagination info
    """
    logger.info(f"Searching supplier invoices for hospital {hospital_id}")
    logger.debug(f"Filters: supplier={supplier_id}, invoice={invoice_number}, po={po_id}, status={payment_status}")
    logger.debug(f"Date range: {start_date} to {end_date}")
    logger.debug(f"Pagination: page={page}, per_page={per_page}")
    
    if session is not None:
        return _search_supplier_invoices(
            session, hospital_id, supplier_id, invoice_number, po_id,
            payment_status, start_date, end_date, page, per_page
        )
    
    with get_db_session() as new_session:
        return _search_supplier_invoices(
            new_session, hospital_id, supplier_id, invoice_number, po_id,
            payment_status, start_date, end_date, page, per_page
        )

# UPDATED _search_supplier_invoices function (simplified):
def _search_supplier_invoices(
    session: Session,
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    po_id: Optional[uuid.UUID] = None,
    invoice_number: Optional[str] = None,
    start_date: Optional[date] = None,
    end_date: Optional[date] = None,
    status: Optional[str] = None,
    page: int = 1,
    per_page: int = 20
) -> Dict:
    """
    UPDATED: Search supplier invoices using stored values only (no GST recalculation)
    """
    try:
        logger.info(f"Searching supplier invoices for hospital {hospital_id}")
        
        # Build the query
        query = session.query(SupplierInvoice).filter_by(
            hospital_id=hospital_id,
        )
        
        # Apply filters
        if supplier_id:
            logger.debug(f"Filtering by supplier: {supplier_id}")
            query = query.filter(SupplierInvoice.supplier_id == supplier_id)
        
        if po_id:
            logger.debug(f"Filtering by purchase order: {po_id}")
            query = query.filter(SupplierInvoice.po_id == po_id)
        
        if invoice_number:
            logger.debug(f"Filtering by invoice number: {invoice_number}")
            query = query.filter(SupplierInvoice.supplier_invoice_number.ilike(f'%{invoice_number}%'))
        
        if status:
            logger.debug(f"Filtering by payment status: {status}")
            query = query.filter(SupplierInvoice.payment_status == status)
        
        if start_date:
            logger.debug(f"Filtering by start date: {start_date}")
            query = query.filter(SupplierInvoice.invoice_date >= start_date)
        
        if end_date:
            logger.debug(f"Filtering by end date: {end_date}")
            query = query.filter(SupplierInvoice.invoice_date <= end_date)
        
        # Get total count
        total = query.count()
        logger.debug(f"Total invoices found: {total}")
        
        # Apply pagination and get results
        invoices = query.order_by(SupplierInvoice.invoice_date.desc()).offset(
            (page - 1) * per_page
        ).limit(per_page).all()
        
        logger.debug(f"Retrieved {len(invoices)} invoices for current page")
        
        # Convert to dictionaries using stored values only (no GST recalculation)
        invoice_list = []
        for invoice in invoices:
            invoice_dict = get_entity_dict(invoice)
            
            # Add supplier name if available
            if hasattr(invoice, 'supplier') and invoice.supplier:
                invoice_dict['supplier_name'] = invoice.supplier.supplier_name
            else:
                # Fetch supplier name separately if not eagerly loaded
                supplier = session.query(Supplier).filter_by(
                    supplier_id=invoice.supplier_id
                ).first()
                if supplier:
                    invoice_dict['supplier_name'] = supplier.supplier_name
                else:
                    invoice_dict['supplier_name'] = 'Unknown Supplier'
            
            # Calculate payment information using stored totals
            payment_total = session.query(func.sum(SupplierPayment.amount)).filter_by(
                invoice_id=invoice.invoice_id
            ).scalar() or 0
            
            payment_total = float(payment_total)
            total_amount = float(invoice.total_amount or 0)
            balance_due = total_amount - payment_total
            
            # Add payment summary using stored invoice total
            invoice_dict['payment_total'] = payment_total
            invoice_dict['balance_due'] = balance_due
            
            # Use stored total_amount (no recalculation)
            # The invoice.total_amount is already calculated and stored during creation
            logger.debug(f"Invoice {invoice.supplier_invoice_number}: "
                        f"Stored total={total_amount}, Paid={payment_total}, Balance={balance_due}")
            
            invoice_list.append(invoice_dict)
        
        logger.info(f"Search completed: {len(invoice_list)} invoices returned using stored values")
        
        return {
            'invoices': invoice_list,
            'total': total,
            'page': page,
            'per_page': per_page,
            'has_more': total > page * per_page,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total_count': total,
                'total_pages': (total + per_page - 1) // per_page
            }
        }
        
    except Exception as e:
        logger.error(f"Error searching supplier invoices: {str(e)}", exc_info=True)
        return {
            'invoices': [],
            'total': 0,
            'page': page,
            'per_page': per_page,
            'has_more': False,
            'pagination': {
                'page': page,
                'per_page': per_page,
                'total_count': 0,
                'total_pages': 0
            }
        }
    
# ===================================================================
# Supplier Payment Management
# ===================================================================

def record_supplier_payment(
    hospital_id: uuid.UUID,
    payment_data: Dict,
    create_gl_entries: bool = True,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Record a payment to a supplier
    
    Args:
        hospital_id: Hospital UUID
        payment_data: Dictionary containing payment information
        create_gl_entries: Whether to create GL entries
        current_user_id: ID of the user recording the payment
        session: Database session (optional)
        
    Returns:
        Dictionary containing created payment details
    """
    logger.info(f"Recording supplier payment for hospital {hospital_id}")
    logger.debug(f"Payment data: {payment_data}")
    logger.debug(f"Create GL entries: {create_gl_entries}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _record_supplier_payment(
            session, hospital_id, payment_data, create_gl_entries, current_user_id
        )
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _record_supplier_payment(
            new_session, hospital_id, payment_data, create_gl_entries, current_user_id
        )
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing supplier payment: {result.get('payment_id')}")
        new_session.commit()
        
        logger.info(f"Successfully recorded supplier payment")
        return result

def _record_supplier_payment(
    session: Session,
    hospital_id: uuid.UUID,
    payment_data: Dict,
    create_gl_entries: bool = True,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to record a payment to a supplier within a session
    """
    try:
        # Validate supplier
        supplier_id = payment_data.get('supplier_id')
        if not supplier_id:
            error_msg = "Supplier ID is required"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Validating supplier {supplier_id}")
        
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Validate invoice if provided
        invoice = None
        invoice_id = payment_data.get('invoice_id')
        
        if invoice_id:
            logger.debug(f"Validating invoice {invoice_id}")
            
            invoice = session.query(SupplierInvoice).filter_by(
                invoice_id=invoice_id,
                hospital_id=hospital_id,
                supplier_id=supplier_id
            ).first()
            
            if not invoice:
                error_msg = f"Supplier invoice with ID {invoice_id} not found for this supplier"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            # Check if invoice is already fully paid
            payment_total = session.query(func.sum(SupplierPayment.amount)).filter_by(
                invoice_id=invoice_id
            ).scalar() or 0
            
            balance_due = invoice.total_amount - Decimal(payment_total)
            logger.debug(f"Invoice balance due: {balance_due}")
            
            if balance_due <= 0:
                error_msg = f"Invoice {invoice.supplier_invoice_number} is already fully paid"
                logger.error(error_msg)
                raise ValueError(error_msg)
                
            # Check if payment amount exceeds balance due
            payment_amount = Decimal(str(payment_data.get('amount', 0)))
            
            if payment_amount > balance_due:
                error_msg = f"Payment amount {payment_amount} exceeds balance due {balance_due}"
                logger.error(error_msg)
                raise ValueError(error_msg)
        
        # Create payment record
        logger.debug("Creating payment record")
        
        payment = SupplierPayment(
            hospital_id=hospital_id,
            supplier_id=supplier_id,
            invoice_id=invoice_id,
            payment_date=payment_data.get('payment_date', datetime.now(timezone.utc)),
            payment_method=payment_data.get('payment_method', 'bank_transfer'),
            currency_code=payment_data.get('currency_code', 'INR'),
            exchange_rate=payment_data.get('exchange_rate', 1.0),
            amount=Decimal(str(payment_data.get('amount', 0))),
            reference_no=payment_data.get('reference_no'),
            status=payment_data.get('status', 'completed'),
            notes=payment_data.get('notes'),
            reconciliation_status='pending'
        )
        
        if current_user_id:
            payment.created_by = current_user_id
            logger.debug(f"Setting created_by to: {current_user_id}")
            
        session.add(payment)
        session.flush()  # To get the payment_id
        
        logger.debug(f"Created payment with ID: {payment.payment_id}")
        
        # Update invoice payment status if applicable
        if invoice:
            logger.debug("Updating invoice payment status")
            
            # Get updated payment total
            new_payment_total = session.query(func.sum(SupplierPayment.amount)).filter_by(
                invoice_id=invoice_id
            ).scalar() or 0
            
            new_balance_due = invoice.total_amount - Decimal(new_payment_total)
            logger.debug(f"New balance due: {new_balance_due}")
            
            # Update payment status
            if new_balance_due <= 0:
                invoice.payment_status = 'paid'
                logger.info(f"Invoice {invoice.supplier_invoice_number} marked as paid")
            else:
                invoice.payment_status = 'partial'
                logger.info(f"Invoice {invoice.supplier_invoice_number} marked as partially paid")
                
            session.flush()
            
        # Create GL entries if requested
        gl_transaction = None
        if create_gl_entries:
            logger.debug("Creating GL entries for payment")
            try:
                gl_result = create_supplier_payment_gl_entries(
                    payment.payment_id,
                    current_user_id,
                    session
                )
                logger.info("GL entries created successfully")
            except Exception as e:
                logger.error(f"Error creating GL entries: {str(e)}")
                if not payment_data.get('skip_gl_error', False):
                    raise
                logger.warning("Continuing despite GL entry error")
            
        # Return the created payment
        result = get_entity_dict(payment)
        result['supplier_name'] = supplier.supplier_name
        
        if invoice:
            result['invoice_number'] = invoice.supplier_invoice_number
            result['invoice_amount'] = float(invoice.total_amount)
            result['payment_status'] = invoice.payment_status
            
        return result
        
    except IntegrityError as e:
        logger.error(f"Database integrity error recording payment: {str(e)}")
        session.rollback()
        raise ValueError(f"Database constraint violation: {str(e)}")
    except Exception as e:
        logger.error(f"Error recording supplier payment: {str(e)}", exc_info=True)
        session.rollback()
        raise

# app/services/supplier_service.py
# Part 4b of 4

def validate_payment_data(payment_data: Dict, hospital_id: uuid.UUID, session: Optional[Session] = None) -> Dict:
    """
    Comprehensive business validation for payment data
    Returns: {'is_valid': bool, 'errors': list, 'warnings': list}
    """
    validation_result = {
        'is_valid': True,
        'errors': [],
        'warnings': []
    }
    
    try:
        if session is None:
            with get_db_session(read_only=True) as new_session:
                return _validate_payment_data_internal(payment_data, hospital_id, new_session, validation_result)
        else:
            return _validate_payment_data_internal(payment_data, hospital_id, session, validation_result)
            
    except Exception as e:
        validation_result['is_valid'] = False
        validation_result['errors'].append(f"Validation error: {str(e)}")
        return validation_result


def _validate_payment_data_internal(payment_data: Dict, hospital_id: uuid.UUID, session: Session, validation_result: Dict) -> Dict:
    """Internal validation logic"""
    
    # === BASIC VALIDATIONS ===
    if not payment_data.get('supplier_id'):
        validation_result['errors'].append("Supplier is required")
    
    if not payment_data.get('branch_id'):
        validation_result['errors'].append("Branch is required")
    
    if not payment_data.get('amount') or float(payment_data.get('amount', 0)) <= 0:
        validation_result['errors'].append("Payment amount must be greater than zero")
    
    if not payment_data.get('payment_method'):
        validation_result['errors'].append("Payment method is required")
    
    # === SUPPLIER VALIDATION ===
    supplier_id = payment_data.get('supplier_id')
    if supplier_id:
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            validation_result['errors'].append("Supplier not found")
        elif supplier.black_listed:
            validation_result['errors'].append(f"Supplier '{supplier.supplier_name}' is blacklisted")
        elif supplier.status != 'active':
            validation_result['errors'].append(f"Supplier '{supplier.supplier_name}' is not active")
    
    # === BRANCH VALIDATION ===
    branch_id = payment_data.get('branch_id')
    if branch_id:
        branch = session.query(Branch).filter_by(
            branch_id=branch_id,
            hospital_id=hospital_id
        ).first()
        
        if not branch:
            validation_result['errors'].append("Branch not found")
        elif not branch.is_active:
            validation_result['errors'].append(f"Branch '{branch.branch_name}' is not active")
    
    # === INVOICE VALIDATION ===
    invoice_id = payment_data.get('invoice_id')
    if invoice_id:
        invoice = session.query(SupplierInvoice).filter_by(
            invoice_id=invoice_id,
            hospital_id=hospital_id,
            supplier_id=supplier_id
        ).first()
        
        if not invoice:
            validation_result['errors'].append("Invoice not found or doesn't belong to selected supplier")
        elif invoice.payment_status == 'cancelled':
            validation_result['errors'].append("Cannot make payment to cancelled invoice")
        else:
            # Check payment amount vs balance due
            payment_amount = float(payment_data.get('amount', 0))
            
            # Get existing payments
            existing_payments = session.query(func.sum(SupplierPayment.amount)).filter_by(
                invoice_id=invoice_id,
                is_reversed=False
            ).scalar() or 0
            
            balance_due = float(invoice.total_amount) - float(existing_payments)
            
            if payment_amount > balance_due:
                validation_result['errors'].append(
                    f"Payment amount ({payment_amount:.2f}) exceeds balance due ({balance_due:.2f})"
                )
    
    # === PAYMENT METHOD SPECIFIC VALIDATIONS ===
    payment_method = payment_data.get('payment_method')
    
    if payment_method == 'mixed':
        # Validate mixed payment amounts
        total_components = (
            float(payment_data.get('cash_amount', 0)) +
            float(payment_data.get('cheque_amount', 0)) +
            float(payment_data.get('bank_transfer_amount', 0)) +
            float(payment_data.get('upi_amount', 0))
        )
        
        total_amount = float(payment_data.get('amount', 0))
        if abs(total_components - total_amount) > 0.01:
            validation_result['errors'].append(
                f"Sum of individual amounts ({total_components:.2f}) must equal total amount ({total_amount:.2f})"
            )
    
    elif payment_method == 'cheque':
        if not payment_data.get('cheque_number'):
            validation_result['errors'].append("Cheque number is required for cheque payments")
        if not payment_data.get('cheque_date'):
            validation_result['errors'].append("Cheque date is required for cheque payments")
        
        # Business rule: Cheque date shouldn't be too old or future
        cheque_date = payment_data.get('cheque_date')
        if cheque_date:
            from datetime import datetime, timedelta
            today = datetime.now().date()
            
            if cheque_date > today:
                validation_result['warnings'].append("Cheque date is in the future")
            elif cheque_date < (today - timedelta(days=180)):
                validation_result['warnings'].append("Cheque is older than 6 months")
    
    elif payment_method == 'bank_transfer':
        if not payment_data.get('bank_reference_number'):
            validation_result['errors'].append("Bank reference number is required for bank transfers")
    
    elif payment_method == 'upi':
        if not payment_data.get('upi_transaction_id'):
            validation_result['errors'].append("UPI transaction ID is required for UPI payments")
    
    # === TDS VALIDATION ===
    if payment_data.get('tds_applicable'):
        tds_rate = float(payment_data.get('tds_rate', 0))
        tds_amount = float(payment_data.get('tds_amount', 0))
        total_amount = float(payment_data.get('amount', 0))
        
        if tds_rate <= 0:
            validation_result['errors'].append("TDS rate must be greater than zero when TDS is applicable")
        
        expected_tds = total_amount * tds_rate / 100
        if abs(tds_amount - expected_tds) > 0.01:
            validation_result['warnings'].append(
                f"TDS amount ({tds_amount:.2f}) doesn't match calculated amount ({expected_tds:.2f})"
            )
    
    # === APPROVAL VALIDATION ===
    if payment_data.get('requires_approval'):
        approval_threshold = float(payment_data.get('approval_threshold', 0))
        payment_amount = float(payment_data.get('amount', 0))
        
        if approval_threshold > 0 and payment_amount < approval_threshold:
            validation_result['warnings'].append(
                f"Payment amount is below approval threshold ({approval_threshold:.2f})"
            )
    
    # === CURRENCY VALIDATION ===
    currency_code = payment_data.get('currency_code', 'INR')
    exchange_rate = float(payment_data.get('exchange_rate', 1.0))
    
    if currency_code != 'INR' and exchange_rate <= 0:
        validation_result['errors'].append("Exchange rate must be greater than zero for foreign currency")
    
    # Set overall validation status
    if validation_result['errors']:
        validation_result['is_valid'] = False
    
    return validation_result


def validate_payment_business_rules(payment_data: Dict, hospital_id: uuid.UUID, session: Optional[Session] = None) -> Dict:
    """
    Additional business rule validations specific to your hospital's policies
    """
    business_rules = {
        'is_valid': True,
        'errors': [],
        'warnings': []
    }
    
    try:
        # === DAILY PAYMENT LIMITS ===
        payment_amount = float(payment_data.get('amount', 0))
        
        # Example: Cash payments over 50,000 require special approval
        if payment_data.get('payment_method') == 'cash' and payment_amount > 50000:
            business_rules['warnings'].append(
                "Cash payments over  Rs.50,000 may require additional documentation"
            )
        
        # === BRANCH SPECIFIC RULES ===
        branch_id = payment_data.get('branch_id')
        if branch_id and session:
            # Example: Check if branch has sufficient cash balance for cash payments
            if payment_data.get('payment_method') == 'cash':
                # This would integrate with your cash management system
                pass
        
        # === SUPPLIER SPECIFIC RULES ===
        supplier_id = payment_data.get('supplier_id')
        if supplier_id and session:
            supplier = session.query(Supplier).filter_by(supplier_id=supplier_id).first()
            if supplier:
                # Example: Check supplier payment terms
                if supplier.payment_terms and 'advance' not in supplier.payment_terms.lower():
                    if not payment_data.get('invoice_id'):
                        business_rules['warnings'].append(
                            "Advance payments may not be allowed for this supplier"
                        )
        
        return business_rules
        
    except Exception as e:
        business_rules['is_valid'] = False
        business_rules['errors'].append(f"Business rule validation error: {str(e)}")
        return business_rules


def enhanced_record_supplier_payment(
    hospital_id: uuid.UUID,
    payment_data: Dict,
    create_gl_entries: bool = True,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Enhanced payment recording with comprehensive validation
    """
    logger.info(f"Recording supplier payment with enhanced validation")
    
    # === STEP 1: VALIDATE PAYMENT DATA ===
    validation_result = validate_payment_data(payment_data, hospital_id, session)
    if not validation_result['is_valid']:
        error_msg = "; ".join(validation_result['errors'])
        raise ValueError(f"Payment validation failed: {error_msg}")
    
    # === STEP 2: BUSINESS RULE VALIDATION ===
    business_validation = validate_payment_business_rules(payment_data, hospital_id, session)
    if not business_validation['is_valid']:
        error_msg = "; ".join(business_validation['errors'])
        raise ValueError(f"Business rule validation failed: {error_msg}")
    
    # === STEP 3: LOG WARNINGS ===
    all_warnings = validation_result.get('warnings', []) + business_validation.get('warnings', [])
    for warning in all_warnings:
        logger.warning(f"Payment warning: {warning}")
    
    # === STEP 4: PROCEED WITH PAYMENT CREATION ===
    return record_supplier_payment(
        hospital_id=hospital_id,
        payment_data=payment_data,
        create_gl_entries=create_gl_entries,
        current_user_id=current_user_id,
        session=session
    )


def get_supplier_payment_history(
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    session: Optional[Session] = None
) -> List[Dict]:
    """
    Get payment history for a supplier
    
    Args:
        supplier_id: Supplier UUID
        hospital_id: Hospital UUID for security
        start_date: Filter by start date
        end_date: Filter by end date
        session: Database session (optional)
        
    Returns:
        List of payment records
    """
    logger.info(f"Getting payment history for supplier {supplier_id}")
    logger.debug(f"Date range: {start_date} to {end_date}")
    
    if session is not None:
        return _get_supplier_payment_history(session, supplier_id, hospital_id, start_date, end_date)
    
    with get_db_session() as new_session:
        return _get_supplier_payment_history(new_session, supplier_id, hospital_id, start_date, end_date)

def _get_supplier_payment_history(
    session: Session,
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None
) -> List[Dict]:
    """
    Internal function to get payment history for a supplier within a session
    """
    try:
        # Validate supplier
        logger.debug(f"Validating supplier {supplier_id}")
        
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug("Building payment history query")
            
        # Query payments
        query = session.query(SupplierPayment).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        )
        
        # Apply date filters
        if start_date:
            logger.debug(f"Filtering by start date: {start_date}")
            query = query.filter(SupplierPayment.payment_date >= start_date)
            
        if end_date:
            logger.debug(f"Filtering by end date: {end_date}")
            query = query.filter(SupplierPayment.payment_date <= end_date)
            
        # Order by date
        query = query.order_by(desc(SupplierPayment.payment_date))
        
        # Execute query
        payments = query.all()
        logger.debug(f"Found {len(payments)} payments")
        
        # Convert to dictionaries with invoice details
        results = []
        
        for payment in payments:
            payment_dict = get_entity_dict(payment)
            
            # Add invoice details if available
            if payment.invoice_id:
                logger.debug(f"Fetching invoice details for payment {payment.payment_id}")
                
                invoice = session.query(SupplierInvoice).filter_by(invoice_id=payment.invoice_id).first()
                if invoice:
                    payment_dict['invoice_number'] = invoice.supplier_invoice_number
                    payment_dict['invoice_date'] = invoice.invoice_date
                    payment_dict['invoice_amount'] = float(invoice.total_amount)
                    
            results.append(payment_dict)
            
        logger.info(f"Retrieved {len(results)} payment records")
            
        return results
        
    except Exception as e:
        logger.error(f"Error getting supplier payment history: {str(e)}", exc_info=True)
        raise

def get_pending_supplier_invoices(
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    session: Optional[Session] = None
) -> List[Dict]:
    """
    Get unpaid or partially paid supplier invoices
    
    Args:
        hospital_id: Hospital UUID
        supplier_id: Filter by supplier ID (optional)
        session: Database session (optional)
        
    Returns:
        List of invoices with payment information
    """
    logger.info(f"Getting pending invoices for hospital {hospital_id}")
    if supplier_id:
        logger.debug(f"Filtering by supplier: {supplier_id}")
    
    if session is not None:
        return _get_pending_supplier_invoices(session, hospital_id, supplier_id)
    
    with get_db_session() as new_session:
        return _get_pending_supplier_invoices(new_session, hospital_id, supplier_id)

def _get_pending_supplier_invoices(
    session: Session,
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None
) -> List[Dict]:
    """
    Internal function to get pending supplier invoices within a session
    """
    try:
        logger.debug("Building pending invoices query")
        
        # Base query for unpaid or partially paid invoices
        query = session.query(SupplierInvoice).filter(
            SupplierInvoice.hospital_id == hospital_id,
            SupplierInvoice.payment_status.in_(['unpaid', 'partial'])
        )
        
        # Apply supplier filter if provided
        if supplier_id:
            logger.debug(f"Filtering by supplier: {supplier_id}")
            query = query.filter(SupplierInvoice.supplier_id == supplier_id)
            
        # Order by date (oldest first)
        query = query.order_by(SupplierInvoice.invoice_date)
        
        # Execute query
        invoices = query.all()
        logger.debug(f"Found {len(invoices)} pending invoices")
        
        # Convert to dictionaries with payment details
        results = []
        
        for invoice in invoices:
            invoice_dict = get_entity_dict(invoice)
            
            # Get supplier information
            logger.debug(f"Getting supplier info for invoice {invoice.invoice_id}")
            
            supplier = session.query(Supplier).filter_by(supplier_id=invoice.supplier_id).first()
            if supplier:
                invoice_dict['supplier_name'] = supplier.supplier_name
                
            # Get payment information
            logger.debug("Calculating payment totals")
            
            payment_total = session.query(func.sum(SupplierPayment.amount)).filter_by(
                invoice_id=invoice.invoice_id
            ).scalar() or 0
            
            payment_total = Decimal(payment_total)
            balance_due = invoice.total_amount - payment_total
            
            invoice_dict['payment_total'] = float(payment_total)
            invoice_dict['balance_due'] = float(balance_due)
            
            # Get payment history
            logger.debug("Getting payment history")
            
            payments = session.query(SupplierPayment).filter_by(
                invoice_id=invoice.invoice_id
            ).order_by(SupplierPayment.payment_date).all()
            
            invoice_dict['payments'] = [get_entity_dict(payment) for payment in payments]
            
            results.append(invoice_dict)
            
        logger.info(f"Retrieved {len(results)} pending invoices")
            
        return results
        
    except Exception as e:
        logger.error(f"Error getting pending supplier invoices: {str(e)}", exc_info=True)
        raise

def delete_supplier(
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Delete a supplier (soft delete by marking as inactive)
    
    Args:
        supplier_id: Supplier UUID
        hospital_id: Hospital UUID for security
        current_user_id: ID of the user deleting the supplier
        session: Database session (optional)
        
    Returns:
        Dictionary containing deletion status
    """
    logger.info(f"Deleting supplier {supplier_id} for hospital {hospital_id}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _delete_supplier(session, supplier_id, hospital_id, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _delete_supplier(new_session, supplier_id, hospital_id, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing supplier deletion: {supplier_id}")
        new_session.commit()
        
        logger.info(f"Successfully deleted supplier")
        return result

def _delete_supplier(
    session: Session,
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to delete a supplier within a session
    """
    try:
        logger.debug(f"Fetching supplier {supplier_id}")
        
        # Get the supplier
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.debug("Checking for active transactions")
        
        # Check if supplier has active purchase orders
        active_pos = session.query(PurchaseOrderHeader).filter_by(
            supplier_id=supplier_id,
            status='approved'
        ).count()
        
        if active_pos > 0:
            error_msg = f"Cannot delete supplier with {active_pos} active purchase orders"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        # Check if supplier has unpaid invoices
        unpaid_invoices = session.query(SupplierInvoice).filter(
            SupplierInvoice.supplier_id == supplier_id,
            SupplierInvoice.payment_status.in_(['unpaid', 'partial'])
        ).count()
        
        if unpaid_invoices > 0:
            error_msg = f"Cannot delete supplier with {unpaid_invoices} unpaid invoices"
            logger.error(error_msg)
            raise ValueError(error_msg)
        
        logger.debug(f"Marking supplier {supplier_id} as inactive")
        
        # Soft delete by marking as inactive
        supplier.status = 'inactive'
        supplier.deleted_flag = True
        
        if current_user_id:
            supplier.updated_by = current_user_id
            supplier.deleted_by = current_user_id
            
        supplier.deleted_at = datetime.now(timezone.utc)
        
        session.flush()
        
        logger.info(f"Supplier {supplier.supplier_name} deleted successfully")
        
        return {
            'supplier_id': str(supplier_id),
            'supplier_name': supplier.supplier_name,
            'status': 'deleted',
            'deleted_at': supplier.deleted_at
        }
        
    except Exception as e:
        logger.error(f"Error deleting supplier: {str(e)}", exc_info=True)
        session.rollback()
        raise

# app/services/supplier_service.py
# Part 4c-1 of 3

# ===================================================================
# Purchase Order Management (Additional functions)
# ===================================================================

def get_po_with_lines(session: Session, po: PurchaseOrderHeader) -> Dict:
    """
    Helper function to get a PO with its line items
    
    Args:
        session: Database session
        po: Purchase order header
        
    Returns:
        Dictionary containing PO details with line items
    """
    try:
        logger.debug(f"Getting PO {po.po_id} with line items")
        
        # Get the line items
        lines = session.query(PurchaseOrderLine).filter_by(po_id=po.po_id).all()
        
        logger.debug(f"Found {len(lines)} line items")
        
        # Get the supplier
        supplier = session.query(Supplier).filter_by(supplier_id=po.supplier_id).first()
        
        if not supplier:
            logger.warning(f"Supplier {po.supplier_id} not found for PO")
        
        # Convert to dictionary
        result = get_entity_dict(po)
        result['line_items'] = [get_entity_dict(line) for line in lines]
        
        if supplier:
            result['supplier_name'] = supplier.supplier_name
            
        return result
        
    except Exception as e:
        logger.error(f"Error getting PO with lines: {str(e)}", exc_info=True)
        raise

def get_purchase_order_by_id(
    po_id: uuid.UUID,
    hospital_id: uuid.UUID,
    session: Optional[Session] = None
) -> Dict:
    """
    Get a purchase order by ID with its line items
    
    Args:
        po_id: Purchase Order UUID
        hospital_id: Hospital UUID for security
        session: Database session (optional)
        
    Returns:
        Dictionary containing PO details with line items
    """
    logger.info(f"Fetching purchase order {po_id} for hospital {hospital_id}")
    
    if session is not None:
        return _get_purchase_order_by_id(session, po_id, hospital_id)
    
    with get_db_session() as new_session:
        return _get_purchase_order_by_id(new_session, po_id, hospital_id)

def _get_purchase_order_by_id(
    session: Session,
    po_id: uuid.UUID,
    hospital_id: uuid.UUID
) -> Dict:
    """
    FIXED: Internal function to get a purchase order by ID within a session
    """
    try:
        logger.debug(f"Querying purchase order {po_id}")
        
        # Get the PO
        po = session.query(PurchaseOrderHeader).filter_by(
            po_id=po_id,
            hospital_id=hospital_id
        ).first()
        
        if not po:
            error_msg = f"Purchase order with ID {po_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.info(f"Found purchase order: {po.po_number}")
            
        # Get PO with lines using stored values only
        po_dict = get_po_with_lines(session, po)
        
        # FIXED: Calculate aggregated totals from stored line_total values with proper debugging
        subtotal = 0  # Sum of taxable amounts
        total_gst = 0  # Sum of GST amounts
        calculated_total = 0  # Sum of line totals
        
        line_items = po_dict.get('line_items', [])
        logger.debug(f"Processing {len(line_items)} line items for total calculation")
        
        for i, line in enumerate(line_items):
            # CRITICAL: Use correct field names and handle None values
            line_taxable = float(line.get('taxable_amount') or 0)
            line_gst = float(line.get('total_gst') or 0)
            line_total_amount = float(line.get('line_total') or 0)
            
            # DEBUGGING: Log each line's contribution
            logger.debug(f"Line {i+1} - {line.get('medicine_name', 'Unknown')}: "
                        f"Units={line.get('units', 0)}, "
                        f"Rate={line.get('pack_purchase_price', 0)}, "
                        f"Taxable={line_taxable}, "
                        f"GST={line_gst}, "
                        f"Total={line_total_amount}")
            
            subtotal += line_taxable
            total_gst += line_gst
            calculated_total += line_total_amount
        
        # Add aggregated totals to PO dictionary
        po_dict['calculated_subtotal'] = subtotal
        po_dict['calculated_total_gst'] = total_gst
        po_dict['calculated_total_with_gst'] = calculated_total
        
        # Get stored header total
        stored_total = float(po_dict.get('total_amount', 0))
        
        # DEBUGGING: Log the calculation results
        logger.info(f"PO {po.po_number} calculation results:")
        logger.info(f"  Calculated Subtotal: {subtotal}")
        logger.info(f"  Calculated GST: {total_gst}")
        logger.info(f"  Calculated Total: {calculated_total}")
        logger.info(f"  Stored Header Total: {stored_total}")
        
        if abs(stored_total - calculated_total) > 0.01:
            logger.warning(f"PO total mismatch - using stored total as authoritative")
        
        # CRITICAL: Verify line items have proper data
        if subtotal == 0 and total_gst > 0:
            logger.error("CALCULATION ERROR: Subtotal is 0 but GST is non-zero - investigating line items")
            for i, line in enumerate(line_items):
                logger.error(f"Line {i+1} raw data: {line}")
        
        logger.debug(f"Final PO totals: Subtotal={subtotal}, GST={total_gst}, Total={stored_total}")
        
        return po_dict
        
    except Exception as e:
        logger.error(f"Error getting purchase order: {str(e)}", exc_info=True)
        raise
    
def search_purchase_orders(
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    po_number: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = 1,
    per_page: int = 20,
    session: Optional[Session] = None
) -> Dict:
    """
    Search purchase orders with filtering and pagination
    
    Args:
        hospital_id: Hospital UUID
        supplier_id: Filter by supplier ID
        po_number: Filter by PO number
        status: Filter by status
        start_date: Filter by start date
        end_date: Filter by end date
        page: Page number for pagination
        per_page: Number of items per page
        session: Database session (optional)
        
    Returns:
        Dictionary containing PO list and pagination info
    """
    logger.info(f"Searching purchase orders for hospital {hospital_id}")
    logger.debug(f"Filters: supplier={supplier_id}, po={po_number}, status={status}")
    logger.debug(f"Date range: {start_date} to {end_date}")
    logger.debug(f"Pagination: page={page}, per_page={per_page}")
    
    if session is not None:
        return _search_purchase_orders(
            session, hospital_id, supplier_id, po_number, 
            status, start_date, end_date, page, per_page
        )
    
    with get_db_session() as new_session:
        return _search_purchase_orders(
            new_session, hospital_id, supplier_id, po_number, 
            status, start_date, end_date, page, per_page
        )

def _search_purchase_orders(
    session: Session,
    hospital_id: uuid.UUID,
    supplier_id: Optional[uuid.UUID] = None,
    po_number: Optional[str] = None,
    status: Optional[str] = None,
    start_date: Optional[datetime] = None,
    end_date: Optional[datetime] = None,
    page: int = 1,
    per_page: int = 20
) -> Dict:
    """
    Internal function to search purchase orders within a session
    """
    try:
        logger.debug("Building purchase order search query")
        
        # Base query
        query = session.query(PurchaseOrderHeader).filter_by(hospital_id=hospital_id)
        
        # Apply filters
        if supplier_id:
            logger.debug(f"Filtering by supplier: {supplier_id}")
            query = query.filter(PurchaseOrderHeader.supplier_id == supplier_id)
            
        if po_number:
            logger.debug(f"Filtering by PO number: {po_number}")
            query = query.filter(PurchaseOrderHeader.po_number.ilike(f"%{po_number}%"))
            
        if status:
            logger.debug(f"Filtering by status: {status}")
            query = query.filter(PurchaseOrderHeader.status == status)
            
        if start_date:
            logger.debug(f"Filtering by start date: {start_date}")
            query = query.filter(PurchaseOrderHeader.po_date >= start_date)
            
        if end_date:
            logger.debug(f"Filtering by end date: {end_date}")
            query = query.filter(PurchaseOrderHeader.po_date <= end_date)
            
        # Count total for pagination
        total_count = query.count()
        logger.debug(f"Total purchase orders found: {total_count}")
        
        # Apply pagination
        query = query.order_by(desc(PurchaseOrderHeader.po_date))
        query = query.offset((page - 1) * per_page).limit(per_page)
        
        # Execute query
        pos = query.all()
        logger.debug(f"Retrieved {len(pos)} purchase orders for current page")
        
        # Convert to dictionaries
        po_list = []
        
        for po in pos:
            po_dict = get_entity_dict(po)
            
            # Get the supplier name
            supplier = session.query(Supplier).filter_by(supplier_id=po.supplier_id).first()
            if supplier:
                po_dict['supplier_name'] = supplier.supplier_name
                
            # Add line count
            line_count = session.query(PurchaseOrderLine).filter_by(po_id=po.po_id).count()
            po_dict['line_count'] = line_count
            
            po_list.append(po_dict)
        
        # Prepare pagination info
        pagination = {
            'page': page,
            'per_page': per_page,
            'total_count': total_count,
            'total_pages': (total_count + per_page - 1) // per_page
        }
        
        logger.info(f"Search completed: {len(po_list)} purchase orders returned")
        
        return {
            'purchase_orders': po_list,
            'pagination': pagination
        }
        
    except Exception as e:
        logger.error(f"Error searching purchase orders: {str(e)}", exc_info=True)
        raise

# app/services/supplier_service.py
# Part 4c-2 of 3

# ===================================================================
# Utility Functions
# ===================================================================

def get_supplier_statistics(
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID,
    session: Optional[Session] = None
) -> Dict:
    """
    Get statistics for a supplier
    
    Args:
        supplier_id: Supplier UUID
        hospital_id: Hospital UUID for security
        session: Database session (optional)
        
    Returns:
        Dictionary containing supplier statistics
    """
    logger.info(f"Getting statistics for supplier {supplier_id}")
    
    if session is not None:
        return _get_supplier_statistics(session, supplier_id, hospital_id)
    
    with get_db_session() as new_session:
        return _get_supplier_statistics(new_session, supplier_id, hospital_id)

def _get_supplier_statistics(
    session: Session,
    supplier_id: uuid.UUID,
    hospital_id: uuid.UUID
) -> Dict:
    """
    Internal function to get supplier statistics within a session
    """
    try:
        # Validate supplier
        logger.debug(f"Validating supplier {supplier_id}")
        
        supplier = session.query(Supplier).filter_by(
            supplier_id=supplier_id,
            hospital_id=hospital_id
        ).first()
        
        if not supplier:
            error_msg = f"Supplier with ID {supplier_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug("Calculating supplier statistics")
        
        # Calculate statistics
        stats = {}
        
        # Total purchase orders
        stats['total_purchase_orders'] = session.query(PurchaseOrderHeader).filter_by(
            supplier_id=supplier_id
        ).count()
        
        # Total invoices
        stats['total_invoices'] = session.query(SupplierInvoice).filter_by(
            supplier_id=supplier_id
        ).count()
        
        # Total payments
        stats['total_payments'] = session.query(SupplierPayment).filter_by(
            supplier_id=supplier_id
        ).count()
        
        # Total business volume
        total_business = session.query(func.sum(SupplierInvoice.total_amount)).filter_by(
            supplier_id=supplier_id
        ).scalar() or 0
        
        stats['total_business_volume'] = float(total_business)
        
        # Outstanding balance
        total_invoiced = session.query(func.sum(SupplierInvoice.total_amount)).filter_by(
            supplier_id=supplier_id
        ).scalar() or 0
        
        total_paid = session.query(func.sum(SupplierPayment.amount)).filter_by(
            supplier_id=supplier_id
        ).scalar() or 0
        
        stats['outstanding_balance'] = float(Decimal(total_invoiced) - Decimal(total_paid))
        
        # Average payment time
        paid_invoices = session.query(SupplierInvoice).filter(
            SupplierInvoice.supplier_id == supplier_id,
            SupplierInvoice.payment_status == 'paid'
        ).all()
        
        if paid_invoices:
            payment_times = []
            for invoice in paid_invoices:
                last_payment = session.query(SupplierPayment).filter_by(
                    invoice_id=invoice.invoice_id
                ).order_by(desc(SupplierPayment.payment_date)).first()
                
                if last_payment:
                    payment_time = (last_payment.payment_date - invoice.invoice_date).days
                    payment_times.append(payment_time)
                    
            stats['average_payment_days'] = sum(payment_times) / len(payment_times) if payment_times else 0
        else:
            stats['average_payment_days'] = 0
            
        # Performance metrics
        stats['supplier_name'] = supplier.supplier_name
        stats['performance_rating'] = supplier.performance_rating
        stats['is_blacklisted'] = supplier.black_listed
        stats['status'] = supplier.status
        
        # Recent activity
        stats['last_po_date'] = None
        last_po = session.query(PurchaseOrderHeader).filter_by(
            supplier_id=supplier_id
        ).order_by(desc(PurchaseOrderHeader.po_date)).first()
        
        if last_po:
            stats['last_po_date'] = last_po.po_date
            
        stats['last_invoice_date'] = None
        last_invoice = session.query(SupplierInvoice).filter_by(
            supplier_id=supplier_id
        ).order_by(desc(SupplierInvoice.invoice_date)).first()
        
        if last_invoice:
            stats['last_invoice_date'] = last_invoice.invoice_date
            
        stats['last_payment_date'] = None
        last_payment = session.query(SupplierPayment).filter_by(
            supplier_id=supplier_id
        ).order_by(desc(SupplierPayment.payment_date)).first()
        
        if last_payment:
            stats['last_payment_date'] = last_payment.payment_date
            
        logger.info(f"Retrieved statistics for supplier {supplier.supplier_name}")
        
        return stats
        
    except Exception as e:
        logger.error(f"Error getting supplier statistics: {str(e)}", exc_info=True)
        raise

def bulk_update_supplier_status(
    supplier_ids: List[uuid.UUID],
    status: str,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Bulk update supplier status
    
    Args:
        supplier_ids: List of supplier UUIDs
        status: New status to set
        hospital_id: Hospital UUID for security
        current_user_id: ID of the user making the update
        session: Database session (optional)
        
    Returns:
        Dictionary containing update results
    """
    logger.info(f"Bulk updating {len(supplier_ids)} suppliers to status: {status}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _bulk_update_supplier_status(
            session, supplier_ids, status, hospital_id, current_user_id
        )
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _bulk_update_supplier_status(
            new_session, supplier_ids, status, hospital_id, current_user_id
        )
        
        # Add explicit commit for this critical operation
        logger.info("Committing bulk supplier status update")
        new_session.commit()
        
        logger.info(f"Successfully updated {result['updated_count']} suppliers")
        return result

def _bulk_update_supplier_status(
    session: Session,
    supplier_ids: List[uuid.UUID],
    status: str,
    hospital_id: uuid.UUID,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to bulk update supplier status within a session
    """
    try:
        # Validate status
        valid_statuses = ['active', 'inactive', 'suspended']
        if status not in valid_statuses:
            error_msg = f"Invalid status '{status}'. Must be one of: {', '.join(valid_statuses)}"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        logger.debug(f"Updating suppliers to status: {status}")
        
        # Update suppliers
        update_time = datetime.now(timezone.utc)
        updated_count = 0
        failed_updates = []
        
        for supplier_id in supplier_ids:
            try:
                logger.debug(f"Updating supplier {supplier_id}")
                
                # Get the supplier
                supplier = session.query(Supplier).filter_by(
                    supplier_id=supplier_id,
                    hospital_id=hospital_id
                ).first()
                
                if not supplier:
                    logger.warning(f"Supplier {supplier_id} not found")
                    failed_updates.append({
                        'supplier_id': str(supplier_id),
                        'error': 'Supplier not found'
                    })
                    continue
                    
                # Update status
                supplier.status = status
                supplier.updated_at = update_time
                
                if current_user_id:
                    supplier.updated_by = current_user_id
                    
                updated_count += 1
                
            except Exception as e:
                logger.error(f"Error updating supplier {supplier_id}: {str(e)}")
                failed_updates.append({
                    'supplier_id': str(supplier_id),
                    'error': str(e)
                })
                
        session.flush()
        
        logger.info(f"Updated {updated_count} suppliers, {len(failed_updates)} failed")
        
        return {
            'updated_count': updated_count,
            'failed_count': len(failed_updates),
            'failed_updates': failed_updates,
            'status': status,
            'update_time': update_time
        }
        
    except Exception as e:
        logger.error(f"Error in bulk supplier status update: {str(e)}", exc_info=True)
        session.rollback()
        raise

def export_supplier_list(
    hospital_id: uuid.UUID,
    filters: Optional[Dict] = None,
    format: str = 'csv',
    session: Optional[Session] = None
) -> Union[str, bytes]:
    """
    Export supplier list to file
    
    Args:
        hospital_id: Hospital UUID
        filters: Optional filters to apply
        format: Export format ('csv' or 'excel')
        session: Database session (optional)
        
    Returns:
        File content as string (CSV) or bytes (Excel)
    """
    logger.info(f"Exporting supplier list for hospital {hospital_id} in {format} format")
    
    if session is not None:
        return _export_supplier_list(session, hospital_id, filters, format)
    
    with get_db_session() as new_session:
        return _export_supplier_list(new_session, hospital_id, filters, format)

def _export_supplier_list(
    session: Session,
    hospital_id: uuid.UUID,
    filters: Optional[Dict] = None,
    format: str = 'csv'
) -> Union[str, bytes]:
    """
    Internal function to export supplier list within a session
    """
    try:
        logger.debug("Building supplier query for export")
        
        # Base query
        query = session.query(Supplier).filter_by(hospital_id=hospital_id)
        
        # Apply filters if provided
        if filters:
            if filters.get('name'):
                query = query.filter(Supplier.supplier_name.ilike(f"%{filters['name']}%"))
                
            if filters.get('category'):
                query = query.filter(Supplier.supplier_category == filters['category'])
                
            if filters.get('status'):
                query = query.filter(Supplier.status == filters['status'])
                
            if filters.get('blacklisted') is not None:
                query = query.filter(Supplier.black_listed == filters['blacklisted'])
                
        # Order by name
        query = query.order_by(Supplier.supplier_name)
        
        # Get all suppliers
        suppliers = query.all()
        logger.debug(f"Exporting {len(suppliers)} suppliers")
        
        # Prepare data for export
        data = []
        headers = [
            'Supplier ID', 'Supplier Name', 'Category', 'Status',
            'Contact Person', 'Phone', 'Email', 'GSTIN',
            'State Code', 'Payment Terms', 'Performance Rating',
            'Black Listed', 'Created Date'
        ]
        
        for supplier in suppliers:
            row = [
                str(supplier.supplier_id),
                supplier.supplier_name,
                supplier.supplier_category or '',
                supplier.status,
                supplier.contact_person_name or '',
                supplier.contact_info.get('phone', '') if supplier.contact_info else '',
                supplier.email or '',
                supplier.gst_registration_number or '',
                supplier.state_code or '',
                supplier.payment_terms or '',
                str(supplier.performance_rating) if supplier.performance_rating else '',
                'Yes' if supplier.black_listed else 'No',
                supplier.created_at.strftime('%Y-%m-%d %H:%M:%S') if supplier.created_at else ''
            ]
            data.append(row)
            
        # Export based on format
        if format == 'csv':
            import csv
            import io
            
            output = io.StringIO()
            writer = csv.writer(output)
            writer.writerow(headers)
            writer.writerows(data)
            
            logger.info("CSV export completed")
            return output.getvalue()
            
        elif format == 'excel':
            import io
            from openpyxl import Workbook
            
            wb = Workbook()
            ws = wb.active
            ws.title = "Suppliers"
            
            # Write headers
            for col, header in enumerate(headers, 1):
                ws.cell(row=1, column=col, value=header)
                
            # Write data
            for row_idx, row_data in enumerate(data, 2):
                for col_idx, value in enumerate(row_data, 1):
                    ws.cell(row=row_idx, column=col_idx, value=value)
                    
            # Save to bytes
            output = io.BytesIO()
            wb.save(output)
            output.seek(0)
            
            logger.info("Excel export completed")
            return output.getvalue()
            
        else:
            error_msg = f"Unsupported export format: {format}"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
    except Exception as e:
        logger.error(f"Error exporting supplier list: {str(e)}", exc_info=True)
        raise

def calculate_purchase_order_line_totals(
    units: float,
    pack_purchase_price: float,
    gst_rate: float,
    supplier_state_code: str,
    hospital_state_code: str,
    units_per_pack: float = 1.0,
    discount_percent: float = 0.0,
    pre_gst_discount: bool = True,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    REMOVE THIS FUNCTION - Replace with calculate_gst_values()
    """
    # This function duplicates the logic now in calculate_gst_values()
    # All callers should use calculate_gst_values() instead
    pass

def calculate_purchase_order_totals(
    line_items: List[Dict],
    currency_code: str = 'INR',
    exchange_rate: float = 1.0,
    additional_charges: Optional[Dict] = None
) -> Dict:
    """
    REMOVE THIS FUNCTION - Replace with simple aggregation
    """
    # This should be replaced with simple sum of line totals from calculate_gst_values()
    pass


def cancel_supplier_invoice(
    invoice_id: uuid.UUID,
    hospital_id: uuid.UUID,
    cancellation_reason: str,
    current_user_id: Optional[str] = None,
    session: Optional[Session] = None
) -> Dict:
    """
    Cancel a supplier invoice
    
    Args:
        invoice_id: Invoice UUID
        hospital_id: Hospital UUID for security
        cancellation_reason: Reason for cancellation
        current_user_id: ID of the user cancelling the invoice
        session: Database session (optional)
        
    Returns:
        Dictionary containing cancelled invoice details
    """
    logger.info(f"Cancelling supplier invoice {invoice_id} for hospital {hospital_id}")
    
    if session is not None:
        logger.debug("Using provided session")
        return _cancel_supplier_invoice(session, invoice_id, hospital_id, cancellation_reason, current_user_id)
    
    logger.debug("Creating new session with explicit commit")
    with get_db_session() as new_session:
        result = _cancel_supplier_invoice(new_session, invoice_id, hospital_id, cancellation_reason, current_user_id)
        
        # Add explicit commit for this critical operation
        logger.info(f"Committing supplier invoice cancellation: {invoice_id}")
        new_session.commit()
        
        logger.info(f"Successfully cancelled supplier invoice")
        return result

def _cancel_supplier_invoice(
    session: Session,
    invoice_id: uuid.UUID,
    hospital_id: uuid.UUID,
    cancellation_reason: str,
    current_user_id: Optional[str] = None
) -> Dict:
    """
    Internal function to cancel a supplier invoice within a session
    """
    try:
        # Get the invoice
        logger.debug(f"Fetching invoice {invoice_id}")
        
        invoice = session.query(SupplierInvoice).filter_by(
            invoice_id=invoice_id,
            hospital_id=hospital_id
        ).first()
        
        if not invoice:
            error_msg = f"Supplier invoice with ID {invoice_id} not found"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # Check if invoice can be cancelled
        if invoice.payment_status == 'cancelled':
            error_msg = f"Invoice {invoice.supplier_invoice_number} is already cancelled"
            logger.error(error_msg)
            raise ValueError(error_msg)
            
        # # Create reverse entries for inventory
        # if create_inventory_reverse_entries:
        #     logger.debug("Creating reverse inventory entries")
        #     try:
        #         # Call inventory service to reverse stock entries
        #         from app.services.inventory_service import reverse_stock_from_supplier_invoice
                
        #         inventory_result = reverse_stock_from_supplier_invoice(
        #             invoice_id=invoice_id,
        #             hospital_id=hospital_id,
        #             current_user_id=current_user_id,
        #             session=session
        #         )
        #         logger.info("Reverse stock entries created successfully")
        #     except Exception as e:
        #         logger.error(f"Error creating reverse stock entries: {str(e)}")
        #         raise
                
        # # Create reverse GL entries
        # if create_gl_reverse_entries:
        #     logger.debug("Creating reverse GL entries")
        #     try:
        #         # Call GL service to reverse accounting entries
        #         from app.services.gl_service import reverse_supplier_invoice_gl_entries
                
        #         gl_result = reverse_supplier_invoice_gl_entries(
        #             invoice_id=invoice_id,
        #             current_user_id=current_user_id,
        #             session=session
        #         )
        #         logger.info("Reverse GL entries created successfully")
        #     except Exception as e:
        #         logger.error(f"Error creating reverse GL entries: {str(e)}")
        #         raise
            
        # Update invoice status and add cancellation info
        invoice.payment_status = 'cancelled'  # Update to cancelled status
        invoice.cancellation_reason = cancellation_reason
        invoice.cancelled_at = datetime.now(timezone.utc)
        invoice.cancelled_by = current_user_id if current_user_id else None
        
        # Update the timestamp
        invoice.updated_at = datetime.now(timezone.utc)
        invoice.updated_by = current_user_id if current_user_id else None
        
        session.flush()
        
        logger.info(f"Supplier invoice {invoice.supplier_invoice_number} cancelled successfully")
        
        # Return the cancelled invoice
        return get_entity_dict(invoice)
        
    except Exception as e:
        logger.error(f"Error cancelling supplier invoice: {str(e)}", exc_info=True)
        session.rollback()
        raise

def calculate_gst_values(
    quantity: float,
    unit_rate: float,
    gst_rate: float,
    discount_percent: float = 0,
    is_free_item: bool = False,
    is_interstate: bool = False,
    conversion_factor: float = 1.0
) -> dict:
    """
    Single source of truth for all GST calculations in the system.
    
    Business Rules Applied:
    1. Rate is vendor's offered price (taxable value)
    2. Discount applied to rate, GST calculated on discounted rate
    3. Free items have zero cost and zero GST
    4. Interstate: IGST, Intrastate: CGST + SGST
    
    Args:
        quantity (float): Quantity in saleable units
        unit_rate (float): Vendor's rate per saleable unit (before discount)
        gst_rate (float): GST rate in percentage (e.g., 18.0 for 18%)
        discount_percent (float, optional): Discount percentage. Defaults to 0.
        is_free_item (bool, optional): Whether this is a free item. Defaults to False.
        is_interstate (bool, optional): Whether transaction is interstate. Defaults to False.
        conversion_factor (float, optional): Sub-units per saleable unit. Defaults to 1.0.
        
    Returns:
        dict: Dictionary containing all calculated values with business-friendly keys
    """
    try:
        # Convert to Decimal for precision in financial calculations
        from decimal import Decimal, ROUND_HALF_UP
        
        quantity_dec = Decimal(str(quantity))
        unit_rate_dec = Decimal(str(unit_rate))
        gst_rate_dec = Decimal(str(gst_rate))
        discount_percent_dec = Decimal(str(discount_percent))
        conversion_factor_dec = Decimal(str(conversion_factor))
        
        # Initialize all values
        base_amount = Decimal('0')
        discount_amount = Decimal('0')
        discounted_rate = Decimal('0')
        taxable_amount = Decimal('0')
        cgst_rate = Decimal('0')
        sgst_rate = Decimal('0')
        igst_rate = Decimal('0')
        cgst_amount = Decimal('0')
        sgst_amount = Decimal('0')
        igst_amount = Decimal('0')
        total_gst_amount = Decimal('0')
        line_total = Decimal('0')
        sub_unit_price = Decimal('0')
        
        # CRITICAL FIX: Handle free items correctly - preserve quantity and conversion factor
        if is_free_item:
            # For free items: financial values are zero, but quantities are preserved
            sub_unit_price = Decimal('0')  # Free items have zero cost per sub-unit
            
            return {
                # Input values (PRESERVED for free items)
                'quantity': float(quantity_dec),                    # KEEP actual quantity
                'unit_rate': 0.0,                                   # Rate is zero for free items
                'discount_percent': 0.0,                           # No discount on free items
                'conversion_factor': float(conversion_factor_dec),  # KEEP conversion factor
                'is_free_item': is_free_item,
                'is_interstate': is_interstate,
                
                # Calculated amounts (all zero for free items)
                'base_amount': 0.0,
                'discount_amount': 0.0,
                'discounted_rate': 0.0,
                'taxable_amount': 0.0,
                
                # GST rates and amounts (all zero for free items)
                'gst_rate': float(gst_rate_dec),
                'cgst_rate': 0.0,
                'sgst_rate': 0.0,
                'igst_rate': 0.0,
                'cgst_amount': 0.0,
                'sgst_amount': 0.0,
                'igst_amount': 0.0,
                'total_gst_amount': 0.0,
                
                # Final amounts (zero cost but preserve inventory data)
                'line_total': 0.0,
                'sub_unit_price': 0.0,
                'total_sub_units': float(quantity_dec * conversion_factor_dec),  # PRESERVE for inventory
                
                # Database field mappings for free items
                'db_mappings': {
                    'units': float(quantity_dec),              # CRITICAL: Preserve actual quantity
                    'pack_purchase_price': 0.0,                # Zero rate for free items
                    'discount_percent': 0.0,
                    'discount_amount': 0.0,
                    'taxable_amount': 0.0,
                    'units_per_pack': float(conversion_factor_dec),  # PRESERVE conversion factor
                    'unit_price': 0.0,
                    'gst_rate': float(gst_rate_dec),
                    'cgst_rate': 0.0,
                    'sgst_rate': 0.0,
                    'igst_rate': 0.0,
                    'cgst': 0.0,
                    'sgst': 0.0,
                    'igst': 0.0,
                    'total_gst': 0.0,
                    'line_total': 0.0,
                    'is_free_item': True
                }
            }
        
        # Step 1: Calculate base amount (quantity × unit rate) - for non-free items
        base_amount = quantity_dec * unit_rate_dec
        
        # Step 2: Apply discount to get discounted rate
        if discount_percent_dec > 0:
            discount_amount = (base_amount * discount_percent_dec / 100).quantize(Decimal('0.01'), ROUND_HALF_UP)
            discounted_rate = (unit_rate_dec * (100 - discount_percent_dec) / 100).quantize(Decimal('0.01'), ROUND_HALF_UP)
        else:
            discount_amount = Decimal('0')
            discounted_rate = unit_rate_dec
        
        # Step 3: Calculate taxable amount (quantity × discounted rate)
        taxable_amount = (quantity_dec * discounted_rate).quantize(Decimal('0.01'), ROUND_HALF_UP)
        
        # Step 4: Determine GST structure based on interstate/intrastate
        if is_interstate:
            # Interstate transaction: Use IGST
            igst_rate = gst_rate_dec
            cgst_rate = Decimal('0')
            sgst_rate = Decimal('0')
        else:
            # Intrastate transaction: Split into CGST + SGST
            cgst_rate = gst_rate_dec / 2
            sgst_rate = gst_rate_dec / 2
            igst_rate = Decimal('0')
        
        # Step 5: Calculate GST amounts on taxable amount
        cgst_amount = (taxable_amount * cgst_rate / 100).quantize(Decimal('0.01'), ROUND_HALF_UP)
        sgst_amount = (taxable_amount * sgst_rate / 100).quantize(Decimal('0.01'), ROUND_HALF_UP)
        igst_amount = (taxable_amount * igst_rate / 100).quantize(Decimal('0.01'), ROUND_HALF_UP)
        total_gst_amount = cgst_amount + sgst_amount + igst_amount
        
        # Step 6: Calculate final line total
        line_total = taxable_amount + total_gst_amount
        
        # Step 7: Calculate sub-unit price for inventory
        if conversion_factor_dec > 0:
            sub_unit_price = discounted_rate / conversion_factor_dec
        else:
            sub_unit_price = discounted_rate
        
        # Return comprehensive calculation results for non-free items
        return {
            # Input values (for reference and validation)
            'quantity': float(quantity_dec),
            'unit_rate': float(unit_rate_dec),
            'discount_percent': float(discount_percent_dec),
            'conversion_factor': float(conversion_factor_dec),
            'is_free_item': is_free_item,
            'is_interstate': is_interstate,
            
            # Base calculations
            'base_amount': float(base_amount),
            'discount_amount': float(discount_amount),
            'discounted_rate': float(discounted_rate),
            'taxable_amount': float(taxable_amount),
            
            # GST rates
            'gst_rate': float(gst_rate_dec),
            'cgst_rate': float(cgst_rate),
            'sgst_rate': float(sgst_rate),
            'igst_rate': float(igst_rate),
            
            # GST amounts
            'cgst_amount': float(cgst_amount),
            'sgst_amount': float(sgst_amount),
            'igst_amount': float(igst_amount),
            'total_gst_amount': float(total_gst_amount),
            
            # Final amounts
            'line_total': float(line_total),
            'sub_unit_price': float(sub_unit_price),
            'total_sub_units': float(quantity_dec * conversion_factor_dec),
            
            # Database field mappings
            'db_mappings': {
                'units': float(quantity_dec),
                'pack_purchase_price': float(unit_rate_dec),
                'discount_percent': float(discount_percent_dec),
                'discount_amount': float(discount_amount),
                'taxable_amount': float(taxable_amount),
                'units_per_pack': float(conversion_factor_dec),
                'unit_price': float(sub_unit_price),
                'gst_rate': float(gst_rate_dec),
                'cgst_rate': float(cgst_rate),
                'sgst_rate': float(sgst_rate),
                'igst_rate': float(igst_rate),
                'cgst': float(cgst_amount),
                'sgst': float(sgst_amount),
                'igst': float(igst_amount),
                'total_gst': float(total_gst_amount),
                'line_total': float(line_total),
                'is_free_item': is_free_item
            }
        }
        
    except (ValueError, TypeError, Exception) as e:
        logger.error(f"Error in calculate_gst_values: {str(e)}")
        # Return safe defaults with preserved quantity for free items
        return {
            'quantity': float(quantity) if quantity else 0.0,  # PRESERVE quantity even in error
            'unit_rate': 0.0 if is_free_item else float(unit_rate) if unit_rate else 0.0,
            'discount_percent': 0.0,
            'conversion_factor': float(conversion_factor) if conversion_factor else 1.0,
            'is_free_item': is_free_item,
            'is_interstate': is_interstate,
            'base_amount': 0.0,
            'discount_amount': 0.0,
            'discounted_rate': 0.0,
            'taxable_amount': 0.0,
            'gst_rate': 0.0,
            'cgst_rate': 0.0,
            'sgst_rate': 0.0,
            'igst_rate': 0.0,
            'cgst_amount': 0.0,
            'sgst_amount': 0.0,
            'igst_amount': 0.0,
            'total_gst_amount': 0.0,
            'line_total': 0.0,
            'sub_unit_price': 0.0,
            'total_sub_units': float(quantity * conversion_factor) if quantity and conversion_factor else 0.0,
            'db_mappings': {
                'units': float(quantity) if quantity else 0.0,  # CRITICAL: Always preserve quantity
                'pack_purchase_price': 0.0 if is_free_item else 0.0,
                'discount_percent': 0.0,
                'discount_amount': 0.0,
                'taxable_amount': 0.0,
                'units_per_pack': float(conversion_factor) if conversion_factor else 1.0,
                'unit_price': 0.0,
                'gst_rate': 0.0,
                'cgst_rate': 0.0,
                'sgst_rate': 0.0,
                'igst_rate': 0.0,
                'cgst': 0.0,
                'sgst': 0.0,
                'igst': 0.0,
                'total_gst': 0.0,
                'line_total': 0.0,
                'is_free_item': is_free_item
            }
        }


def apply_gst_calculations_to_line(line_obj, calculations):
    """
    ENHANCED: Helper function to apply GST calculations with detailed logging
    """
    try:
        db_mappings = calculations.get('db_mappings', {})
        
        # Log what we're applying
        medicine_name = getattr(line_obj, 'medicine_name', 'Unknown')
        logger.debug(f"Applying calculations to line: {medicine_name}")
        
        # Apply calculations using original field names
        line_obj.units = db_mappings.get('units', 0)
        line_obj.pack_purchase_price = db_mappings.get('pack_purchase_price', 0)
        line_obj.discount_percent = db_mappings.get('discount_percent', 0)
        line_obj.discount_amount = db_mappings.get('discount_amount', 0)
        line_obj.taxable_amount = db_mappings.get('taxable_amount', 0)
        line_obj.units_per_pack = db_mappings.get('units_per_pack', 1)
        line_obj.unit_price = db_mappings.get('unit_price', 0)
        line_obj.gst_rate = db_mappings.get('gst_rate', 0)
        line_obj.cgst_rate = db_mappings.get('cgst_rate', 0)
        line_obj.sgst_rate = db_mappings.get('sgst_rate', 0)
        line_obj.igst_rate = db_mappings.get('igst_rate', 0)
        line_obj.cgst = db_mappings.get('cgst', 0)
        line_obj.sgst = db_mappings.get('sgst', 0)
        line_obj.igst = db_mappings.get('igst', 0)
        line_obj.total_gst = db_mappings.get('total_gst', 0)
        line_obj.line_total = db_mappings.get('line_total', 0)
        
        # Apply free item flag if applicable
        if hasattr(line_obj, 'is_free_item'):
            line_obj.is_free_item = db_mappings.get('is_free_item', False)
        
        # CRITICAL: Log what was actually stored
        logger.debug(f"STORED to DB - {medicine_name}: "
                    f"Units={line_obj.units}, "
                    f"Rate={line_obj.pack_purchase_price}, "
                    f"Taxable={line_obj.taxable_amount}, "
                    f"GST={line_obj.total_gst}, "
                    f"Total={line_obj.line_total}")
        
        # CRITICAL: Verify non-zero values were stored
        if float(line_obj.units) > 0 and float(line_obj.taxable_amount) == 0:
            logger.error(f"ERROR: Line {medicine_name} has quantity but zero taxable amount!")
            logger.error(f"Input data: {db_mappings}")
        
    except Exception as e:
        logger.error(f"Error applying GST calculations to line object: {str(e)}")
        raise

def apply_gst_calculations_to_invoice_line(invoice_line_obj, calculations):
    """
    Helper function to apply GST calculations to invoice line objects
    """
    try:
        db_mappings = calculations.get('db_mappings', {})
        
        # Log what we're applying
        medicine_name = getattr(invoice_line_obj, 'medicine_name', 'Unknown')
        logger.debug(f"Applying calculations to invoice line: {medicine_name}")
        
        # Apply calculations using invoice line field names
        invoice_line_obj.units = db_mappings.get('units', 0)
        invoice_line_obj.pack_purchase_price = db_mappings.get('pack_purchase_price', 0)
        invoice_line_obj.discount_percent = db_mappings.get('discount_percent', 0)
        invoice_line_obj.discount_amount = db_mappings.get('discount_amount', 0)
        invoice_line_obj.taxable_amount = db_mappings.get('taxable_amount', 0)
        invoice_line_obj.units_per_pack = db_mappings.get('units_per_pack', 1)
        invoice_line_obj.unit_price = db_mappings.get('unit_price', 0)
        invoice_line_obj.gst_rate = db_mappings.get('gst_rate', 0)
        invoice_line_obj.cgst_rate = db_mappings.get('cgst_rate', 0)
        invoice_line_obj.sgst_rate = db_mappings.get('sgst_rate', 0)
        invoice_line_obj.igst_rate = db_mappings.get('igst_rate', 0)
        invoice_line_obj.cgst = db_mappings.get('cgst', 0)
        invoice_line_obj.sgst = db_mappings.get('sgst', 0)
        invoice_line_obj.igst = db_mappings.get('igst', 0)
        invoice_line_obj.total_gst = db_mappings.get('total_gst', 0)
        invoice_line_obj.line_total = db_mappings.get('line_total', 0)
        
        # Apply free item flag if applicable
        if hasattr(invoice_line_obj, 'is_free_item'):
            invoice_line_obj.is_free_item = db_mappings.get('is_free_item', False)
        
        # CRITICAL: Log what was actually stored
        logger.debug(f"STORED to DB - {medicine_name}: "
                    f"Units={invoice_line_obj.units}, "
                    f"Rate={invoice_line_obj.pack_purchase_price}, "
                    f"Taxable={invoice_line_obj.taxable_amount}, "
                    f"GST={invoice_line_obj.total_gst}, "
                    f"Total={invoice_line_obj.line_total}")
        
    except Exception as e:
        logger.error(f"Error applying GST calculations to invoice line object: {str(e)}")
        raise

def validate_purchase_order_data(po_data: Dict) -> Dict[str, List[str]]:
    """
    Validate complete purchase order data
    
    Args:
        po_data: Dictionary containing PO data
        
    Returns:
        Dictionary with validation results
    """
    validation_result = {
        'is_valid': True,
        'errors': [],
        'line_errors': {}
    }
    
    try:
        # Header validations
        if not po_data.get('supplier_id'):
            validation_result['errors'].append("Supplier is required")
            
        if not po_data.get('line_items') or len(po_data.get('line_items', [])) == 0:
            validation_result['errors'].append("At least one line item is required")
            
        # Line item validations
        line_items = po_data.get('line_items', [])
        has_line_errors = False
        
        for idx, line_data in enumerate(line_items):
            line_errors = validate_po_line_item(line_data)
            
            if line_errors:
                validation_result['line_errors'][f'line_{idx + 1}'] = line_errors
                has_line_errors = True
        
        # Overall validation status
        if validation_result['errors'] or has_line_errors:
            validation_result['is_valid'] = False
            
    except Exception as e:
        validation_result['is_valid'] = False
        validation_result['errors'].append(f"Validation error: {str(e)}")
        
    return validation_result
